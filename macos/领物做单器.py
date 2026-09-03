#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import csv
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
import math
import os
import re
import subprocess
import sys
import threading
import time
import tkinter as tk
import traceback
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from queue import Empty, Queue
from typing import Any, Iterable
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from urllib.parse import quote, urlsplit


def detach_console_for_windowed_python() -> None:
    if os.name != "nt":
        return
    exe_name = Path(sys.executable).name.lower()
    if exe_name not in {"pythonw.exe", "pyw.exe"}:
        return
    try:
        from ctypes import windll

        windll.kernel32.FreeConsole()
    except Exception:
        pass


detach_console_for_windowed_python()

import requests
import playwright._impl._transport as playwright_transport
from playwright.sync_api import Browser, BrowserContext, Page, sync_playwright


def configure_no_console_runtime() -> None:
    node_options = os.environ.get("NODE_OPTIONS", "")
    if "--no-deprecation" not in node_options:
        os.environ["NODE_OPTIONS"] = f"{node_options} --no-deprecation".strip()

    # In pythonw/pyw.exe mode Playwright's Node driver can inherit stderr and
    # open a console for Node warnings. Send it to DEVNULL instead.
    playwright_transport._get_stderr_fileno = lambda: subprocess.DEVNULL


configure_no_console_runtime()


def configure_dpi_awareness() -> None:
    if os.name != "nt":
        return
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            from ctypes import windll as fallback_windll

            fallback_windll.user32.SetProcessDPIAware()
        except Exception:
            pass


configure_dpi_awareness()


NO_WINDOW_FLAGS = getattr(subprocess, "CREATE_NO_WINDOW", 0)


def get_app_runtime_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def get_app_support_dir() -> Path:
    return Path.home() / "Library" / "Application Support" / "领物做单器"


def get_default_auth_state_file() -> Path:
    return get_app_support_dir() / "auth-state-v1.json"


def get_app_settings_file() -> Path:
    return get_app_support_dir() / "settings.json"


def load_app_settings() -> dict[str, Any]:
    try:
        payload = json.loads(get_app_settings_file().read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def save_app_settings(settings: dict[str, Any]) -> None:
    settings_file = get_app_settings_file()
    settings_file.parent.mkdir(parents=True, exist_ok=True)
    temp_file = settings_file.with_suffix(".json.tmp")
    temp_file.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_file.replace(settings_file)


def open_path(path_or_url: Any) -> None:
    target = str(path_or_url)
    if os.name == "nt":
        os.startfile(target)
        return
    subprocess.Popen(["open", target])


DEFAULT_SHORTCUT = ""
DEFAULT_DEBUG_PORT = 18800
DEFAULT_COMPOSITION_XLSX = ""
COMPOSITION_DB_FOLDER = Path.home() / "Documents"
COMPOSITION_DB_SUFFIXES = {".csv", ".xlsx", ".xlsm"}
# 数据库匹配字段：SKU ID 是唯一键，优先于 SKC/SPU。
# 表头比较会自动忽略空格和下划线，因此同时兼容“SKU ID”“SKU_ID”等写法。
COMPOSITION_DB_JOIN_FIELDS = ("SKU_ID", "SKU", "SKC_ID", "SPU_ID")
PRODUCT_NO_FIELD_ALIASES = (
    "货号",
    "款号",
    "商品货号",
    "product_no",
    "productNo",
    "product_sn",
    "productSn",
    "product_number",
    "productNumber",
    "product_code",
    "productCode",
    "goods_no",
    "goodsNo",
    "goods_sn",
    "goodsSn",
    "goods_code",
    "goodsCode",
    "style_no",
    "styleNo",
    "article_no",
    "articleNo",
    "item_no",
    "itemNo",
    "spu_code",
    "spuCode",
)
PRODUCT_NO_POLYESTER_FALLBACK_START = (2026, 7, 1)
PRODUCT_NO_POLYESTER_FALLBACK_START_TEXT = "2026-07-01"
DEFAULT_AUTH_STATE_FILE = str(get_default_auth_state_file())
AUTH_SYNC_PORTS = (18321, 18888)
AUTH_SYNC_LISTEN_SECONDS = 180
AUTH_SOURCE_AUTO = "auto"
AUTH_SOURCE_BROWSER = "browser"
AUTH_SOURCE_FILE = "file"
DEFAULT_AUTH_SOURCE = AUTH_SOURCE_FILE
AUTH_SOURCE_LABELS = {
    "本地同步文件": AUTH_SOURCE_FILE,
}
DEFAULT_ORDER_URLS = [
    "https://user.landwu.com/#/Orderlist",
    "https://usersource.landwu.com/#/Orderlist",
]
STATUS_TEXT_MAP = {
    1: "待编辑",
    2: "待付款",
    3: "已支付",
    4: "排单中",
    5: "生产中",
    6: "已发货",
    7: "已取消",
}
GUI_STATUS_TABS = (1, 2, 3)
WAYBILL_MONITORED_STATUSES = (3, 4, 5, 6)
WAYBILL_FAILED_EXPRESS_STATUS = 3
LOW_BALANCE_ALERT_THRESHOLD = 400.0
SIZE_TARGET_OPTIONS = ("", "通用尺码", "涤纶", "棉", "人棉")
# 一格滚轮对应的滚动像素；Text/Canvas 用像素滚动，避免整张卡片一次跳过去
SCROLL_PIXELS_PER_UNIT = 60
APP_VERSION = "2026.09.03.4"
UPDATE_REPOSITORY = "Frank-jpeg/landwu-order-tool"
UPDATE_BRANCH = "main"
UPDATE_SOURCE_PATH = "macos/领物做单器.py"

LANDWU_AUTH_SYNC_USERSCRIPT = r"""// ==UserScript==
// @name         Landwu-桥接同步登录态-双兼容18888_18321-v7
// @namespace    https://user.landwu.com/
// @version      2026.04.17.2
// @description  无感自动同步当前 Landwu 登录态到本地速卖通专用上传器(18888)和领物TEMU上传器(18321)
// @match        https://user.landwu.com/*
// @grant        GM_xmlhttpRequest
// @connect      127.0.0.1
// @run-at       document-idle
// ==/UserScript==

(function () {
  'use strict';

  const LOCAL_SERVERS = [
    { url: 'http://127.0.0.1:18888', label: '速卖通18888' },
    { url: 'http://127.0.0.1:18321', label: 'TEMU18321' },
  ];
  const BADGE_ID = 'landwu-bridge-badge-v7-dual';
  let lastFingerprint = '';
  let collapseTimer = null;
  let lastStatusKey = '';

  function readAuth() {
    const token = localStorage.getItem('access_token') || '';
    const userInfo = JSON.parse(localStorage.getItem('user_info') || '{}');
    const sessionMatch = document.cookie.match(/(?:^|;\s*)(laravel_session=[^;]+)/);

    return {
      token,
      factoryId: userInfo.factory_id ? String(userInfo.factory_id) : '',
      masterFactoryId: userInfo.factory_id ? `6${userInfo.factory_id}` : '',
      session: sessionMatch ? sessionMatch[1] : '',
      username: userInfo.username || userInfo.nickname || '',
      companyName: userInfo.company_name || '',
      source: 'scriptcat-auto',
    };
  }

  function getFingerprint(auth) {
    return [auth.token, auth.factoryId, auth.masterFactoryId, auth.username].join('|');
  }

  function expandBadge() {
    const badge = document.getElementById(BADGE_ID);
    if (!badge) return;
    badge.dataset.collapsed = '0';
    badge.style.transform = 'translateX(0)';
    badge.style.opacity = '0.96';
  }

  function collapseBadge() {
    const badge = document.getElementById(BADGE_ID);
    if (!badge) return;
    badge.dataset.collapsed = '1';
    badge.style.transform = 'translateX(calc(100% - 16px))';
    badge.style.opacity = '0.76';
  }

  function getBadge() {
    let badge = document.getElementById(BADGE_ID);
    if (badge) return badge;

    badge = document.createElement('div');
    badge.id = BADGE_ID;
    badge.style.cssText = [
      'position:fixed',
      'right:10px',
      'bottom:12px',
      'z-index:999999',
      'width:220px',
      'background:rgba(17,24,39,.88)',
      'color:#fff',
      'padding:6px 10px',
      'border-radius:10px',
      'font-size:12px',
      'line-height:1.4',
      'box-shadow:0 8px 24px rgba(0,0,0,.18)',
      'opacity:.96',
      'transition:transform .22s ease, opacity .22s ease',
      'pointer-events:auto',
      'cursor:default',
      'white-space:normal',
      'word-break:break-all',
      'user-select:none',
    ].join(';');
    badge.addEventListener('mouseenter', () => {
      if (collapseTimer) {
        clearTimeout(collapseTimer);
        collapseTimer = null;
      }
      expandBadge();
    });
    badge.addEventListener('mouseleave', () => {
      if (badge.dataset.canCollapse === '1') {
        collapseBadge();
      }
    });
    document.body.appendChild(badge);
    return badge;
  }

  function setBadge(text, color = '#16a34a', options = {}) {
    const {
      autoCollapse = true,
      collapseDelay = 2200,
      keepExpanded = false,
      statusKey = '',
      silentIfSame = false,
    } = options;

    if (statusKey) {
      if (silentIfSame && lastStatusKey === statusKey) return;
      lastStatusKey = statusKey;
    }

    const badge = getBadge();
    if (collapseTimer) {
      clearTimeout(collapseTimer);
      collapseTimer = null;
    }

    badge.textContent = text;
    badge.style.border = `1px solid ${color}`;
    badge.dataset.canCollapse = keepExpanded ? '0' : '1';
    expandBadge();

    if (autoCollapse && !keepExpanded) {
      collapseTimer = setTimeout(() => {
        collapseBadge();
      }, collapseDelay);
    }
  }

  function postAuth(server, auth) {
    return new Promise((resolve) => {
      GM_xmlhttpRequest({
        method: 'POST',
        url: `${server.url}/api/auth/sync`,
        headers: {
          'Content-Type': 'application/json',
        },
        data: JSON.stringify(auth),
        onload: (response) => {
          try {
            const json = JSON.parse(response.responseText || '{}');
            if (!json.ok) {
              resolve({ ok: false, server, reason: 'sync-failed' });
              return;
            }
            resolve({ ok: true, server, json });
          } catch (error) {
            resolve({ ok: false, server, reason: 'parse-failed' });
          }
        },
        onerror: () => {
          resolve({ ok: false, server, reason: 'connect-failed' });
        },
      });
    });
  }

  async function syncAuth(force = false) {
    const auth = readAuth();
    if (!auth.token || !auth.factoryId) {
      setBadge('桥接登录态：未登录', '#dc2626', {
        autoCollapse: true,
        collapseDelay: 1800,
        statusKey: 'not-logged-in',
        silentIfSame: !force,
      });
      return;
    }

    const fingerprint = getFingerprint(auth);
    if (!force && fingerprint === lastFingerprint) {
      return;
    }

    const results = await Promise.all(LOCAL_SERVERS.map((server) => postAuth(server, auth)));
    const successTargets = results.filter((item) => item.ok);

    if (successTargets.length) {
      lastFingerprint = fingerprint;
      const targetsText = successTargets.map((item) => item.server.label).join(' + ');
      setBadge(`桥接登录态：已同步到 ${targetsText} ${auth.username || auth.companyName || ''}`, '#16a34a', {
        autoCollapse: true,
        collapseDelay: 2200,
        statusKey: `synced:${targetsText}:${auth.username || auth.companyName || ''}`,
      });
      return;
    }

    setBadge('桥接登录态：本地服务未启动', '#dc2626', {
      autoCollapse: true,
      collapseDelay: 1500,
      statusKey: 'service-offline',
      silentIfSame: !force,
    });
  }

  function boot() {
    syncAuth(true);
    setInterval(() => syncAuth(false), 15000);
    window.addEventListener('focus', () => syncAuth(false));
    window.addEventListener('storage', () => syncAuth(true));
  }

  setTimeout(boot, 1200);
})();
"""


def setup_stdio() -> None:
    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        if stream and hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")


setup_stdio()


def normalize_source_text(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def get_source_version(text: str) -> str:
    match = re.search(r'^APP_VERSION\s*=\s*["\']([^"\']+)["\']', text, flags=re.MULTILINE)
    return match.group(1).strip() if match else "未知版本"


def fetch_remote_update_source() -> dict[str, str]:
    source_url = (
        f"https://raw.githubusercontent.com/{UPDATE_REPOSITORY}/{UPDATE_BRANCH}/"
        f"{quote(UPDATE_SOURCE_PATH, safe='/')}"
    )
    try:
        response = requests.get(source_url, timeout=60)
        response.raise_for_status()
        source = response.content.decode("utf-8")
    except requests.RequestException as exc:
        raise RuntimeError(f"无法下载公开更新文件：{exc}") from exc
    except UnicodeDecodeError as exc:
        raise RuntimeError("GitHub 返回的更新文件编码无效。") from exc

    if not source.strip():
        raise RuntimeError("GitHub 更新文件为空，已取消更新。")
    try:
        compile(source, UPDATE_SOURCE_PATH, "exec")
    except SyntaxError as exc:
        raise RuntimeError(f"GitHub 更新文件语法无效：{exc}") from exc

    return {
        "source": source,
        "version": get_source_version(source),
        "sha": str(response.headers.get("ETag") or ""),
    }


def apply_remote_update_source(source: str) -> Path:
    target_path = Path(__file__).resolve()
    compile(source, str(target_path), "exec")
    temp_path = target_path.with_suffix(target_path.suffix + ".download")
    try:
        temp_path.write_text(source, encoding="utf-8")
        os.replace(temp_path, target_path)
    except Exception as exc:  # noqa: BLE001
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise RuntimeError(f"写入更新失败：{exc}") from exc
    return target_path


@dataclass
class BrowserConfig:
    shortcut_path: str
    target_path: str
    arguments: str
    working_directory: str
    user_data_dir: str
    profile_directory: str


@dataclass
class LandwuSession:
    browser_config: BrowserConfig
    browser_pid: int
    origin: str
    host: str
    href: str
    access_token: str
    factory_id: int
    master_factory_id: str
    session_cookie: str
    user_info: dict[str, Any]
    user_agent: str
    page_title: str
    page_url: str
    auth_source: str


@dataclass
class RuntimeHandle:
    config: BrowserConfig
    port: int
    proc: subprocess.Popen[str] | None
    launched: bool


def print_help_text() -> str:
    return "\n".join(
        [
            "Landwu 浏览器做单脚本（Python 单文件版）",
            "",
            "命令行用法：",
            f"  python {Path(__file__).name} auth",
            f"  python {Path(__file__).name} apply-logistics",
            f"  python {Path(__file__).name} apply-logistics --commit",
            f"  python {Path(__file__).name} process-until-review --commit-logistics",
            "",
            "双击运行：",
            "  不带参数会打开可视化界面。",
            "",
            "默认行为：",
            "  1. 复用 C:\\AI用览器.lnk 的 Chrome 登录态",
            f"  2. 默认复用调试端口 {DEFAULT_DEBUG_PORT}",
            "  3. 默认执行完保留浏览器窗口",
            "  4. 支付和改物流都只有显式确认后才会真实提交",
        ]
    )


def json_out(data: Any) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def parse_ids(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [item.strip() for item in re.split(r"[\s,，;；]+", str(raw)) if item.strip()]


def normalize_order_no(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def normalize_sku(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", "", str(value).strip())


def normalize_db_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return ""
    text = text.replace("\t", "").replace("\u3000", " ").strip()
    if re.fullmatch(r"[+-]?\d+(\.0+)?", text):
        return text.split(".")[0]
    if re.fullmatch(r"[+-]?\d+(\.\d+)?[eE][+-]?\d+", text):
        try:
            dec = Decimal(text)
            if dec == dec.to_integral():
                return str(dec.quantize(Decimal("1")))
            return format(dec.normalize(), "f").rstrip("0").rstrip(".")
        except InvalidOperation:
            return text
    return re.sub(r"\s+", "", text)


OPTION_ID_FIELDS = (
    "id",
    "option_id",
    "optionId",
    "size_id",
    "sizeId",
    "value_id",
    "valueId",
    "code_id",
    "codeId",
    "key",
)
OPTION_VALUE_FIELDS = (
    "name_zh",
    "zh_name",
    "cn_name",
    "cnName",
    "display_name",
    "displayName",
    "label",
    "title",
    "name",
    "text",
    "value",
    "size",
    "code",
    "value_code",
    "valueCode",
)


def normalize_option_id(value: Any) -> str:
    return normalize_db_key(value)


def _option_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _option_field_value(option: dict[str, Any], fields: Iterable[str]) -> Any:
    aliases = {_option_key(field) for field in fields}
    for key, value in option.items():
        if _option_key(key) in aliases and value not in (None, ""):
            return value
    return None


def _option_id_from_value(option: Any, fallback: Any = "") -> str:
    if isinstance(option, dict):
        value = _option_field_value(option, OPTION_ID_FIELDS)
        if value not in (None, ""):
            return normalize_option_id(value)
    return normalize_option_id(fallback)


def _option_search_values(option: Any) -> list[str]:
    if not isinstance(option, dict):
        text = str(option or "").strip()
        return [text] if text else []
    id_aliases = {_option_key(field) for field in OPTION_ID_FIELDS}
    values: list[str] = []
    ordered_keys = {_option_key(field): field for field in OPTION_VALUE_FIELDS}
    for field in OPTION_VALUE_FIELDS:
        value = _option_field_value(option, (field,))
        if value in (None, ""):
            continue
        text = str(value).strip()
        if text and text not in values:
            values.append(text)
    for key, value in option.items():
        if _option_key(key) in id_aliases or _option_key(key) not in ordered_keys:
            continue
        if isinstance(value, (dict, list, tuple, set)):
            continue
        text = str(value or "").strip()
        if text and text not in values:
            values.append(text)
    return values


def option_display_name(option: Any, fallback: Any = "") -> str:
    values = _option_search_values(option)
    if values:
        return values[0]
    text = str(fallback or "").strip()
    return text


def iter_named_options(options: Any) -> Iterable[tuple[str, Any]]:
    if isinstance(options, dict):
        for key, option in options.items():
            yield _option_id_from_value(option, key), option
    elif isinstance(options, list):
        for option in options:
            yield _option_id_from_value(option), option


def find_named_option(options: Any, target: str) -> dict[str, Any]:
    wanted = str(target or "").strip()
    if not wanted:
        return {}
    for option_id, option in iter_named_options(options):
        if wanted in _option_search_values(option):
            return {"id": option_id, "name": option_display_name(option, wanted), "raw": option}
    return {}


def find_option_by_id(options: Any, option_id: Any) -> dict[str, Any]:
    wanted = normalize_option_id(option_id)
    if not wanted:
        return {}
    for current_id, option in iter_named_options(options):
        if current_id == wanted:
            return {"id": current_id, "name": option_display_name(option), "raw": option}
    return {}


def format_money(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "-"
    try:
        return f"{float(raw):,.2f}"
    except Exception:
        return raw


def parse_money_amount(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace(",", "").replace("￥", "").replace("¥", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def parse_order_no_set(raw: str | None) -> set[str]:
    return {normalize_order_no(item) for item in parse_ids(raw) if normalize_order_no(item)}


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def normalize_composition_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text)


def infer_size_from_composition(value: Any) -> str:
    text = normalize_composition_text(value)
    if not text:
        return ""
    # 顺序不能改：人棉包含“棉”，必须优先判断。
    if "人棉" in text:
        return "人棉"
    if any(token in text for token in ("涤纶", "聚酯", "聚脂", "polyester")):
        return "涤纶"
    if "棉" in text or "cotton" in text:
        return "棉"
    return ""


def get_product_no_from_sources(*sources: dict[str, Any]) -> str:
    for source in sources:
        if not isinstance(source, dict):
            continue
        for key in PRODUCT_NO_FIELD_ALIASES:
            value = source.get(key)
            if value not in (None, ""):
                return str(value).strip()
    return ""


def parse_product_no_upload_date(value: Any) -> tuple[str, tuple[int, int, int]] | None:
    text = str(value or "")
    for match in re.finditer(r"20\d{6}", text):
        token = match.group(0)
        try:
            parsed = time.strptime(token, "%Y%m%d")
        except ValueError:
            continue
        date_tuple = (parsed.tm_year, parsed.tm_mon, parsed.tm_mday)
        date_text = f"{parsed.tm_year:04d}-{parsed.tm_mon:02d}-{parsed.tm_mday:02d}"
        return date_text, date_tuple
    return None


def infer_polyester_fallback_from_product_no(value: Any) -> dict[str, str] | None:
    product_no = str(value or "").strip()
    parsed = parse_product_no_upload_date(product_no)
    if not product_no or not parsed:
        return None
    date_text, date_tuple = parsed
    if date_tuple < PRODUCT_NO_POLYESTER_FALLBACK_START:
        return None
    return {
        "query": product_no,
        "composition": f"货号 {product_no} 日期 {date_text} >= {PRODUCT_NO_POLYESTER_FALLBACK_START_TEXT}，默认涤纶",
        "target_size": "涤纶",
        "db_field": "货号日期兜底",
        "db_file": "",
        "product_no": product_no,
        "product_no_date": date_text,
    }


def load_composition_xlsx(path_text: str) -> dict[str, Any]:
    try:
        import openpyxl
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("缺少 openpyxl，无法读取 xlsx。") from exc

    path_obj = Path(path_text)
    if not path_obj.exists():
        raise RuntimeError(f"表格不存在：{path_obj}")

    workbook = openpyxl.load_workbook(path_obj, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(item or "").strip() for item in next(rows, [])]
    lowered = [item.lower() for item in headers]

    def find_col(*names: str) -> int:
        for name in names:
            name_lower = name.lower()
            for index, header in enumerate(lowered):
                if name_lower in header:
                    return index
        return -1

    sku_col = find_col("sku")
    composition_col = find_col("成分", "材质")
    order_col = find_col("订单", "订单id", "订单号")
    if sku_col < 0 or composition_col < 0:
        raise RuntimeError(f"表格必须包含 SKU 和 成分 列。当前表头：{', '.join(headers)}")

    mapping: dict[str, dict[str, str]] = {}
    conflicts: list[dict[str, str]] = []
    total_rows = 0
    matched_rows = 0
    unknown_rows = 0
    for row in rows:
        total_rows += 1
        sku = normalize_sku(row[sku_col] if sku_col < len(row) else "")
        composition_value = row[composition_col] if composition_col < len(row) else ""
        composition = str(composition_value or "").strip()
        if not sku or not composition:
            continue
        target_size = infer_size_from_composition(composition)
        order_value = row[order_col] if 0 <= order_col < len(row) else ""
        order_no = str(order_value or "").strip()
        if target_size:
            matched_rows += 1
        else:
            unknown_rows += 1
        item = {"sku": sku, "composition": composition, "target_size": target_size, "order_no": order_no}
        old = mapping.get(sku)
        if old and old.get("target_size") and target_size and old.get("target_size") != target_size:
            conflicts.append(
                {
                    "sku": sku,
                    "old": old.get("composition", ""),
                    "oldTarget": old.get("target_size", ""),
                    "new": composition,
                    "newTarget": target_size,
                }
            )
            continue
        if not old or target_size:
            mapping[sku] = item

    return {
        "path": str(path_obj),
        "mapping": mapping,
        "totalRows": total_rows,
        "matchedRows": matched_rows,
        "unknownRows": unknown_rows,
        "conflicts": conflicts,
    }


def clean_table_header(value: Any) -> str:
    return str(value or "").replace("\n", "").replace("\r", "").strip()


def find_table_column(headers: list[str], aliases: Iterable[str]) -> int:
    cleaned = [clean_table_header(header) for header in headers]
    casefold_map = {header.casefold(): index for index, header in enumerate(cleaned)}
    normalized_map = {
        re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", header.casefold()): index
        for index, header in enumerate(cleaned)
    }
    for alias in aliases:
        alias_text = clean_table_header(alias)
        if alias_text in cleaned:
            return cleaned.index(alias_text)
        found = casefold_map.get(alias_text.casefold())
        if found is not None:
            return found
        normalized_alias = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", alias_text.casefold())
        found = normalized_map.get(normalized_alias)
        if found is not None:
            return found
    return -1


def read_csv_rows(path_obj: Path) -> tuple[list[str], list[list[Any]]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with path_obj.open("r", encoding=encoding, newline="") as handle:
                rows = list(csv.reader(handle))
            if not rows:
                return [], []
            return [clean_table_header(item) for item in rows[0]], rows[1:]
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    raise RuntimeError(f"读取 CSV 失败：{path_obj}，{last_error}")


def read_xlsx_rows(path_obj: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        import openpyxl
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("缺少 openpyxl，无法读取数据库 xlsx。") from exc
    workbook = openpyxl.load_workbook(path_obj, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean_table_header(item) for item in next(iterator, [])]
    return headers, [list(row) for row in iterator]


def read_composition_db_table(path_obj: Path) -> tuple[list[str], list[list[Any]]]:
    suffix = path_obj.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path_obj)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path_obj)
    raise RuntimeError(f"暂不支持数据库文件类型：{path_obj.name}")


def choose_db_composition(row: list[Any], ingredient_col: int, material_col: int) -> str:
    ingredient_value = row[ingredient_col] if 0 <= ingredient_col < len(row) else ""
    ingredient = str(ingredient_value or "").strip()
    if ingredient:
        return ingredient
    material_value = row[material_col] if 0 <= material_col < len(row) else ""
    material = str(material_value or "").strip()
    return material


def load_composition_db_mapping(query_values: Iterable[Any], db_folder: Path = COMPOSITION_DB_FOLDER) -> dict[str, Any]:
    query_keys = {normalize_db_key(value) for value in query_values if normalize_db_key(value)}
    if not query_keys:
        raise RuntimeError("当前订单缺少可查询的 SKU/SKC/商品ID。")
    if not db_folder.exists():
        raise RuntimeError(f"数据库文件夹不存在：{db_folder}")

    db_files = sorted(
        path
        for path in db_folder.iterdir()
        if path.is_file() and path.suffix.lower() in COMPOSITION_DB_SUFFIXES and not path.name.startswith("~$")
    )
    if not db_files:
        raise RuntimeError(f"数据库文件夹没有可用 csv/xlsx 文件：{db_folder}")

    mapping: dict[str, dict[str, str]] = {}
    used_files = 0
    skipped_files: list[str] = []
    scanned_rows = 0
    for db_path in db_files:
        try:
            headers, rows = read_composition_db_table(db_path)
        except Exception:
            skipped_files.append(db_path.name)
            continue
        join_cols = [(field, find_table_column(headers, [field])) for field in COMPOSITION_DB_JOIN_FIELDS]
        join_cols = [(field, index) for field, index in join_cols if index >= 0]
        if not join_cols:
            skipped_files.append(db_path.name)
            continue
        ingredient_col = find_table_column(headers, ["成分", "成份"])
        if ingredient_col < 0 and len(headers) >= 8:
            ingredient_col = 7
        material_col = find_table_column(headers, ["材质"])
        if material_col < 0 and len(headers) >= 7:
            material_col = 6
        if ingredient_col < 0 and material_col < 0:
            skipped_files.append(db_path.name)
            continue
        used_files += 1
        for row in rows:
            scanned_rows += 1
            for field, index in join_cols:
                key = normalize_db_key(row[index] if index < len(row) else "")
                if not key or key not in query_keys or key in mapping:
                    continue
                composition = choose_db_composition(row, ingredient_col, material_col)
                target_size = infer_size_from_composition(composition)
                mapping[key] = {
                    "query": key,
                    "composition": composition or "未匹配",
                    "target_size": target_size,
                    "db_field": field,
                    "db_file": db_path.name,
                }
    return {
        "mapping": mapping,
        "dbFolder": str(db_folder),
        "dbFileCount": used_files,
        "skippedFiles": skipped_files,
        "scannedRows": scanned_rows,
    }


def text_has_payment_risk(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    negative_terms = ("失败", "错误", "不可", "不能", "无法", "不支持", "余额不足", "未通过", "已取消", "不存在", "无效", "风险")
    if any(term in value for term in negative_terms):
        return True
    return "异常" in value and all(term not in value for term in ("无异常", "没有异常", "未发现异常"))


def inspect_payment_preview_expected_ids(preview: Any, expected_ids: Iterable[str]) -> dict[str, Any]:
    expected = {str(item).strip() for item in expected_ids if str(item).strip()}
    found: set[str] = set()
    saw_order_id_field = False
    if not expected:
        return {"found": found, "sawOrderIdField": saw_order_id_field}

    def is_order_id_key(key_norm: str) -> bool:
        return ("order" in key_norm and "id" in key_norm) or key_norm in {"ids", "idlist"}

    def scan_scalar(value: Any) -> None:
        if isinstance(value, bool) or value is None:
            return
        if isinstance(value, float) and value.is_integer():
            text = str(int(value))
        else:
            text = str(value)
        for expected_id in expected:
            if re.search(rf"(?<!\d){re.escape(expected_id)}(?!\d)", text):
                found.add(expected_id)

    def scan_value(value: Any) -> None:
        if isinstance(value, dict):
            for child in value.values():
                scan_value(child)
        elif isinstance(value, (list, tuple, set)):
            for child in value:
                scan_value(child)
        else:
            scan_scalar(value)

    def walk(obj: Any) -> None:
        nonlocal saw_order_id_field
        if isinstance(obj, dict):
            for key, value in obj.items():
                key_norm = normalize_key(key)
                if is_order_id_key(key_norm):
                    saw_order_id_field = True
                if "id" in key_norm or "order" in key_norm:
                    scan_value(value)
                if isinstance(value, (dict, list, tuple, set)):
                    walk(value)
        elif isinstance(obj, (list, tuple, set)):
            for value in obj:
                walk(value)

    walk(preview)
    return {"found": found, "sawOrderIdField": saw_order_id_field}


def validate_payment_preview(preview: Any, expected_ids: Iterable[str]) -> dict[str, Any]:
    issues: list[str] = []
    expected = [str(item) for item in expected_ids if str(item)]
    if not expected:
        issues.append("没有待支付订单")
    if not isinstance(preview, dict) or not preview:
        issues.append("支付预检返回为空")

    good_bool_keys = {
        "ok",
        "success",
        "pass",
        "passed",
        "canpay",
        "allowpay",
        "allowedpay",
        "available",
        "valid",
        "checked",
    }
    bad_true_keys = {
        "error",
        "haserror",
        "fail",
        "failed",
        "invalid",
        "disabled",
        "risk",
        "hasrisk",
        "rejected",
        "abnormal",
        "hasabnormal",
        "unpayable",
    }
    bad_list_tokens = ("error", "fail", "invalid", "abnormal", "unpay", "reject", "disable", "risk")

    def add_issue(message: str) -> None:
        if message not in issues and len(issues) < 20:
            issues.append(message)

    def walk(obj: Any, path: str = "preview") -> None:
        if isinstance(obj, dict):
            for key, value in obj.items():
                key_text = str(key)
                key_norm = normalize_key(key)
                child_path = f"{path}.{key_text}"
                if isinstance(value, bool):
                    if key_norm in good_bool_keys and not value:
                        add_issue(f"{child_path}=false")
                    if key_norm in bad_true_keys and value:
                        add_issue(f"{child_path}=true")
                elif isinstance(value, (int, float)):
                    if key_norm in good_bool_keys and value == 0:
                        add_issue(f"{child_path}=0")
                    if key_norm in bad_true_keys and value != 0:
                        add_issue(f"{child_path}={value}")
                elif isinstance(value, str):
                    if text_has_payment_risk(value):
                        add_issue(f"{child_path}: {value[:80]}")
                elif isinstance(value, (list, tuple)) and value and any(token in key_norm for token in bad_list_tokens):
                    add_issue(f"{child_path} 非空")

                if isinstance(value, (dict, list, tuple)):
                    walk(value, child_path)
        elif isinstance(obj, (list, tuple)):
            for index, value in enumerate(obj):
                walk(value, f"{path}[{index}]")
        elif isinstance(obj, str) and text_has_payment_risk(obj):
            add_issue(f"{path}: {obj[:80]}")

    walk(preview)
    preview_id_info = inspect_payment_preview_expected_ids(preview, expected)
    preview_ids = preview_id_info["found"]
    missing_ids = sorted(set(expected) - preview_ids)
    if (preview_ids or preview_id_info["sawOrderIdField"]) and missing_ids:
        add_issue("支付预检未覆盖订单ID：" + ",".join(missing_ids[:5]))
    return {"ok": not issues, "issues": issues, "checkedIds": expected, "previewIds": sorted(preview_ids)}


def require_payment_preview_ok(preview: Any, expected_ids: Iterable[str]) -> dict[str, Any]:
    validation = validate_payment_preview(preview, expected_ids)
    if not validation["ok"]:
        raise RuntimeError("支付预检未通过，已阻止支付：" + "；".join(validation["issues"][:5]))
    return validation


def summarize_download_failure(download_data: dict[str, Any]) -> str:
    if download_data.get("success"):
        return ""
    message = str(download_data.get("message") or "图片下载不完整")
    manifest = str(download_data.get("manifest") or "")
    if manifest:
        message += f"\n详情：{manifest}"
    return message


def require_download_success(download_data: dict[str, Any]) -> None:
    message = summarize_download_failure(download_data)
    if message:
        raise RuntimeError(message)


def sanitize_name(text: str) -> str:
    value = str(text or "").strip()
    value = re.sub(r'[\\/:*?"<>|]', "_", value)
    value = re.sub(r"[ ._]+$", "", value)
    return value or "unknown"


def guess_ext(url_text: str) -> str:
    lowered = urlsplit(url_text).path.lower()
    for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        if lowered.endswith(ext):
            return ext
    return ".jpg"


def unique_path(path_obj: Path) -> Path:
    if not path_obj.exists():
        return path_obj
    index = 2
    while True:
        candidate = path_obj.with_name(f"{path_obj.stem}__{index}{path_obj.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def format_local_stamp() -> str:
    return time.strftime("%Y-%m-%d-%H%M")


def get_default_output_dir() -> Path:
    return Path.home() / "Desktop" / f"{time.strftime('%Y-%m-%d')}-landwu-orders"


def parse_chrome_argument(arguments_text: str, name: str) -> str:
    pattern = re.compile(rf'--{re.escape(name)}=(?:"([^"]+)"|(\S+))', re.IGNORECASE)
    match = pattern.search(arguments_text or "")
    if not match:
        return ""
    return (match.group(1) or match.group(2) or "").strip()


def resolve_shortcut(shortcut_path: str) -> BrowserConfig:
    escaped = shortcut_path.replace("'", "''")
    ps_script = "\n".join(
        [
            "$ws = New-Object -ComObject WScript.Shell",
            f"$lnk = $ws.CreateShortcut('{escaped}')",
            "$obj = [pscustomobject]@{",
            "  TargetPath = $lnk.TargetPath",
            "  Arguments = $lnk.Arguments",
            "  WorkingDirectory = $lnk.WorkingDirectory",
            "}",
            "$obj | ConvertTo-Json -Depth 3",
        ]
    )
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", ps_script],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=NO_WINDOW_FLAGS,
        check=True,
    )
    payload = json.loads(result.stdout)
    arguments = payload.get("Arguments") or ""
    return BrowserConfig(
        shortcut_path=shortcut_path,
        target_path=payload.get("TargetPath") or "",
        arguments=arguments,
        working_directory=payload.get("WorkingDirectory") or "",
        user_data_dir=parse_chrome_argument(arguments, "user-data-dir"),
        profile_directory=parse_chrome_argument(arguments, "profile-directory") or "Default",
    )


def resolve_browser_config(args: argparse.Namespace) -> BrowserConfig:
    if args.browser_path and args.user_data_dir:
        return BrowserConfig(
            shortcut_path=args.shortcut or "",
            target_path=args.browser_path,
            arguments="",
            working_directory=str(Path(args.browser_path).parent),
            user_data_dir=args.user_data_dir,
            profile_directory=args.profile_directory or "Default",
        )

    config = resolve_shortcut(args.shortcut or DEFAULT_SHORTCUT)
    if not config.target_path:
        raise RuntimeError(f"快捷方式无效：{args.shortcut}")
    if not config.user_data_dir:
        raise RuntimeError(f"快捷方式里没有 --user-data-dir：{args.shortcut}")
    if args.browser_path:
        config.target_path = args.browser_path
    if args.user_data_dir:
        config.user_data_dir = args.user_data_dir
    if args.profile_directory:
        config.profile_directory = args.profile_directory
    return config


def wait_for_debug_port(port: int, proc: subprocess.Popen[str], timeout: float = 20.0) -> None:
    deadline = time.time() + timeout
    last_error = ""
    while time.time() < deadline:
        if proc.poll() is not None:
            raise RuntimeError("AI浏览器可能已在使用这个登录态，请先关闭它的所有窗口后重试")
        try:
            response = requests.get(f"http://127.0.0.1:{port}/json/version", timeout=1)
            if response.ok:
                return
            last_error = f"HTTP {response.status_code}"
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
        time.sleep(0.5)
    raise RuntimeError(f"浏览器调试端口未就绪：{last_error}")


def is_debug_port_ready(port: int) -> bool:
    try:
        response = requests.get(f"http://127.0.0.1:{port}/json/version", timeout=1)
        return bool(response.ok)
    except Exception:  # noqa: BLE001
        return False


def launch_chrome(config: BrowserConfig, order_url: str, port: int) -> subprocess.Popen[str]:
    args = [
        config.target_path,
        f"--remote-debugging-port={port}",
        f'--user-data-dir={config.user_data_dir}',
        f"--profile-directory={config.profile_directory}",
        "--no-first-run",
        "--no-default-browser-check",
        "--new-window",
        order_url,
    ]
    creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0) | NO_WINDOW_FLAGS
    proc = subprocess.Popen(
        args,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        stdin=subprocess.DEVNULL,
        creationflags=creationflags,
        text=True,
    )
    wait_for_debug_port(port, proc)
    return proc


def terminate_process_tree(pid: int) -> None:
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/T", "/F"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=NO_WINDOW_FLAGS,
        check=False,
    )


def ensure_browser_runtime(args: argparse.Namespace) -> RuntimeHandle:
    config = resolve_browser_config(args)
    port = int(getattr(args, "debug_port", DEFAULT_DEBUG_PORT))
    if is_debug_port_ready(port):
        return RuntimeHandle(config=config, port=port, proc=None, launched=False)
    proc = launch_chrome(config, DEFAULT_ORDER_URLS[0], port)
    return RuntimeHandle(config=config, port=port, proc=proc, launched=True)


def connect_browser(port: int) -> tuple[Any, Browser, BrowserContext]:
    playwright = sync_playwright().start()
    browser = playwright.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
    contexts = browser.contexts
    if not contexts:
        playwright.stop()
        raise RuntimeError("未获取到浏览器上下文")
    return playwright, browser, contexts[0]


def pick_order_page(context: BrowserContext) -> Page:
    for page in context.pages:
        url = page.url
        if "landwu.com" in url and "#/Orderlist" in url:
            return page

    page = context.pages[0] if context.pages else context.new_page()
    last_error: Exception | None = None
    for url in DEFAULT_ORDER_URLS:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(1500)
            return page
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    raise RuntimeError(str(last_error) if last_error else "无法打开 Landwu 订单页")


def extract_session_payload(page: Page) -> dict[str, Any]:
    for _ in range(10):
        payload = page.evaluate(
            """() => {
                const rawUser = localStorage.getItem('user_info') || '{}';
                let parsedUser = {};
                try {
                    parsedUser = JSON.parse(rawUser);
                } catch (error) {}
                return {
                    origin: location.origin,
                    host: location.host,
                    href: location.href,
                    accessToken: localStorage.getItem('access_token') || '',
                    userInfo: parsedUser,
                    userAgent: navigator.userAgent,
                    sessionCookie: (document.cookie.match(/(?:^|;\\s*)(laravel_session=[^;]+)/) || [])[1] || '',
                };
            }"""
        )
        if payload.get("accessToken") and int(payload.get("userInfo", {}).get("factory_id") or 0) > 0:
            return payload
        page.wait_for_timeout(1000)
    raise RuntimeError("页面状态不明确，已暂停")


def load_auth_state_payload(auth_file: str) -> dict[str, Any]:
    path_obj = Path(auth_file or DEFAULT_AUTH_STATE_FILE)
    if not path_obj.exists():
        raise RuntimeError(f"登录态文件不存在：{path_obj}")
    try:
        auth = json.loads(path_obj.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"登录态文件读取失败：{path_obj}") from exc

    token = str(auth.get("token") or auth.get("access_token") or "").strip()
    factory_id = str(auth.get("factoryId") or auth.get("factory_id") or "").strip()
    master_factory_id = str(auth.get("masterFactoryId") or auth.get("master_factory_id") or "").strip()
    if not master_factory_id and factory_id:
        master_factory_id = f"6{factory_id}"
    if not token:
        raise RuntimeError("登录态文件缺少 token，请先打开 Landwu 页面让同步脚本刷新一次。")
    if not factory_id:
        raise RuntimeError("登录态文件缺少 factoryId，请先打开 Landwu 页面让同步脚本刷新一次。")

    user_info = {
        "username": auth.get("username") or "",
        "nickname": auth.get("username") or "",
        "company_name": auth.get("companyName") or "",
        "factory_id": factory_id,
        "synced_at": auth.get("syncedAt") or "",
    }
    return {
        "origin": "https://user.landwu.com",
        "host": "user.landwu.com",
        "href": "auth-state-v1.json",
        "accessToken": token,
        "factoryId": factory_id,
        "masterFactoryId": master_factory_id,
        "sessionCookie": str(auth.get("session") or "").strip(),
        "userInfo": user_info,
        "userAgent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
        ),
        "pageTitle": "本地同步登录态",
        "pageUrl": str(path_obj),
        "authSource": AUTH_SOURCE_FILE,
    }


def normalize_synced_auth_payload(payload: dict[str, Any]) -> dict[str, Any]:
    token = str(payload.get("token") or payload.get("access_token") or payload.get("accessToken") or "").strip()
    factory_id = str(payload.get("factoryId") or payload.get("factory_id") or "").strip()
    master_factory_id = str(payload.get("masterFactoryId") or payload.get("master_factory_id") or "").strip()
    if not master_factory_id and factory_id:
        master_factory_id = f"6{factory_id}"
    if not token:
        raise ValueError("同步数据缺少 token")
    if not factory_id:
        raise ValueError("同步数据缺少 factoryId")
    return {
        "token": token,
        "factoryId": factory_id,
        "masterFactoryId": master_factory_id,
        "session": str(payload.get("session") or payload.get("sessionCookie") or "").strip(),
        "username": str(payload.get("username") or payload.get("nickname") or "").strip(),
        "companyName": str(payload.get("companyName") or payload.get("company_name") or "").strip(),
        "source": str(payload.get("source") or "landwu-helper-sync").strip(),
        "syncedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
    }


class AuthSyncReceiver:
    def __init__(self, auth_file_getter, on_synced=None, ports: Iterable[int] = AUTH_SYNC_PORTS):
        self.auth_file_getter = auth_file_getter
        self.on_synced = on_synced
        self.ports = list(ports)
        self.servers: list[ThreadingHTTPServer] = []
        self.threads: list[threading.Thread] = []

    def start(self) -> list[int]:
        if self.servers:
            return [int(server.server_port) for server in self.servers]
        started_ports: list[int] = []
        owner = self

        class Handler(BaseHTTPRequestHandler):
            def log_message(self, _format: str, *args: Any) -> None:
                return

            def send_json(self, status: int, data: dict[str, Any]) -> None:
                body = json.dumps(data, ensure_ascii=False).encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Access-Control-Allow-Headers", "Content-Type")
                self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

            def do_OPTIONS(self) -> None:
                self.send_json(200, {"ok": True})

            def do_GET(self) -> None:
                if self.path.split("?", 1)[0] == "/health":
                    self.send_json(200, {"ok": True, "name": "landwu-auth-sync"})
                else:
                    self.send_json(404, {"ok": False, "message": "not found"})

            def do_POST(self) -> None:
                if self.path.split("?", 1)[0] != "/api/auth/sync":
                    self.send_json(404, {"ok": False, "message": "not found"})
                    return
                try:
                    length = int(self.headers.get("Content-Length") or "0")
                    if length <= 0 or length > 1024 * 1024:
                        raise ValueError("同步数据长度异常")
                    raw = self.rfile.read(length).decode("utf-8")
                    payload = json.loads(raw or "{}")
                    if not isinstance(payload, dict):
                        raise ValueError("同步数据格式错误")
                    auth = normalize_synced_auth_payload(payload)
                    auth_file = Path(owner.auth_file_getter() or DEFAULT_AUTH_STATE_FILE)
                    auth_file.parent.mkdir(parents=True, exist_ok=True)
                    auth_file.write_text(json.dumps(auth, ensure_ascii=False, indent=2), encoding="utf-8")
                    if owner.on_synced:
                        owner.on_synced(auth, auth_file, self.server.server_port)
                    self.send_json(200, {"ok": True, "data": {"username": auth.get("username"), "factoryId": auth.get("factoryId")}})
                except Exception as exc:  # noqa: BLE001
                    self.send_json(400, {"ok": False, "message": str(exc)})

        for port in self.ports:
            try:
                server = ThreadingHTTPServer(("127.0.0.1", int(port)), Handler)
                server.daemon_threads = True
                thread = threading.Thread(target=server.serve_forever, daemon=True, name=f"landwu-auth-sync-{port}")
                thread.start()
                self.servers.append(server)
                self.threads.append(thread)
                started_ports.append(int(port))
            except OSError:
                continue
        return started_ports

    def stop(self) -> None:
        for server in self.servers:
            try:
                server.shutdown()
                server.server_close()
            except Exception:
                pass
        self.servers.clear()
        self.threads.clear()


def looks_like_auth_error(exc: Exception) -> bool:
    text = str(exc).lower()
    tokens = (
        "登录态",
        "登录",
        "未登录",
        "重新登录",
        "鉴权",
        "认证",
        "授权",
        "unauthorized",
        "forbidden",
        "token",
        "401",
        "403",
        "factoryid",
        "factory id",
    )
    return any(token in text for token in tokens)


class LandwuClient:
    def __init__(self, session: LandwuSession):
        self.session = session
        self.http = requests.Session()

    def headers(self) -> dict[str, str]:
        token = self.session.access_token
        headers = {
            "Authorization": f"Bearer {token}",
            "X-CSRF-TOKEN": f"Bearer {token}",
            "m-master-factory-id": f"factory:{self.session.master_factory_id or f'6{self.session.factory_id}'}",
            "Content-Type": "application/json;charset=UTF-8",
            "lange": "zh-CN",
            "User-Agent": self.session.user_agent,
            "Referer": f"{self.session.origin}/#/Orderlist",
        }
        if self.session.session_cookie:
            headers["Cookie"] = self.session.session_cookie
        return headers

    def post(self, api_path: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
        body = dict(payload or {})
        body.setdefault("api_token", self.session.access_token)
        body.setdefault("lange", "zh-CN")
        response = self.http.post(
            f"{self.session.origin}/api{api_path}",
            headers=self.headers(),
            json=body,
            timeout=60,
        )
        try:
            data = response.json()
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"{api_path} 返回非 JSON：{response.text[:200]}") from exc
        if response.status_code >= 400:
            raise RuntimeError(f"{api_path} HTTP {response.status_code}：{data.get('msg') or response.text[:200]}")
        if data.get("code") is not None and int(data.get("code")) != 1:
            raise RuntimeError(f"{api_path} 失败：{data.get('msg') or data.get('message') or response.text[:200]}")
        return data

    def get(self, api_path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        query = dict(params or {})
        query.setdefault("api_token", self.session.access_token)
        query.setdefault("lange", "zh")
        response = self.http.get(
            f"{self.session.origin}/api{api_path}",
            headers=self.headers(),
            params=query,
            timeout=60,
        )
        try:
            data = response.json()
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"{api_path} 返回非 JSON：{response.text[:200]}") from exc
        if response.status_code >= 400:
            raise RuntimeError(f"{api_path} HTTP {response.status_code}：{data.get('msg') or response.text[:200]}")
        if data.get("code") is not None and int(data.get("code")) != 1:
            raise RuntimeError(f"{api_path} 失败：{data.get('msg') or data.get('message') or response.text[:200]}")
        return data

    def get_platform_tags(self) -> list[dict[str, Any]]:
        result = self.post("/plat/getOrderPlatformTags")
        return result.get("data") or []

    def get_jit_tag(self) -> dict[str, Any]:
        tags = self.get_platform_tags()
        for item in tags:
            if str(item.get("name") or "").strip().upper() == "JIT":
                return item
        raise RuntimeError("未找到 JIT 平台标签")

    def get_user_profile(self) -> dict[str, Any]:
        result = self.post("/user/getProfile")
        return (result.get("data") or {}).get("userinfo") or {}

    def get_company_list(self, plat_id: int | str) -> list[dict[str, Any]]:
        result = self.post("/logistic/getCompanyList", {"plat_id": str(plat_id)})
        data = result.get("data") or []
        if isinstance(data, dict):
            data = data.get("data") or data.get("list") or []
        normalized = []
        for item in data:
            copied = dict(item)
            copied["id"] = copied.get("id") or copied.get("logistic_id")
            normalized.append(copied)
        return normalized

    def find_temu_company(self, plat_id: int | str, keyword: str = "temu") -> dict[str, Any]:
        lower_keyword = str(keyword or "temu").lower()
        companies = self.get_company_list(plat_id)
        for item in companies:
            name = str(item.get("name") or "").lower()
            code = str(item.get("code") or "").lower()
            if lower_keyword in name or lower_keyword in code:
                return item
        raise RuntimeError(f"plat_id={plat_id} 下未找到包含 {keyword} 的物流")

    def get_order_company_preview(
        self,
        plat_id: int | str,
        order_ids: Iterable[str],
        logistic_id: int | str,
    ) -> dict[str, Any]:
        result = self.post(
            "/logistic/getOrderCompany",
            {
                "plat_id": str(plat_id),
                "order_id": ",".join(str(item) for item in order_ids),
                "logistic_id": str(logistic_id),
            },
        )
        return result.get("data") or {}

    def get_order_edit_detail(self, order_detail_id: int | str) -> dict[str, Any]:
        result = self.get("/order/getEditDetail", {"ids": str(order_detail_id), "type": 1})
        return ((result.get("data") or {}).get("data") or {})

    def get_order_product_info(self, product_id: int | str) -> dict[str, Any]:
        result = self.get("/order/getProductInfo", {"productId": str(product_id)})
        return ((result.get("data") or {}).get("data") or {})

    def get_order_size_state(
        self,
        order_detail_id: int | str,
        *,
        list_detail: dict[str, Any] | None = None,
        product_id: int | str | None = None,
    ) -> dict[str, Any]:
        """读取领物的尺码选项，用选项 ID 作为当前值的唯一判断依据。"""
        edit_detail = self.get_order_edit_detail(order_detail_id)
        current = edit_detail.get("data") or {}
        size_options = edit_detail.get("size") or edit_detail.get("sizes") or {}
        resolved_product_id = (
            edit_detail.get("product_id")
            or current.get("product_id")
            or product_id
            or (list_detail or {}).get("product_id")
            or (list_detail or {}).get("productId")
        )
        product_info: dict[str, Any] = {}
        if resolved_product_id and not size_options:
            product_info = self.get_order_product_info(resolved_product_id)
            size_options = product_info.get("size") or product_info.get("sizes") or {}

        current_size_id = normalize_option_id(
            current.get("size_id")
            or current.get("sizeId")
            or (list_detail or {}).get("size_id")
            or (list_detail or {}).get("sizeId")
        )
        current_option = find_option_by_id(size_options, current_size_id)
        if not current_option:
            current_size = current.get("size") or (list_detail or {}).get("size") or ""
            current_option = find_named_option(size_options, current_size)
            current_size_id = current_option.get("id") or current_size_id
        current_name = current_option.get("name") or str(current.get("size") or (list_detail or {}).get("size") or "").strip()
        return {
            "orderDetailId": str(order_detail_id),
            "productId": resolved_product_id,
            "currentSizeId": normalize_option_id(current_size_id),
            "currentSizeName": str(current_name or "").strip(),
            "sizeOptions": size_options,
            "editDetail": edit_detail,
            "productInfo": product_info,
        }

    def inspect_payment_size_states(
        self,
        rows: list[dict[str, Any]],
        selected_order_ids: Iterable[str],
    ) -> dict[str, Any]:
        selected = {str(item) for item in selected_order_ids if str(item)}
        generic_by_order: dict[str, dict[str, Any]] = {}
        unresolved_by_order: dict[str, dict[str, Any]] = {}
        for row in rows:
            order_id = str(row.get("order_id") or "")
            if not order_id or order_id not in selected:
                continue
            for detail in row.get("detail") or []:
                if not isinstance(detail, dict):
                    continue
                detail_id = detail.get("id") or detail.get("order_detail_id") or detail.get("item_id")
                if not detail_id:
                    continue
                sku = normalize_sku(
                    detail.get("sku")
                    or detail.get("sku_id")
                    or detail.get("skuId")
                    or detail.get("productSku")
                    or detail.get("product_sku")
                )
                try:
                    state = self.get_order_size_state(
                        detail_id,
                        list_detail=detail,
                        product_id=detail.get("product_id") or detail.get("productId"),
                    )
                except Exception as exc:  # noqa: BLE001
                    unresolved_by_order.setdefault(
                        order_id,
                        {"order_id": order_id, "order_no": str(row.get("order_no") or ""), "skus": []},
                    )["skus"].append(sku or "未识别 SKU")
                    continue
                options = state.get("sizeOptions") or {}
                generic_option = find_named_option(options, "通用尺码")
                current_id = normalize_option_id(state.get("currentSizeId"))
                generic_id = normalize_option_id(generic_option.get("id"))
                current_name = str(state.get("currentSizeName") or "").strip()
                is_generic = bool(
                    (current_id and generic_id and current_id == generic_id)
                    or current_name == "通用尺码"
                )
                if is_generic:
                    generic_by_order.setdefault(
                        order_id,
                        {"order_id": order_id, "order_no": str(row.get("order_no") or ""), "skus": []},
                    )["skus"].append(sku or "未识别 SKU")
                elif not current_id or not options:
                    unresolved_by_order.setdefault(
                        order_id,
                        {"order_id": order_id, "order_no": str(row.get("order_no") or ""), "skus": []},
                    )["skus"].append(sku or "未识别 SKU")
        return {
            "genericSizeOrders": list(generic_by_order.values()),
            "unresolvedSizeOrders": list(unresolved_by_order.values()),
        }

    @staticmethod
    def find_named_option_id(options: Any, target: str) -> str:
        return str(find_named_option(options, target).get("id") or "")

    @staticmethod
    def format_named_options(options: Any) -> str:
        names = [option_display_name(option) for _option_id, option in iter_named_options(options)]
        names = [name for name in names if name]
        return "、".join(names[:12])

    def change_order_detail_size(
        self,
        *,
        order_detail_id: int | str,
        target_size: str,
        relation_type: int = 1,
    ) -> dict[str, Any]:
        edit_detail = self.get_order_edit_detail(order_detail_id)
        size_map = edit_detail.get("size") or {}
        color_map = edit_detail.get("color") or {}
        current = edit_detail.get("data") or {}
        product_id = edit_detail.get("product_id") or current.get("product_id")
        target = str(target_size or "").strip()
        target_size_id = self.find_named_option_id(size_map, target)
        size_source = "getEditDetail"
        product_info: dict[str, Any] = {}
        if not target_size_id and product_id:
            product_info = self.get_order_product_info(product_id)
            searched_size_map = product_info.get("size") or {}
            searched_target_size_id = self.find_named_option_id(searched_size_map, target)
            if searched_target_size_id:
                size_map = searched_size_map
                target_size_id = searched_target_size_id
                size_source = "getProductInfo"
        if not target_size_id:
            option_text = self.format_named_options(size_map)
            extra = f"；当前可选：{option_text}" if option_text else ""
            raise RuntimeError(f"订单明细 {order_detail_id} 找不到尺码：{target}（已搜索产品编码{extra}）")

        current_size_id = normalize_option_id(current.get("size_id") or current.get("sizeId"))
        current_option = find_option_by_id(size_map, current_size_id)
        current_size = str(current_option.get("name") or current.get("size") or "").strip()
        if current_size_id and normalize_option_id(target_size_id) == current_size_id:
            return {
                "orderDetailId": str(order_detail_id),
                "sku": str(current.get("sku") or ""),
                "fromSize": current_size,
                "fromSizeId": current_size_id,
                "toSize": target,
                "toSizeId": normalize_option_id(target_size_id),
                "skipped": True,
                "reason": "当前尺码 ID 已是目标尺码 ID，未重复提交",
            }
        current_color = str(current.get("colour") or current.get("color") or "").strip()
        current_color_id = str(current.get("colour_id") or current.get("color_id") or "").strip()
        color_source = "orderDetail"
        if not current_color_id and current_color:
            current_color_id = self.find_named_option_id(color_map, current_color)
            color_source = "getEditDetail"
        if not current_color_id and current_color and product_id:
            if not product_info:
                product_info = self.get_order_product_info(product_id)
            current_color_id = self.find_named_option_id(product_info.get("color") or {}, current_color)
            color_source = "getProductInfo"
        if not current_color_id:
            raise RuntimeError(f"订单明细 {order_detail_id} 找不到原颜色ID，已阻止提交，避免颜色被默认值覆盖")

        payload = {
            "productId": product_id,
            "colourId": current_color_id,
            "sizeId": target_size_id,
            "buyNumber": current.get("buy_number") or 1,
            "is_img_custom": current.get("is_img_custom") or "",
            "fabric_id": current.get("fabric_id") or "",
            "order_detail_id": current.get("id") or order_detail_id,
            "order_id": current.get("order_id"),
            "isSave": 1,
            "type": relation_type,
            "lange": "zh",
        }
        response = self.post("/order/relateOrderDetailSave", payload)
        return {
            "orderDetailId": str(order_detail_id),
            "sku": str(current.get("sku") or ""),
            "fromSize": current_size,
            "fromSizeId": current_size_id,
            "toSize": target,
            "toSizeId": normalize_option_id(target_size_id),
            "skipped": False,
            "sizeSource": size_source,
            "color": current_color,
            "colorId": current_color_id,
            "colorSource": color_source,
            "payload": payload,
            "response": response,
        }

    def filter_live_payment_size_targets(self, items: list[dict[str, Any]]) -> dict[str, Any]:
        # 尺码编辑器同时展示 JIT 和 VMI；校验也必须覆盖全部待付款订单。
        rows = self.iter_orders(status=2, limit=100)
        live_detail_ids: set[str] = set()
        for row in rows:
            for detail in row.get("detail") or []:
                if not isinstance(detail, dict):
                    continue
                detail_id = detail.get("id") or detail.get("order_detail_id") or detail.get("item_id")
                if detail_id:
                    live_detail_ids.add(str(detail_id))

        live_items: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        for item in items:
            order_detail_id = str(item.get("order_detail_id") or "")
            if order_detail_id and order_detail_id in live_detail_ids:
                live_items.append(item)
            else:
                skipped.append(
                    {
                        "orderNo": item.get("order_no"),
                        "sku": item.get("sku"),
                        "orderDetailId": order_detail_id,
                        "targetSize": item.get("target_size"),
                        "error": "订单状态已变化或明细不存在，已跳过",
                    }
                )
        return {"items": live_items, "skipped": skipped, "liveOrderCount": len(rows)}

    def change_order_detail_sizes(self, items: list[dict[str, Any]], *, relation_type: int = 1) -> dict[str, Any]:
        live_filter = self.filter_live_payment_size_targets(items)
        items = live_filter["items"]
        results: list[dict[str, Any]] = []
        failed: list[dict[str, Any]] = list(live_filter["skipped"])
        for item in items:
            order_detail_id = item.get("order_detail_id")
            target_size = str(item.get("target_size") or "").strip()
            try:
                result = self.change_order_detail_size(
                    order_detail_id=order_detail_id,
                    target_size=target_size,
                    relation_type=relation_type,
                )
                result["orderNo"] = item.get("order_no")
                results.append(result)
            except Exception as exc:  # noqa: BLE001
                failed.append(
                    {
                        "orderNo": item.get("order_no"),
                        "sku": item.get("sku"),
                        "orderDetailId": str(order_detail_id or ""),
                        "targetSize": target_size,
                        "error": str(exc),
                    }
                )
        success_count = sum(1 for result in results if not result.get("skipped"))
        skipped_count = sum(1 for result in results if result.get("skipped"))
        return {
            "message": f"尺码修改完成：提交成功 {success_count}，已是目标 {skipped_count}，失败 {len(failed)}",
            "successCount": success_count,
            "skippedCount": skipped_count,
            "failedCount": len(failed),
            "skippedStaleCount": len(live_filter["skipped"]),
            "results": results,
            "failed": failed,
        }

    def list_orders(
        self,
        *,
        status: int | str,
        page: int = 1,
        limit: int = 100,
        plat_order_type: int | str | None = None,
        extra: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        payload: dict[str, Any] = {"status": str(status), "page": str(page), "limit": str(limit)}
        if plat_order_type is not None:
            payload["plat_order_type"] = str(plat_order_type)
        if extra:
            payload.update(extra)
        result = self.post("/order/getUserOrderList", payload)
        return result.get("data") or {}

    def iter_orders(
        self,
        *,
        status: int | str,
        plat_order_type: int | str | None = None,
        limit: int = 100,
        max_pages: int = 50,
        extra: dict[str, Any] | None = None,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for page in range(1, max_pages + 1):
            chunk = self.list_orders(
                status=status,
                page=page,
                limit=limit,
                plat_order_type=plat_order_type,
                extra=extra,
            )
            data = chunk.get("data") or []
            rows.extend(data)
            if not data or page >= int(chunk.get("last_page") or page):
                break
        return rows

    def get_order_total(self, *, status: int | str, plat_order_type: int | str | None = None) -> tuple[int, dict[str, Any]]:
        data = self.list_orders(status=status, page=1, limit=1, plat_order_type=plat_order_type)
        return int(data.get("total") or 0), data

    def get_today_waybill_failed_summary(self) -> dict[str, Any]:
        paytime = [f"{time.strftime('%Y-%m-%d')} 00:00:00", time.strftime("%Y-%m-%d %H:%M:%S")]
        failed_rows: list[dict[str, Any]] = []
        for status in WAYBILL_MONITORED_STATUSES:
            rows = self.iter_orders(status=status, limit=100, extra={"paytime": paytime})
            failed_rows.extend(row for row in rows if int(row.get("express_status") or 0) == WAYBILL_FAILED_EXPRESS_STATUS)
        order_nos = [str(row.get("order_no") or "") for row in failed_rows if row.get("order_no")]
        return {
            "count": len(failed_rows),
            "orderNos": order_nos,
            "paytime": paytime,
            "statuses": list(WAYBILL_MONITORED_STATUSES),
        }

    def get_tag_by_name(self, name: str) -> dict[str, Any] | None:
        target = str(name or "").strip().upper()
        for item in self.get_platform_tags():
            if str(item.get("name") or "").strip().upper() == target:
                return item
        return None

    def build_tag_name_map(self) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for item in self.get_platform_tags():
            tag_id = str(item.get("id"))
            tag_name = str(item.get("name") or "").strip()
            if tag_name:
                mapping[tag_id] = tag_name
        return mapping

    def get_dashboard_summary(self) -> dict[str, Any]:
        tags = self.get_platform_tags()
        wait_edit_total, wait_edit_data = self.get_order_total(status=1)
        wait_pay_total, _wait_pay_data = self.get_order_total(status=2)
        paid_total, _paid_data = self.get_order_total(status=3)
        profile_info: dict[str, Any] = {}
        try:
            profile_info = self.get_user_profile()
        except Exception as exc:  # noqa: BLE001
            profile_info = {"balanceError": str(exc)}

        jit_tag = next((item for item in tags if str(item.get("name") or "").strip().upper() == "JIT"), None)
        vmi_tag = next((item for item in tags if str(item.get("name") or "").strip().upper() == "VMI"), None)

        jit_total = self.get_order_total(status=1, plat_order_type=jit_tag["id"])[0] if jit_tag else 0
        vmi_total = self.get_order_total(status=1, plat_order_type=vmi_tag["id"])[0] if vmi_tag else 0
        wait_pay_jit_total = self.get_order_total(status=2, plat_order_type=jit_tag["id"])[0] if jit_tag else 0
        waybill_failed = self.get_today_waybill_failed_summary()
        unknown_total = max(wait_edit_total - jit_total - vmi_total, 0)

        status_list = wait_edit_data.get("statusList") or {}
        return {
            "account": {
                "username": profile_info.get("username") or self.session.user_info.get("username"),
                "nickname": profile_info.get("nickname") or self.session.user_info.get("nickname"),
                "factoryId": self.session.factory_id,
                "origin": self.session.origin,
                "pageTitle": self.session.page_title,
                "balance": profile_info.get("money"),
                "balanceError": profile_info.get("balanceError"),
            },
            "counts": {
                "waitEdit": wait_edit_total,
                "waitPay": wait_pay_total,
                "paid": paid_total,
                "waitPayJit": wait_pay_jit_total,
                "jit": jit_total,
                "vmi": vmi_total,
                "other": unknown_total,
                "waybillFailed": waybill_failed.get("count") or 0,
            },
            "waybillFailed": waybill_failed,
            "tags": {
                "jit": jit_tag,
                "vmi": vmi_tag,
            },
            "statusList": status_list,
        }

    def get_gui_order_rows(self, *, statuses: Iterable[int] = (1,), limit: int = 100) -> list[dict[str, Any]]:
        tag_map = self.build_tag_name_map()
        rows: list[dict[str, Any]] = []
        for status in statuses:
            for row in self.iter_orders(status=status, limit=limit):
                row_status = int(row.get("status") or 0)
                plat_order_type = row.get("plat_order_type")
                tag_name = tag_map.get(str(plat_order_type), "")
                if not tag_name and plat_order_type is not None:
                    if str(plat_order_type) == "0":
                        tag_name = "VMI"
                    else:
                        tag_name = f"标签{plat_order_type}"
                rows.append(
                    {
                        "id": str(row.get("id") or ""),
                        "order_id": str(row.get("order_id") or ""),
                        "order_no": str(row.get("order_no") or ""),
                        "status": row_status,
                        "status_text": STATUS_TEXT_MAP.get(row_status, str(row.get("status_text") or row_status)),
                        "plat_order_type": str(plat_order_type if plat_order_type is not None else ""),
                        "tag_name": tag_name or "未标记",
                        "shop_name": str(row.get("shop_name") or ""),
                        "plat_name": str(row.get("plat_name") or ""),
                        "buy_number_count": int(row.get("buy_number_count") or 0),
                        "buy_type_count": int(row.get("buy_type_count") or 0),
                        "express_name": str(row.get("express_name") or ""),
                        "remarks": str(row.get("remarks") or ""),
                        "plattime": str(row.get("plattime") or ""),
                        "detail": row.get("detail") or [],
                    }
                )
        return rows

    def get_gui_snapshot(self) -> dict[str, Any]:
        orders = self.get_gui_order_rows(statuses=GUI_STATUS_TABS)
        orders_by_status: dict[str, list[dict[str, Any]]] = {str(status): [] for status in GUI_STATUS_TABS}
        for row in orders:
            orders_by_status.setdefault(str(row.get("status") or ""), []).append(row)
        return {
            "summary": self.get_dashboard_summary(),
            "orders": orders,
            "ordersByStatus": orders_by_status,
        }

    def build_change_express_payloads(
        self,
        rows: list[dict[str, Any]],
        *,
        plat_id: int | str,
        express_price: Any,
        express_id: int | str,
        express_company: str,
    ) -> list[dict[str, Any]]:
        payloads = []
        for row in rows:
            payloads.append(
                {
                    "plat_id": str(plat_id),
                    "order_id": str(row["order_id"]),
                    "express_price": str(express_price if express_price is not None else 0),
                    "express_id": str(express_id),
                    "express_company": str(express_company),
                }
            )
        return payloads

    def apply_change_express(self, payloads: list[dict[str, Any]], *, commit: bool) -> list[dict[str, Any]]:
        results = []
        for payload in payloads:
            if not commit:
                results.append({"payload": payload, "status": "dry-run"})
                continue
            response = self.post("/logistic/changeExpressSave", payload)
            results.append({"payload": payload, "status": "committed", "response": response})
        return results

    def auto_apply_logistics(
        self,
        *,
        status: int = 1,
        limit: int = 100,
        commit: bool = False,
        target_keyword: str = "temu",
        order_ids: Iterable[str] | None = None,
    ) -> dict[str, Any]:
        jit_tag = self.get_jit_tag()
        rows = self.iter_orders(status=status, plat_order_type=jit_tag["id"], limit=limit)
        selected_ids = {str(item) for item in order_ids or [] if str(item)}
        if selected_ids:
            rows = [row for row in rows if str(row.get("order_id") or "") in selected_ids]
        if not rows:
            return {
                "message": "待编辑无 JIT，已跳过改物流" if not selected_ids else "选中的订单里无 JIT，已跳过改物流",
                "jitTag": jit_tag,
                "matchedOrders": 0,
                "selectedCount": len(selected_ids),
                "commitRequested": commit,
                "commitExecuted": False,
            }

        plat_ids = sorted({str(row.get("plat_id") or "") for row in rows if row.get("plat_id") is not None})
        if len(plat_ids) != 1:
            raise RuntimeError(f"待编辑 JIT 存在多个 plat_id：{','.join(plat_ids)}")

        plat_id = plat_ids[0]
        selected_company = self.find_temu_company(plat_id, target_keyword)
        preview = self.get_order_company_preview(
            plat_id,
            [str(row["order_id"]) for row in rows],
            selected_company["id"],
        )
        express = preview.get("express") or {}
        if isinstance(express, list):
            express = express[0] if express else {}
        ok_num = int(preview.get("ok_num") or 0)
        no_num = int(preview.get("no_num") or 0)
        preview_all_ok = ok_num == len(rows) and no_num == 0
        commit_executed = bool(commit and preview_all_ok)
        payloads = self.build_change_express_payloads(
            rows,
            plat_id=plat_id,
            express_price=express.get("price", 0),
            express_id=express.get("id") or selected_company.get("id"),
            express_company=express.get("name") or selected_company.get("name"),
        )
        result = self.apply_change_express(payloads, commit=commit_executed)
        return {
            "jitTag": jit_tag,
            "platId": plat_id,
            "matchedOrders": len(rows),
            "orderIds": [str(row.get("order_id") or "") for row in rows if row.get("order_id")],
            "orderNos": [str(row.get("order_no") or "") for row in rows if row.get("order_no")],
            "selectedCount": len(selected_ids),
            "okNum": ok_num,
            "noNum": no_num,
            "requiredOkNum": len(rows),
            "commitRequested": commit,
            "commitExecuted": commit_executed,
            "selectedCompany": {
                "id": selected_company.get("id"),
                "logistic_id": selected_company.get("logistic_id"),
                "name": selected_company.get("name"),
                "code": selected_company.get("code"),
            },
            "preview": preview,
            "result": result,
            "message": (
                "改物流预检未全量通过，未提交"
                if commit and not commit_executed
                else "改物流已提交"
                if commit_executed
                else "改物流 dry-run 完成"
            ),
        }

    def get_check_order(self, order_ids: Iterable[str]) -> dict[str, Any]:
        result = self.post(
            "/order/getCheckOrder",
            {
                "ids": ",".join(str(item) for item in order_ids),
                "lange": "zh",
            },
        )
        return result.get("data") or {}

    def order_pay(self, order_ids: Iterable[str], *, commit: bool, force: bool = False) -> dict[str, Any]:
        payload: dict[str, Any] = {"ids": ",".join(str(item) for item in order_ids)}
        if force:
            payload["is_force"] = 1
        if not commit:
            return {"status": "dry-run", "payload": payload}
        return self.post("/order/orderPay", payload)

    def resolve_payment_orders(
        self,
        *,
        ids: list[str] | None = None,
        status: int = 2,
        limit: int = 100,
        plat_order_type: int | str | None = None,
        exclude_order_nos: set[str] | None = None,
    ) -> dict[str, Any]:
        selected_ids = {str(item) for item in ids or [] if str(item)}
        rows = self.iter_orders(status=status, plat_order_type=plat_order_type, limit=limit)
        if selected_ids:
            rows = [row for row in rows if str(row.get("order_id") or "") in selected_ids]

        excludes = {normalize_order_no(item) for item in exclude_order_nos or set() if normalize_order_no(item)}
        payable_rows = []
        excluded_rows = []
        seen_order_nos = set()
        for row in rows:
            order_no = normalize_order_no(row.get("order_no"))
            if order_no:
                seen_order_nos.add(order_no)
            if order_no and order_no in excludes:
                excluded_rows.append(row)
            else:
                payable_rows.append(row)

        return {
            "rows": payable_rows,
            "ids": [str(row["order_id"]) for row in payable_rows if row.get("order_id")],
            "excludedRows": excluded_rows,
            "excludedOrderNos": [str(row.get("order_no") or "") for row in excluded_rows],
            "excludeNotFound": sorted(excludes - seen_order_nos),
            "totalMatchedBeforeExclude": len(rows),
        }

    def wait_for_orders_in_status(
        self,
        *,
        order_ids: Iterable[str],
        status: int = 2,
        plat_order_type: int | str | None = None,
        timeout_seconds: int = 60,
        interval_seconds: int = 3,
    ) -> dict[str, Any]:
        target_ids = {str(item) for item in order_ids if str(item)}
        deadline = time.time() + timeout_seconds
        last_rows: list[dict[str, Any]] = []
        while True:
            rows = self.iter_orders(status=status, plat_order_type=plat_order_type, limit=100)
            last_rows = rows
            found_rows = [row for row in rows if str(row.get("order_id") or "") in target_ids]
            found_ids = {str(row.get("order_id") or "") for row in found_rows}
            missing_ids = sorted(target_ids - found_ids)
            if not missing_ids:
                return {
                    "ok": True,
                    "rows": found_rows,
                    "foundIds": sorted(found_ids),
                    "missingIds": [],
                    "status": status,
                    "platOrderType": str(plat_order_type or ""),
                }
            if time.time() >= deadline:
                return {
                    "ok": False,
                    "rows": found_rows,
                    "foundIds": sorted(found_ids),
                    "missingIds": missing_ids,
                    "status": status,
                    "platOrderType": str(plat_order_type or ""),
                    "lastSeenCount": len(last_rows),
                }
            time.sleep(interval_seconds)

    def download_order_images(
        self,
        *,
        status: int = 2,
        output_dir: str = "",
        order_ids: Iterable[str] | None = None,
        plat_order_type: int | str | None = None,
    ) -> dict[str, Any]:
        rows = self.iter_orders(status=status, plat_order_type=plat_order_type, limit=100)
        selected_ids = {str(item) for item in order_ids or [] if str(item)}
        if selected_ids:
            rows = [row for row in rows if str(row.get("order_id") or "") in selected_ids]
        if not rows:
            return {
                "message": "待付款无单" if not selected_ids else "选中的订单里无待付款单",
                "success": False,
                "outputDir": "",
                "manifest": "",
                "downloaded": 0,
                "failed": 0,
                "records": [],
            }

        if output_dir:
            dir_path = Path(output_dir)
        else:
            dir_path = get_default_output_dir()
        dir_path.mkdir(parents=True, exist_ok=True)

        records: list[dict[str, Any]] = []
        for row in rows:
            order_no = str(row.get("order_no") or "")
            shop_name = str(row.get("shop_name") or "")
            for detail in row.get("detail") or []:
                image_url = str(detail.get("image") or detail.get("pic") or "")
                sku = str(detail.get("sku") or "")

                if not image_url:
                    records.append(
                        {
                            "order_no": order_no,
                            "sku": sku,
                            "shop_name": shop_name,
                            "image_url": "",
                            "file": "",
                            "status": "skip:no-image",
                            "bytes": 0,
                        }
                    )
                    continue

                if image_url.startswith("//"):
                    image_url = f"https:{image_url}"

                filename = "__".join(
                    [sanitize_name(order_no), sanitize_name(sku), sanitize_name(f"Temu_{shop_name}")]
                ) + guess_ext(image_url)
                destination = unique_path(dir_path / filename)

                try:
                    response = self.http.get(
                        image_url,
                        headers={
                            "User-Agent": self.session.user_agent,
                            "Referer": f"{self.session.origin}/#/Orderlist",
                        },
                        timeout=60,
                    )
                    response.raise_for_status()
                    destination.write_bytes(response.content)
                    records.append(
                        {
                            "order_no": order_no,
                            "sku": sku,
                            "shop_name": shop_name,
                            "image_url": image_url,
                            "file": str(destination),
                            "status": "ok",
                            "bytes": len(response.content),
                        }
                    )
                except Exception as exc:  # noqa: BLE001
                    records.append(
                        {
                            "order_no": order_no,
                            "sku": sku,
                            "shop_name": shop_name,
                            "image_url": image_url,
                            "file": "",
                            "status": f"error:{type(exc).__name__}",
                            "bytes": 0,
                        }
                    )

        manifest = dir_path / "api_download_manifest.json"
        manifest.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
        downloaded = sum(1 for item in records if item["status"] == "ok")
        failed_records = [item for item in records if item["status"] != "ok"]
        success = bool(records) and downloaded > 0 and not failed_records
        return {
            "message": (
                "待付款图片已下载，等待验图"
                if success
                else f"图片下载不完整：成功 {downloaded}，失败或缺图 {len(failed_records)}"
            ),
            "success": success,
            "outputDir": str(dir_path),
            "manifest": str(manifest),
            "downloaded": downloaded,
            "failed": len(failed_records),
            "records": records,
        }


def with_auth_file_session(args: argparse.Namespace, worker) -> Any:
    payload = load_auth_state_payload(getattr(args, "auth_file", DEFAULT_AUTH_STATE_FILE))
    session = LandwuSession(
        browser_config=BrowserConfig("", "", "", "", "", ""),
        browser_pid=0,
        origin=payload["origin"],
        host=payload["host"],
        href=payload["href"],
        access_token=payload["accessToken"],
        factory_id=int(payload["factoryId"] or 0),
        master_factory_id=str(payload["masterFactoryId"] or ""),
        session_cookie=str(payload.get("sessionCookie") or ""),
        user_info=payload.get("userInfo") or {},
        user_agent=payload.get("userAgent") or "Mozilla/5.0",
        page_title=payload.get("pageTitle") or "本地同步登录态",
        page_url=payload.get("pageUrl") or "",
        auth_source=AUTH_SOURCE_FILE,
    )
    client = LandwuClient(session)
    return worker(session, client)


def with_browser_session(args: argparse.Namespace, worker) -> Any:
    runtime = ensure_browser_runtime(args)
    playwright = None
    browser = None
    try:
        playwright, browser, context = connect_browser(runtime.port)
        page = pick_order_page(context)
        payload = extract_session_payload(page)
        factory_id = int(payload["userInfo"].get("factory_id") or 0)
        session = LandwuSession(
            browser_config=runtime.config,
            browser_pid=runtime.proc.pid if runtime.proc else 0,
            origin=payload["origin"],
            host=payload["host"],
            href=payload["href"],
            access_token=payload["accessToken"],
            factory_id=factory_id,
            master_factory_id=f"6{factory_id}",
            session_cookie=str(payload.get("sessionCookie") or ""),
            user_info=payload.get("userInfo") or {},
            user_agent=payload.get("userAgent") or "Mozilla/5.0",
            page_title=page.title(),
            page_url=page.url,
            auth_source=AUTH_SOURCE_BROWSER,
        )
        client = LandwuClient(session)
        return worker(session, client)
    finally:
        if playwright is not None:
            try:
                playwright.stop()
            except Exception:  # noqa: BLE001
                pass
        if args.close_on_finish and runtime.launched and runtime.proc is not None:
            terminate_process_tree(runtime.proc.pid)


def with_landwu_session(args: argparse.Namespace, worker) -> Any:
    return with_auth_file_session(args, worker)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(add_help=True, description="Landwu 做单脚本（macOS 单文件版）")
    parser.add_argument("--auth-source", choices=[AUTH_SOURCE_FILE], default=DEFAULT_AUTH_SOURCE)
    parser.add_argument("--auth-file", default=DEFAULT_AUTH_STATE_FILE)

    sub = parser.add_subparsers(dest="command")

    sub.add_parser("auth")
    sub.add_parser("tags")

    company = sub.add_parser("company-list")
    company.add_argument("--plat-id", required=True)

    list_parser = sub.add_parser("list")
    list_parser.add_argument("--status", required=True, type=int)
    list_parser.add_argument("--page", type=int, default=1)
    list_parser.add_argument("--limit", type=int, default=20)
    list_parser.add_argument("--plat-order-type", type=int)

    apply_parser = sub.add_parser("apply-logistics")
    apply_parser.add_argument("--status", type=int, default=1)
    apply_parser.add_argument("--limit", type=int, default=100)
    apply_parser.add_argument("--commit", action="store_true")
    apply_parser.add_argument("--target-keyword", default="temu")

    download_parser = sub.add_parser("download-images")
    download_parser.add_argument("--status", type=int, default=2)
    download_parser.add_argument("--output-dir")

    pay_parser = sub.add_parser("pay-orders")
    pay_parser.add_argument("--ids")
    pay_parser.add_argument("--status", type=int, default=2)
    pay_parser.add_argument("--limit", type=int, default=100)
    pay_parser.add_argument("--commit", action="store_true")
    pay_parser.add_argument("--force", action="store_true")

    process_parser = sub.add_parser("process-until-review")
    process_parser.add_argument("--limit", type=int, default=100)
    process_parser.add_argument("--commit-logistics", action="store_true")
    process_parser.add_argument("--target-keyword", default="temu")
    process_parser.add_argument("--output-dir")

    return parser


def build_runtime_args(
    *,
    auth_source: str = DEFAULT_AUTH_SOURCE,
    auth_file: str = DEFAULT_AUTH_STATE_FILE,
) -> argparse.Namespace:
    return argparse.Namespace(
        auth_source=auth_source,
        auth_file=auth_file,
    )


class LandwuGuiApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"Landwu 做单助手 Pro v{APP_VERSION}")
        self.root.geometry("1180x720")
        self.root.minsize(1000, 600)

        self.style = ttk.Style()
        self._setup_style()

        app_settings = load_app_settings()
        self.auth_file_var = tk.StringVar(value=DEFAULT_AUTH_STATE_FILE)
        self.output_dir_var = tk.StringVar(value=str(get_default_output_dir()))
        self.composition_db_folder_var = tk.StringVar(
            value=str(app_settings.get("composition_db_folder") or COMPOSITION_DB_FOLDER)
        )
        self.status_var = tk.StringVar(value="待命")
        self.account_var = tk.StringVar(value="-")
        self.factory_var = tk.StringVar(value="-")
        self.wait_edit_var = tk.StringVar(value="0")
        self.wait_pay_var = tk.StringVar(value="0")
        self.paid_var = tk.StringVar(value="0")
        self.jit_var = tk.StringVar(value="0")
        self.vmi_var = tk.StringVar(value="0")
        self.other_var = tk.StringVar(value="0")
        self.waybill_failed_var = tk.StringVar(value="0")
        self.bulk_uncheck_order_nos_var = tk.StringVar(value="")
        self.selection_hint_var = tk.StringVar(value="待付款页第一列小方框打勾表示会支付，点一下取消。")
        self.auto_refresh_enabled_var = tk.BooleanVar(value=False)
        self.auto_refresh_mode_var = tk.StringVar(value="每隔分钟")
        self.auto_refresh_interval_var = tk.StringVar(value="10")
        self.auto_refresh_time_var = tk.StringVar(value="09:00")
        self.auto_refresh_status_var = tk.StringVar(value="自动刷新关闭")
        self.summary_line_var = tk.StringVar(value="账号：-    待编辑：0    JIT：0    VMI：0    待付款 JIT：0    已支付：0")
        self.balance_var = tk.StringVar(value="-")
        self.toolbar_title_var = tk.StringVar(value="待编辑操作")
        self.waybill_failed_card: tk.Frame | None = None
        self.waybill_failed_title_label: ttk.Label | None = None
        self.waybill_failed_value_label: ttk.Label | None = None
        self.low_balance_alerted = False

        self.queue: Queue[tuple[str, str, Any]] = Queue()
        self.image_queue: Queue[tuple[tk.Widget, Any, list[Any], str | None]] = Queue()
        self.image_executor = ThreadPoolExecutor(max_workers=6, thread_name_prefix="landwu-image")
        self.busy = False
        self.current_success_handler = None
        self.current_error_handler = None
        self.current_show_error_popup = True
        self.pending_update_source = ""
        self.buttons: list[ttk.Button] = []
        self.trees: dict[int, ttk.Treeview] = {}
        self._mousewheel_targets: dict[str, tk.Widget] = {}
        self._mousewheel_hover_target: tk.Widget | None = None
        self._mousewheel_dispatcher_installed = False
        self._touchpad_scroll_supported: bool | None = None
        self._scroll_pixel_remainder: dict[str, int] = {}
        self.tab_statuses = list(GUI_STATUS_TABS)
        self.order_rows_by_status_iid: dict[int, dict[str, dict[str, Any]]] = {status: {} for status in GUI_STATUS_TABS}
        self.unchecked_edit_order_ids: set[str] = set()
        self.edit_check_vars: dict[str, tk.BooleanVar] = {}
        self.edit_card_frames: dict[str, tk.Frame] = {}
        self.selected_edit_iid = ""
        self.unchecked_payment_order_ids: set[str] = set()
        self.payment_card_images: list[Any] = []
        self.payment_check_vars: dict[str, tk.BooleanVar] = {}
        self.payment_card_frames: dict[str, tk.Frame] = {}
        self.selected_payment_iid = ""
        self.payment_render_generation = 0
        self.size_editor_window: tk.Toplevel | None = None
        self.preview_window: tk.Toplevel | None = None
        self.settings_window: tk.Toplevel | None = None
        self.log_window: tk.Toplevel | None = None
        self.log_text: ScrolledText | None = None
        self.log_lines: list[str] = []
        self.auto_refresh_after_id: str | None = None
        self.auto_refresh_next_ts: float | None = None
        self.auth_sync_status_var = tk.StringVar(value="登录态接收：未启动，需要同步时点“接收登录态3分钟”")
        self.auth_sync_receiver = AuthSyncReceiver(
            lambda: self.auth_file_var.get().strip() or DEFAULT_AUTH_STATE_FILE,
            self._on_auth_synced,
        )
        self.auth_sync_stop_after_id: str | None = None

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(150, self._process_queue)
        self.root.after(100, self._process_image_queue)
        self.root.after(300, lambda: self.refresh_summary(startup=True))

    def _setup_style(self) -> None:
        if "clam" in self.style.theme_names():
            self.style.theme_use("clam")

        self.root.configure(background="#EEF2F6")
        self.style.configure(".", font=("Microsoft YaHei UI", 9), background="#EEF2F6")
        self.style.configure("TFrame", background="#EEF2F6")
        self.style.configure("Surface.TFrame", background="#FFFFFF")
        self.style.configure("Toolbar.TFrame", background="#FFFFFF")
        self.style.configure("Status.TFrame", background="#17212B")
        self.style.configure("TLabel", background="#FFFFFF", foreground="#1F2933")
        self.style.configure("AppTitle.TLabel", background="#FFFFFF", foreground="#102A43", font=("Microsoft YaHei UI", 13, "bold"))
        self.style.configure("Version.TLabel", background="#FFFFFF", foreground="#7B8794", font=("Segoe UI", 8))
        self.style.configure("Summary.TLabel", background="#FFFFFF", foreground="#52606D", font=("Microsoft YaHei UI", 9))
        self.style.configure("BalanceTitle.TLabel", background="#F0F9FF", foreground="#64748B", font=("Microsoft YaHei UI", 8))
        self.style.configure("BalanceValue.TLabel", background="#F0F9FF", foreground="#0369A1", font=("Segoe UI", 13, "bold"))
        self.style.configure("SettingsBalance.TLabel", background="#FFFFFF", foreground="#0369A1", font=("Segoe UI", 11, "bold"))
        self.style.configure("WaybillOkTitle.TLabel", background="#F8FAFC", foreground="#64748B", font=("Microsoft YaHei UI", 8))
        self.style.configure("WaybillOkValue.TLabel", background="#F8FAFC", foreground="#334155", font=("Segoe UI", 13, "bold"))
        self.style.configure("WaybillFailTitle.TLabel", background="#FEF2F2", foreground="#991B1B", font=("Microsoft YaHei UI", 8))
        self.style.configure("WaybillFailValue.TLabel", background="#FEF2F2", foreground="#DC2626", font=("Segoe UI", 13, "bold"))
        self.style.configure("ToolbarTitle.TLabel", background="#FFFFFF", foreground="#243B53", font=("Microsoft YaHei UI", 10, "bold"))
        self.style.configure("Hint.TLabel", background="#EEF2F6", foreground="#6C757D", font=("Microsoft YaHei UI", 8, "bold"))
        self.style.configure("StatusText.TLabel", background="#17212B", foreground="#E6EDF3", font=("Microsoft YaHei UI", 9))
        self.style.configure("MetricTitle.TLabel", background="#F8FAFC", foreground="#6B7280", font=("Microsoft YaHei UI", 8))
        self.style.configure("MetricNum.TLabel", background="#F8FAFC", foreground="#0D6EFD", font=("Segoe UI", 13, "bold"))
        self.style.configure("TButton", font=("Microsoft YaHei UI", 9), padding=(8, 4), focuscolor="")
        self.style.configure(
            "Accent.TButton",
            font=("Microsoft YaHei UI", 9, "bold"),
            foreground="#FFFFFF",
            background="#0D6EFD",
            bordercolor="#0D6EFD",
            lightcolor="#0D6EFD",
            darkcolor="#0D6EFD",
        )
        self.style.map("Accent.TButton", background=[("active", "#0B5ED7"), ("disabled", "#9EC5FE")])
        self.style.configure(
            "Match.TButton",
            font=("Microsoft YaHei UI", 11, "bold"),
            padding=(18, 7),
            foreground="#FFFFFF",
            background="#16A34A",
            bordercolor="#16A34A",
            lightcolor="#16A34A",
            darkcolor="#16A34A",
        )
        self.style.map("Match.TButton", background=[("active", "#15803D"), ("disabled", "#86EFAC")])
        self.style.configure(
            "Danger.TButton",
            font=("Microsoft YaHei UI", 9, "bold"),
            foreground="#FFFFFF",
            background="#DC3545",
            bordercolor="#DC3545",
            lightcolor="#DC3545",
            darkcolor="#DC3545",
        )
        self.style.map("Danger.TButton", background=[("active", "#BB2D3B"), ("disabled", "#F1AEB5")])
        self.style.configure("Ghost.TButton", font=("Microsoft YaHei UI", 9), padding=(8, 4), background="#F8FAFC")
        self.style.configure("Treeview", font=("Microsoft YaHei UI", 9), rowheight=28, borderwidth=0)
        self.style.configure(
            "Treeview.Heading",
            font=("Microsoft YaHei UI", 9, "bold"),
            background="#F8FAFC",
            foreground="#495057",
            padding=3,
        )
        self.style.map("Treeview", background=[("selected", "#E7F1FF")], foreground=[("selected", "#000000")])

    def _build_ui(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)

        header = ttk.Frame(self.root, style="Surface.TFrame", padding=(10, 8))
        header.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        header.columnconfigure(1, weight=1)
        title_frame = ttk.Frame(header, style="Surface.TFrame")
        title_frame.grid(row=0, column=0, sticky="w")
        ttk.Label(title_frame, text="Landwu 做单助手", style="AppTitle.TLabel").pack(side="left")
        ttk.Label(title_frame, text=f"v{APP_VERSION}", style="Version.TLabel").pack(side="left", padx=(8, 0))
        ttk.Label(header, textvariable=self.summary_line_var, style="Summary.TLabel").grid(row=0, column=1, sticky="w", padx=(16, 8))
        refresh_btn = ttk.Button(header, text="刷新订单", command=self.refresh_summary, style="Accent.TButton")
        self.waybill_failed_card = tk.Frame(header, bg="#F8FAFC", highlightbackground="#CBD5E1", highlightthickness=1, padx=10, pady=3)
        self.waybill_failed_card.grid(row=0, column=2, sticky="e", padx=(8, 4))
        self.waybill_failed_title_label = ttk.Label(self.waybill_failed_card, text="运单失败", style="WaybillOkTitle.TLabel")
        self.waybill_failed_title_label.pack(anchor="w")
        self.waybill_failed_value_label = ttk.Label(
            self.waybill_failed_card,
            textvariable=self.waybill_failed_var,
            style="WaybillOkValue.TLabel",
        )
        self.waybill_failed_value_label.pack(anchor="w")
        refresh_btn.grid(row=0, column=3, sticky="e", padx=(4, 0))
        self.buttons.append(refresh_btn)
        ttk.Button(header, text="设置", command=self.open_settings_window, style="Ghost.TButton").grid(row=0, column=4, sticky="e", padx=(6, 0))
        ttk.Button(header, text="日志", command=self.open_log_window, style="Ghost.TButton").grid(row=0, column=5, sticky="e", padx=(6, 0))

        action_frame = ttk.Frame(self.root, style="Toolbar.TFrame", padding=(10, 7))
        action_frame.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))
        action_frame.columnconfigure(10, weight=1)
        ttk.Label(action_frame, textvariable=self.toolbar_title_var, style="ToolbarTitle.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))

        self.toolbar_widgets_by_status: dict[int, list[tk.Widget]] = {1: [], 2: [], 3: []}
        self.toolbar_widgets_by_status[1].append(
            self._add_button(action_frame, 0, 1, "一键 JIT 改物流", self.commit_logistics, style="Accent.TButton")
        )
        self.toolbar_widgets_by_status[1].append(
            self._add_button(action_frame, 0, 2, "改物流并下载图片", self.process_until_review, style="Accent.TButton")
        )
        self.toolbar_widgets_by_status[2].append(
            self._add_button(action_frame, 0, 1, "预检并支付 JIT", self.commit_payment, style="Danger.TButton")
        )
        self.toolbar_widgets_by_status[2].append(
            self._add_button(action_frame, 0, 2, "修改成分尺码", self.open_payment_size_editor, style="Ghost.TButton")
        )
        self.toolbar_widgets_by_status[2].append(
            self._add_button(action_frame, 0, 3, "一键匹配并提交成分", self.quick_match_submit_payment_sizes, style="Match.TButton")
        )

        payment_bulk = ttk.Frame(action_frame, style="Toolbar.TFrame")
        payment_bulk.grid(row=0, column=4, columnspan=6, sticky="ew", padx=(10, 0))
        payment_bulk.columnconfigure(1, weight=1)
        ttk.Label(payment_bulk, text="取消勾选订单号", background="#FFFFFF", foreground="#52606D").grid(row=0, column=0, sticky="w")
        ttk.Entry(payment_bulk, textvariable=self.bulk_uncheck_order_nos_var, width=34).grid(row=0, column=1, sticky="ew", padx=(6, 4))
        bulk_btn = ttk.Button(payment_bulk, text="取消勾选", command=self.bulk_uncheck_payment_orders)
        bulk_btn.grid(row=0, column=2, sticky="e")
        self.buttons.append(bulk_btn)
        self.toolbar_widgets_by_status[2].append(payment_bulk)

        paid_hint = ttk.Label(action_frame, text="已支付页仅查看订单，可右键复制订单号。", background="#FFFFFF", foreground="#6C757D")
        paid_hint.grid(row=0, column=1, columnspan=8, sticky="w")
        self.toolbar_widgets_by_status[3].append(paid_hint)

        table_frame = ttk.Frame(self.root)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=8, pady=(0, 4))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        hint_label = ttk.Label(table_frame, textvariable=self.selection_hint_var, style="Hint.TLabel")
        hint_label.grid(row=0, column=0, sticky="w", pady=(0, 2))

        columns = ("order_no", "status_text", "tag_name", "shop_name", "buy_number_count", "buy_type_count", "plattime")
        headings = {
            "order_no": "WB订单号",
            "status_text": "状态",
            "tag_name": "标签",
            "shop_name": "店铺",
            "buy_number_count": "件数",
            "buy_type_count": "SKU数",
            "plattime": "平台时间",
        }
        widths = {
            "order_no": 180,
            "status_text": 80,
            "tag_name": 80,
            "shop_name": 140,
            "buy_number_count": 60,
            "buy_type_count": 60,
            "plattime": 160,
        }
        self.notebook = ttk.Notebook(table_frame)
        self.notebook.grid(row=1, column=0, sticky="nsew")
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tree_select)
        for status in GUI_STATUS_TABS:
            tab = ttk.Frame(self.notebook)
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(0, weight=1)
            if status == 1:
                self._create_edit_cards_view(tab)
            elif status == 2:
                self._create_payment_cards_view(tab)
            else:
                self._create_order_tree(tab, status, columns, headings, widths)
            self.notebook.add(tab, text=STATUS_TEXT_MAP.get(status, str(status)))
        self.tree = self.trees.get(3)

        status_frame = ttk.Frame(self.root, style="Status.TFrame", padding=(10, 6))
        status_frame.grid(row=3, column=0, sticky="ew", padx=8, pady=(0, 8))
        status_frame.columnconfigure(0, weight=1)
        ttk.Label(status_frame, textvariable=self.status_var, style="StatusText.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(status_frame, textvariable=self.auto_refresh_status_var, style="StatusText.TLabel").grid(row=0, column=1, sticky="e")

        self._show_toolbar_for_status(1)
        self._log("界面已启动。请先点击“刷新统计”获取最新数据。")

    def _build_metric(self, parent: ttk.LabelFrame, column: int, title: str, variable: tk.StringVar) -> None:
        card = tk.Frame(
            parent,
            bg="#F8F9FA",
            highlightbackground="#E9ECEF",
            highlightcolor="#E9ECEF",
            highlightthickness=1,
            padx=8,
            pady=4,
        )
        card.grid(row=0, column=column, sticky="nsew", padx=3, pady=2)
        ttk.Label(card, text=title, foreground="#6C757D", background="#F8F9FA", font=("Microsoft YaHei UI", 8)).pack(anchor="w")
        ttk.Label(card, textvariable=variable, style="MetricNum.TLabel", background="#F8F9FA").pack(anchor="w")

    def _add_button(self, parent: ttk.Frame, row: int, column: int, text: str, command, style: str = "TButton") -> ttk.Button:
        button = ttk.Button(parent, text=text, command=command, style=style)
        button.grid(row=row, column=column, sticky="ew", padx=4, pady=2)
        self.buttons.append(button)
        return button

    def _show_toolbar_for_status(self, status: int) -> None:
        title_map = {
            1: "待编辑操作",
            2: "待付款操作",
            3: "已支付查看",
        }
        self.toolbar_title_var.set(title_map.get(status, "订单操作"))
        for widgets in getattr(self, "toolbar_widgets_by_status", {}).values():
            for widget in widgets:
                try:
                    widget.grid_remove()
                except Exception:
                    pass
        for widget in getattr(self, "toolbar_widgets_by_status", {}).get(status, []):
            try:
                widget.grid()
            except Exception:
                pass

    def _create_order_tree(
        self,
        parent: ttk.Frame,
        status: int,
        columns: tuple[str, ...],
        headings: dict[str, str],
        widths: dict[str, int],
    ) -> None:
        tree = ttk.Treeview(parent, columns=columns, show="headings", selectmode="extended", height=10)
        for key in columns:
            tree.heading(key, text=headings[key])
            tree.column(key, width=widths[key], anchor="center")
        tree_scroll = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=tree_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll.grid(row=0, column=1, sticky="ns")
        tree.bind("<Button-3>", lambda event, tree_status=status: self._show_order_context_menu(event, tree_status))
        tree.bind("<Control-c>", lambda event: self.copy_selected_order_nos())
        tree.bind("<Control-C>", lambda event: self.copy_selected_order_nos())
        tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        # macOS / Windows / Linux 鼠标滚轮滚动 Treeview
        def _tree_mousewheel(event, _tree=tree):
            units = self._mousewheel_units(event)
            if units:
                _tree.yview_scroll(units, "units")
            return "break"
        tree.bind("<MouseWheel>", _tree_mousewheel)
        tree.bind("<Button-4>", _tree_mousewheel)
        tree.bind("<Button-5>", _tree_mousewheel)
        self._bind_touchpad_scroll(tree, lambda event, _tree=tree: self._scroll_touchpad_target(_tree, event))
        self.trees[status] = tree

    def _create_payment_cards_view(self, parent: ttk.Frame) -> None:
        text = tk.Text(
            parent,
            wrap="word",
            bg="#F8F9FA",
            fg="#1F2933",
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground="#CBD5E1",
            highlightcolor="#0D6EFD",
            padx=8,
            pady=6,
            cursor="arrow",
        )
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)
        text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        text.tag_configure("payment_title", font=("Microsoft YaHei UI", 9, "bold"), foreground="#1F2933")
        text.tag_configure("payment_meta", font=("Microsoft YaHei UI", 9), foreground="#52606D")
        text.tag_configure("payment_sku", font=("Microsoft YaHei UI", 9, "bold"), foreground="#52606D", justify="center")
        text.tag_configure("payment_card_gap", spacing1=8, spacing3=8)
        text.tag_configure("payment_selected", background="#E7F1FF")
        text.tag_configure("payment_error", foreground="#B02A37")
        text.bind("<Configure>", self._resize_payment_text_cards, add="+")
        self._bind_widget_mousewheel(text, text)
        self.payment_text = text
        self.payment_canvas = text
        self.payment_body = text
        self._set_payment_cards_empty("待付款订单会显示在这里，每单直接带图片预览。")

    def _create_edit_cards_view(self, parent: ttk.Frame) -> None:
        text = tk.Text(
            parent,
            wrap="word",
            bg="#F8F9FA",
            fg="#1F2933",
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightbackground="#CBD5E1",
            highlightcolor="#0D6EFD",
            padx=8,
            pady=6,
            cursor="arrow",
        )
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)
        text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        text.tag_configure("edit_title", font=("Microsoft YaHei UI", 9, "bold"), foreground="#1F2933")
        text.tag_configure("edit_meta", font=("Microsoft YaHei UI", 9), foreground="#52606D")
        text.tag_configure("edit_card_gap", spacing1=8, spacing3=8)
        text.tag_configure("edit_selected", background="#E7F1FF")
        self._bind_widget_mousewheel(text, text)
        self.edit_text = text
        self.edit_canvas = text
        self.edit_body = text
        self._set_edit_cards_empty("待编辑订单会显示在这里，打勾的 JIT 才会进入一键流程。")

    def _on_card_canvas_mousewheel(self, event) -> str | None:
        status = self._active_status()
        if status == 1 and hasattr(self, "edit_canvas"):
            return self._scroll_mousewheel_target(self.edit_canvas, event)
        if status == 2 and hasattr(self, "payment_canvas"):
            return self._scroll_mousewheel_target(self.payment_canvas, event)
        return None

    def _show_order_context_menu(self, event, status: int) -> str:
        if status == 2:
            menu = tk.Menu(self.root, tearoff=False)
            menu.add_command(label="复制订单号", command=self.copy_selected_order_nos)
            menu.add_command(label="预览订单图片", command=self.preview_selected_payment_order_images)
            menu.tk_popup(event.x_root, event.y_root)
            return "break"

        tree = self.trees.get(status)
        if tree is None:
            return "break"
        iid = tree.identify_row(event.y)
        if iid and iid not in tree.selection():
            tree.selection_set(iid)
            tree.focus(iid)
            self._on_tree_select()
        menu = tk.Menu(self.root, tearoff=False)
        menu.add_command(label="复制订单号", command=self.copy_selected_order_nos)
        menu.tk_popup(event.x_root, event.y_root)
        return "break"

    def copy_selected_order_nos(self, _event=None) -> str:
        order_nos = [str(row.get("order_no") or "").strip() for row in self.get_selected_rows()]
        order_nos = [item for item in order_nos if item]
        if not order_nos:
            return "break"
        text = "\n".join(order_nos)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.status_var.set(f"已复制 {len(order_nos)} 个订单号")
        return "break"

    def _current_output_dir(self) -> Path:
        raw = self.output_dir_var.get().strip()
        return Path(raw) if raw else get_default_output_dir()

    def _find_downloaded_images_for_order(self, order_no: str) -> list[dict[str, Any]]:
        output_dir = self._current_output_dir()
        order_key = normalize_order_no(order_no)
        found: list[dict[str, Any]] = []
        seen_files: set[str] = set()

        manifest = output_dir / "api_download_manifest.json"
        if manifest.exists():
            try:
                payload = json.loads(manifest.read_text(encoding="utf-8"))
                records = payload.get("records") if isinstance(payload, dict) else payload
                if isinstance(records, list):
                    for record in records:
                        if not isinstance(record, dict):
                            continue
                        if normalize_order_no(record.get("order_no")) != order_key:
                            continue
                        file_path = Path(str(record.get("file") or ""))
                        if record.get("status") == "ok" and file_path.exists():
                            key = str(file_path.resolve())
                            if key not in seen_files:
                                found.append({"file": file_path, "sku": str(record.get("sku") or "")})
                                seen_files.add(key)
            except Exception as exc:  # noqa: BLE001
                self._log(f"读取图片清单失败：{manifest}", str(exc))

        if not found and output_dir.exists():
            prefix = sanitize_name(order_no) + "__"
            suffixes = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
            for file_path in sorted(output_dir.glob(prefix + "*")):
                if file_path.is_file() and file_path.suffix.lower() in suffixes:
                    found.append({"file": file_path, "sku": ""})
        return found

    def _normalize_image_url(self, image_url: Any) -> str:
        value = str(image_url or "").strip()
        if not value:
            return ""
        if value.startswith("//"):
            return f"https:{value}"
        return value

    def _resize_image_url(self, image_url: str, size: int = 1200) -> str:
        url = self._normalize_image_url(image_url)
        if not url:
            return ""
        resize_arg = f"image/resize,l_{size}/imageslim"
        if "x-image-process=" in url:
            return re.sub(r"x-image-process=[^&]+", f"x-image-process={resize_arg}", url)
        separator = "&" if "?" in url else "?"
        return f"{url}{separator}x-image-process={resize_arg}"

    def _server_images_from_row(self, row: dict[str, Any]) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        seen_urls: set[str] = set()
        order_no = str(row.get("order_no") or "")
        for detail in row.get("detail") or []:
            if not isinstance(detail, dict):
                continue
            image_url = self._normalize_image_url(detail.get("image") or detail.get("pic"))
            if not image_url or image_url in seen_urls:
                continue
            seen_urls.add(image_url)
            items.append(
                {
                    "order_no": order_no,
                    "sku": str(detail.get("sku") or detail.get("product_sku") or detail.get("goods_sku") or ""),
                    "url": image_url,
                    "thumb_url": self._resize_image_url(image_url, 600),
                    "high_url": self._resize_image_url(image_url, 1200),
                    "source": "server",
                }
            )
        return items

    def _image_items_for_row(self, row: dict[str, Any]) -> list[dict[str, Any]]:
        server_items = self._server_images_from_row(row)
        if server_items:
            return server_items
        order_no = str(row.get("order_no") or "")
        return [{**item, "source": "local"} for item in self._find_downloaded_images_for_order(order_no)]

    def _open_image(self, item: dict[str, Any], *, high_res: bool = False):
        from io import BytesIO
        from PIL import Image

        candidates = []
        if high_res:
            candidates.extend([item.get("high_url"), item.get("thumb_url"), item.get("url")])
        else:
            candidates.extend([item.get("thumb_url"), item.get("url")])
        for image_url in [str(item or "").strip() for item in candidates if str(item or "").strip()]:
            try:
                response = requests.get(image_url, timeout=12)
                response.raise_for_status()
                return Image.open(BytesIO(response.content))
            except Exception:
                continue
        return Image.open(Path(str(item.get("file") or "")))

    def _trim_light_border(self, image):
        from PIL import Image, ImageChops

        rgb = image.convert("RGB")
        background = Image.new("RGB", rgb.size, (255, 255, 255))
        diff = ImageChops.difference(rgb, background).convert("L")
        mask = diff.point(lambda value: 255 if value > 18 else 0)
        bbox = mask.getbbox()
        if not bbox:
            return image
        left, top, right, bottom = bbox
        pad = 18
        left = max(left - pad, 0)
        top = max(top - pad, 0)
        right = min(right + pad, image.width)
        bottom = min(bottom + pad, image.height)
        return image.crop((left, top, right, bottom))

    def _mousewheel_units(self, event) -> int:
        delta = int(getattr(event, "delta", 0) or 0)
        if delta:
            if abs(delta) >= 120:
                return int(-delta / 120)
            return -1 if delta > 0 else 1
        number = int(getattr(event, "num", 0) or 0)
        if number == 4:
            return -1
        if number == 5:
            return 1
        return 0

    def _scroll_target_pixels(self, target: tk.Widget, pixels: int) -> str | None:
        """按像素滚动；部件不支持 pixels 时按累计量回退到 units。"""
        if not pixels:
            return None
        try:
            target.yview_scroll(pixels, "pixels")
            return "break"
        except tk.TclError:
            pass
        key = str(target)
        pending = self._scroll_pixel_remainder.get(key, 0) + pixels
        units = int(pending / SCROLL_PIXELS_PER_UNIT)
        self._scroll_pixel_remainder[key] = pending - units * SCROLL_PIXELS_PER_UNIT
        if not units:
            return "break"
        try:
            target.yview_scroll(units, "units")
        except tk.TclError:
            return None
        return "break"

    def _scroll_mousewheel_target(self, target: tk.Widget, event) -> str | None:
        units = self._mousewheel_units(event)
        if not units:
            return None
        # Text 的一个 unit 是「显示行」，而每张订单卡片就是一整行；卡片比可视区高时整行滚不动。
        # Canvas 同理会按 yscrollincrement 跳格，所以这两类统一改成像素滚动。
        if isinstance(target, (tk.Text, tk.Canvas)):
            return self._scroll_target_pixels(target, units * SCROLL_PIXELS_PER_UNIT)
        try:
            target.yview_scroll(units, "units")
        except tk.TclError:
            return None
        return "break"

    @staticmethod
    def _touchpad_scroll_deltas(event) -> tuple[int, int]:
        """拆开 <TouchpadScroll> 的 %D，与 Tk 的 tk::PreciseScrollDeltas 一致。"""
        packed = int(getattr(event, "delta", 0) or 0)
        delta_x = packed >> 16
        low = packed & 0xFFFF
        delta_y = low if low < 0x8000 else low - 0x10000
        return delta_x, delta_y

    def _scroll_touchpad_target(self, target: tk.Widget, event) -> str | None:
        _delta_x, delta_y = self._touchpad_scroll_deltas(event)
        if not delta_y:
            return None
        return self._scroll_target_pixels(target, -delta_y)

    def _find_mousewheel_target(self, widget: tk.Widget | None) -> tk.Widget | None:
        current = widget
        while current is not None:
            target = self._mousewheel_targets.get(str(current))
            if target is not None:
                try:
                    if target.winfo_exists():
                        return target
                except tk.TclError:
                    self._mousewheel_targets.pop(str(current), None)
                    return None
            current = getattr(current, "master", None)
        return None

    def _event_in_widget(self, event, widget: tk.Widget) -> bool:
        try:
            x_root = int(getattr(event, "x_root", 0) or 0)
            y_root = int(getattr(event, "y_root", 0) or 0)
            left = widget.winfo_rootx()
            top = widget.winfo_rooty()
            return left <= x_root <= left + widget.winfo_width() and top <= y_root <= top + widget.winfo_height()
        except tk.TclError:
            return False

    def _active_canvas_under_event(self, event) -> tk.Widget | None:
        status = self._active_status()
        candidates: list[tk.Widget] = []
        if status == 1 and hasattr(self, "edit_canvas"):
            candidates.append(self.edit_canvas)
        if status == 2 and hasattr(self, "payment_canvas"):
            candidates.append(self.payment_canvas)
        if status == 3 and self.tree is not None:
            candidates.append(self.tree)
        for candidate in candidates:
            if self._event_in_widget(event, candidate):
                return candidate
        return None

    def _resolve_scroll_target(self, event) -> tk.Widget | None:
        target = None
        hover_target = getattr(self, "_mousewheel_hover_target", None)
        if hover_target is not None:
            try:
                if hover_target.winfo_exists():
                    target = hover_target
            except tk.TclError:
                self._mousewheel_hover_target = None
        if target is None:
            target = self._find_mousewheel_target(getattr(event, "widget", None))
        if target is None:
            target = self._active_canvas_under_event(event)
        return target

    def _dispatch_mousewheel(self, event) -> str | None:
        target = self._resolve_scroll_target(event)
        if target is None:
            return None
        return self._scroll_mousewheel_target(target, event)

    def _dispatch_touchpad_scroll(self, event) -> str | None:
        target = self._resolve_scroll_target(event)
        if target is None:
            return None
        return self._scroll_touchpad_target(target, event)

    def _bind_touchpad_scroll(self, widget: tk.Widget, handler) -> bool:
        """绑定 <TouchpadScroll>。Tk 8.6 没有这个事件，绑定失败就静默跳过。"""
        if self._touchpad_scroll_supported is False:
            return False
        try:
            widget.bind("<TouchpadScroll>", handler)
        except tk.TclError:
            self._touchpad_scroll_supported = False
            return False
        self._touchpad_scroll_supported = True
        return True

    def _ensure_mousewheel_dispatcher(self) -> None:
        if self._mousewheel_dispatcher_installed:
            return
        self.root.bind_all("<MouseWheel>", self._dispatch_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._dispatch_mousewheel, add="+")
        self.root.bind_all("<Button-5>", self._dispatch_mousewheel, add="+")
        if self._touchpad_scroll_supported is not False:
            try:
                self.root.bind_all("<TouchpadScroll>", self._dispatch_touchpad_scroll, add="+")
                self._touchpad_scroll_supported = True
            except tk.TclError:
                self._touchpad_scroll_supported = False
        self._mousewheel_dispatcher_installed = True

    def _forget_mousewheel_widget(self, widget_key: str) -> None:
        self._mousewheel_targets.pop(widget_key, None)
        self._scroll_pixel_remainder.pop(widget_key, None)

    def _bind_widget_mousewheel(self, widget: tk.Widget, canvas: tk.Canvas) -> None:
        widget_key = str(widget)
        self._mousewheel_targets[widget_key] = canvas
        self._ensure_mousewheel_dispatcher()
        def _handler(event, target=canvas):
            result = self._scroll_mousewheel_target(target, event)
            # macOS: 阻止事件继续传播，强制处理
            if result == "break":
                return "break"
            return result

        def _touchpad_handler(event, target=canvas):
            # Tk 9 起，macOS 的触控板/妙控鼠标发的是 <TouchpadScroll> 而不是 <MouseWheel>。
            # Text、Treeview 等自带类绑定还能滚，但卡片里的 Frame/Label（尤其是图片 Label）
            # 没有任何类绑定，不显式绑定就完全滚不动。
            return self._scroll_touchpad_target(target, event)

        def _focus_scroll_target(_event=None, target=canvas) -> None:
            try:
                self._mousewheel_hover_target = target
            except tk.TclError:
                pass

        def _blur_scroll_target(_event=None, target=canvas) -> None:
            if self._mousewheel_hover_target is target:
                self._mousewheel_hover_target = None

        # macOS 需要直接绑定到 widget，优先级更高
        widget.bind("<Enter>", _focus_scroll_target, add="+")
        widget.bind("<Leave>", _blur_scroll_target, add="+")
        widget.bind("<MouseWheel>", _handler)  # 移除 add="+" 让绑定优先级更高
        widget.bind("<Button-4>", _handler)
        widget.bind("<Button-5>", _handler)
        self._bind_touchpad_scroll(widget, _touchpad_handler)
        widget.bind("<Destroy>", lambda _event, key=widget_key: self._forget_mousewheel_widget(key), add="+")

    def _prepare_image_for_label(
        self,
        item: dict[str, Any],
        max_size: tuple[int, int],
        *,
        trim: bool = False,
        high_res: bool = False,
    ):
        from PIL import Image

        image = self._open_image(item, high_res=high_res)
        if trim:
            image = self._trim_light_border(image)
        image.thumbnail(max_size, Image.Resampling.LANCZOS)
        return image.copy()

    def _load_image_async(
        self,
        item: dict[str, Any],
        label: tk.Widget,
        max_size: tuple[int, int],
        owner_refs: list[Any],
        *,
        trim: bool = False,
        high_res: bool = False,
    ) -> None:
        def done(future) -> None:
            image = None
            error = None
            try:
                image = future.result()
            except Exception as exc:  # noqa: BLE001
                error = f"图片打开失败：{exc}"
            self.image_queue.put((label, image, owner_refs, error))

        future = self.image_executor.submit(
            self._prepare_image_for_label,
            item,
            max_size,
            trim=trim,
            high_res=high_res,
        )
        future.add_done_callback(done)

    def preview_selected_payment_order_images(self) -> str:
        if self._active_status() != 2:
            messagebox.showinfo("预览图片", "请先切换到待付款标签页，并选中订单。")
            return "break"

        rows = self.get_selected_rows()
        if not rows:
            messagebox.showinfo("预览图片", "请先选中一个待付款订单。")
            return "break"

        image_items: list[dict[str, Any]] = []
        missing_order_nos: list[str] = []
        for row in rows:
            order_no = str(row.get("order_no") or "").strip()
            if not order_no:
                continue
            items = self._image_items_for_row(row)
            if not items:
                missing_order_nos.append(order_no)
                continue
            for item in items:
                image_items.append({"order_no": order_no, **item})

        if not image_items:
            messagebox.showinfo(
                "预览图片",
                "\n".join(
                    [
                        "未找到这些订单的服务器图片或已下载图片。",
                        f"订单：{self._format_order_no_preview([str(row.get('order_no') or '') for row in rows])}",
                        f"目录：{self._current_output_dir()}",
                        "",
                        "如果接口没有返回图链，请先执行“下载图片”。",
                    ]
                ),
            )
            return "break"

        self._show_image_preview_window(image_items, missing_order_nos)
        return "break"

    def _show_image_preview_window(self, image_items: list[dict[str, Any]], missing_order_nos: list[str]) -> None:
        order_nos = []
        for item in image_items:
            order_no = str(item.get("order_no") or "")
            if order_no and order_no not in order_nos:
                order_nos.append(order_no)

        if self.preview_window is not None and self.preview_window.winfo_exists():
            self.preview_window.destroy()

        window = tk.Toplevel(self.root)
        self.preview_window = window
        window.title(f"订单图片预览 - {len(order_nos)} 单")
        window.geometry("940x720")
        window.configure(background="#F8F9FA")
        window.preview_images = []
        window.protocol("WM_DELETE_WINDOW", lambda: self._close_preview_window(window))

        toolbar = ttk.Frame(window)
        toolbar.pack(fill="x", padx=8, pady=6)
        source_text = "优先服务器高清图；无图链时使用本地已下载图片"
        ttk.Label(toolbar, text=source_text, background="#F8F9FA").pack(side="left")
        ttk.Button(toolbar, text="打开目录", command=lambda: open_path(self._current_output_dir())).pack(side="right")

        if missing_order_nos:
            ttk.Label(
                window,
                text=f"未找到图片：{self._format_order_no_preview(missing_order_nos)}",
                foreground="#B02A37",
                background="#F8F9FA",
            ).pack(fill="x", padx=8, pady=(0, 4))

        canvas = tk.Canvas(window, bg="#F8F9FA", highlightthickness=0)
        scrollbar = ttk.Scrollbar(window, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        body_window = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(body_window, width=event.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        scrollbar.pack(side="right", fill="y", padx=(0, 8), pady=(0, 8))
        self._bind_widget_mousewheel(canvas, canvas)
        self._bind_widget_mousewheel(body, canvas)

        max_width = 860
        max_height = 520
        for item in image_items:
            header = f"{item.get('order_no') or '-'}"
            if item.get("sku"):
                header += f"  SKU：{item['sku']}"
            header += "  服务器图" if item.get("url") else f"  {Path(str(item.get('file') or '')).name}"
            row_header = ttk.Frame(body)
            row_header.pack(fill="x", padx=8, pady=(8, 3))
            self._bind_widget_mousewheel(row_header, canvas)
            header_label = ttk.Label(row_header, text=header, font=("Microsoft YaHei UI", 9, "bold"), background="#F8F9FA")
            header_label.pack(side="left")
            self._bind_widget_mousewheel(header_label, canvas)
            high_url = str(item.get("high_url") or item.get("url") or "")
            if high_url:
                high_btn = ttk.Button(row_header, text="查看高清", command=lambda url=high_url: open_path(url))
                high_btn.pack(side="right")
                self._bind_widget_mousewheel(high_btn, canvas)
            image_label = ttk.Label(body, text="图片加载中...", background="#F8F9FA", foreground="#6C757D")
            image_label.pack(anchor="w", padx=8, pady=(0, 8))
            self._bind_widget_mousewheel(image_label, canvas)
            self._load_image_async(item, image_label, (max_width, max_height), window.preview_images, high_res=True)

    def _close_preview_window(self, window: tk.Toplevel | None = None) -> None:
        target = window or self.preview_window
        if target is not None:
            try:
                target.destroy()
            except Exception:
                pass
        if target is self.preview_window:
            self.preview_window = None

    def _choose_output_dir(self) -> None:
        chosen = filedialog.askdirectory()
        if chosen:
            self.output_dir_var.set(chosen)

    def _choose_auth_file(self) -> None:
        chosen = filedialog.askopenfilename(
            title="选择 auth-state-v1.json",
            initialdir=str(Path(DEFAULT_AUTH_STATE_FILE).parent),
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
        )
        if chosen:
            self.auth_file_var.set(chosen)

    def _start_auth_sync_receiver(self) -> None:
        ports = self.auth_sync_receiver.start()
        if ports:
            text = "登录态接收服务：已启动 " + " / ".join(str(port) for port in ports)
            self.auth_sync_status_var.set(text)
            self._log(text)
            if self.auth_sync_stop_after_id:
                try:
                    self.root.after_cancel(self.auth_sync_stop_after_id)
                except Exception:
                    pass
            self.auth_sync_stop_after_id = self.root.after(AUTH_SYNC_LISTEN_SECONDS * 1000, self._stop_auth_sync_receiver_timeout)
        else:
            text = "登录态接收服务：18321/18888 均被占用，已跳过"
            self.auth_sync_status_var.set(text)
            self._log(text)

    def _stop_auth_sync_receiver_timeout(self) -> None:
        self.auth_sync_stop_after_id = None
        self._stop_auth_sync_receiver("登录态接收：已超时关闭")

    def _stop_auth_sync_receiver(self, message: str = "登录态接收：已关闭") -> None:
        if self.auth_sync_stop_after_id:
            try:
                self.root.after_cancel(self.auth_sync_stop_after_id)
            except Exception:
                pass
            self.auth_sync_stop_after_id = None
        try:
            self.auth_sync_receiver.stop()
        except Exception:
            pass
        self.auth_sync_status_var.set(message)
        self._log(message)

    def _on_auth_synced(self, auth: dict[str, Any], auth_file: Path, port: int) -> None:
        username = auth.get("username") or auth.get("companyName") or "-"
        message = f"登录态已同步：{username}，端口 {port}，文件 {auth_file}；接收服务已关闭"
        try:
            self.root.after(0, lambda: self._stop_auth_sync_receiver(message))
        except Exception:
            pass

    def show_auth_help(self) -> None:
        messagebox.showinfo(
            "登录态说明",
            "\n".join(
                [
                    "macOS 版本仅使用本地同步文件，不读取浏览器登录态：",
                    "1. 点击“接收登录态3分钟”启动本机接收服务。",
                    "2. 点击“一键复制同步脚本”，粘贴到 Tampermonkey 后保存并启用。",
                    "",
                    "如何生成同步文件：",
                    "1. 在浏览器打开 https://user.landwu.com/ 并登录。",
                    "2. 页面右下角提示“桥接登录态：已同步”后，就会写入当前同步文件。",
                    "",
                    "注意：",
                    "auth-state-v1.json 等同登录凭证，不要发给无关人员。",
                    "如果刷新订单提示登录失效，重新打开 Landwu 页面同步一次即可。",
                ]
            ),
        )

    def copy_auth_sync_userscript(self) -> None:
        self.root.clipboard_clear()
        self.root.clipboard_append(LANDWU_AUTH_SYNC_USERSCRIPT)
        self.root.update_idletasks()
        self.status_var.set("登录态同步 userscript 已复制到剪贴板")
        self._log("已复制内置 Landwu 登录态同步 userscript。")
        messagebox.showinfo(
            "已复制同步脚本",
            "请在 Tampermonkey 新建脚本，粘贴后保存并启用。\n\n"
            "随后点击“接收登录态3分钟”，再打开或刷新 Landwu 页面即可同步。",
        )

    def check_for_update(self) -> None:
        def task() -> dict[str, Any]:
            local_source = Path(__file__).read_text(encoding="utf-8")
            remote = fetch_remote_update_source()
            available = normalize_source_text(local_source) != normalize_source_text(remote["source"])
            self.pending_update_source = remote["source"] if available else ""
            return {
                "available": available,
                "localVersion": APP_VERSION,
                "remoteVersion": remote["version"],
                "sha": remote["sha"],
            }

        def on_success(payload: dict[str, Any]) -> None:
            if not payload.get("available"):
                messagebox.showinfo("检查更新", f"当前已是最新版本（{APP_VERSION}）。")
                return

            remote_version = str(payload.get("remoteVersion") or "未知版本")
            confirmed = messagebox.askyesno(
                "发现新版本",
                f"当前版本：{APP_VERSION}\n最新版本：{remote_version}\n\n现在更新并重启吗？",
            )
            if not confirmed:
                self.status_var.set("已取消更新")
                return

            source = self.pending_update_source
            if not source:
                messagebox.showerror("安装更新", "更新内容已丢失，请重新检查更新。")
                return

            def install_task() -> dict[str, str]:
                target_path = apply_remote_update_source(source)
                return {"path": str(target_path), "version": remote_version}

            def restart_app(result: dict[str, str]) -> None:
                target_path = Path(result["path"])
                try:
                    subprocess.Popen(
                        [sys.executable, str(target_path)],
                        cwd=str(target_path.parent),
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        stdin=subprocess.DEVNULL,
                    )
                except Exception as exc:  # noqa: BLE001
                    messagebox.showerror("更新已安装", f"新版本已写入，但自动重启失败：{exc}")
                    return
                self.root.after(150, self._on_close)

            self._run_task("安装更新", install_task, on_success=restart_app)

        self._run_task("检查更新", task, on_success=on_success)

    def open_settings_window(self) -> None:
        if self.settings_window is not None:
            try:
                if self.settings_window.winfo_exists():
                    self.settings_window.lift()
                    self.settings_window.focus_force()
                    return
            except Exception:
                self.settings_window = None

        window = tk.Toplevel(self.root)
        self.settings_window = window
        window.title("设置")
        window.geometry("820x450")
        window.minsize(760, 405)
        window.configure(background="#EEF2F6")
        window.transient(self.root)

        def close_window() -> None:
            self.settings_window = None
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", close_window)

        form = ttk.Frame(window, style="Surface.TFrame", padding=14)
        form.pack(fill="both", expand=True, padx=10, pady=10)
        form.columnconfigure(1, weight=1)

        ttk.Label(form, text="图片目录").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.output_dir_var).grid(row=0, column=1, columnspan=3, sticky="ew", padx=(8, 0), pady=4)
        ttk.Button(form, text="选择目录", command=self._choose_output_dir).grid(row=0, column=4, columnspan=2, sticky="w", padx=(12, 0), pady=4)

        ttk.Label(form, text="成分数据库").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.composition_db_folder_var).grid(row=1, column=1, columnspan=3, sticky="ew", padx=(8, 0), pady=4)
        ttk.Button(form, text="选择目录", command=self._choose_composition_db_folder).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=(12, 0), pady=4
        )

        ttk.Label(form, text="登录态来源").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Label(form, text="本地同步文件", background="#FFFFFF", foreground="#495057").grid(row=2, column=1, sticky="w", padx=(8, 0), pady=4)
        ttk.Button(form, text="登录态说明", command=self.show_auth_help).grid(row=2, column=2, sticky="w", padx=(8, 0), pady=4)
        ttk.Button(form, text="查看登录态", command=self.fetch_auth).grid(row=2, column=3, sticky="w", padx=(8, 0), pady=4)
        ttk.Label(form, text="下单余额").grid(row=2, column=4, sticky="e", padx=(12, 4), pady=4)
        ttk.Label(form, textvariable=self.balance_var, style="SettingsBalance.TLabel").grid(row=2, column=5, sticky="w", pady=4)

        ttk.Label(form, text="同步文件").grid(row=3, column=0, sticky="w", pady=4)
        ttk.Entry(form, textvariable=self.auth_file_var).grid(row=3, column=1, columnspan=2, sticky="ew", padx=(8, 0), pady=4)
        ttk.Button(form, text="选择文件", command=self._choose_auth_file).grid(row=3, column=3, sticky="w", padx=(8, 0), pady=4)
        ttk.Button(form, text="接收登录态3分钟", command=self._start_auth_sync_receiver).grid(
            row=3, column=4, columnspan=2, sticky="w", padx=(8, 0), pady=4
        )
        ttk.Label(form, text="同步脚本").grid(row=4, column=0, sticky="w", pady=4)
        ttk.Button(form, text="一键复制同步脚本", command=self.copy_auth_sync_userscript).grid(
            row=4, column=1, columnspan=2, sticky="w", padx=(8, 0), pady=4
        )
        ttk.Label(form, text="复制后粘贴到 Tampermonkey 并启用。", background="#FFFFFF", foreground="#6C757D").grid(
            row=4, column=3, columnspan=3, sticky="w", padx=(8, 0), pady=4
        )
        ttk.Label(form, textvariable=self.auth_sync_status_var, background="#FFFFFF", foreground="#6C757D").grid(
            row=5, column=1, columnspan=5, sticky="w", padx=(8, 0), pady=(0, 4)
        )

        ttk.Separator(form).grid(row=6, column=0, columnspan=6, sticky="ew", pady=12)
        ttk.Checkbutton(
            form,
            text="开启自动刷新",
            variable=self.auto_refresh_enabled_var,
            command=self.apply_auto_refresh_settings,
        ).grid(row=7, column=0, sticky="w", pady=4)
        ttk.Label(form, text="模式").grid(row=7, column=1, sticky="e", pady=4)
        ttk.Combobox(
            form,
            textvariable=self.auto_refresh_mode_var,
            values=("每隔分钟", "每天时间"),
            width=10,
            state="readonly",
        ).grid(row=7, column=2, sticky="w", padx=(8, 0), pady=4)
        ttk.Label(form, text="间隔分钟").grid(row=7, column=3, sticky="e", padx=(10, 0), pady=4)
        ttk.Entry(form, textvariable=self.auto_refresh_interval_var, width=8).grid(row=7, column=4, sticky="w", padx=(8, 0), pady=4)
        ttk.Label(form, text="每天 HH:MM").grid(row=8, column=3, sticky="e", padx=(10, 0), pady=4)
        ttk.Entry(form, textvariable=self.auto_refresh_time_var, width=8).grid(row=8, column=4, sticky="w", padx=(8, 0), pady=4)
        ttk.Button(form, text="应用定时", command=self.apply_auto_refresh_settings).grid(row=8, column=5, sticky="e", pady=4)
        ttk.Label(form, textvariable=self.auto_refresh_status_var, background="#FFFFFF", foreground="#6C757D").grid(
            row=9, column=0, columnspan=6, sticky="w", pady=(10, 0)
        )

        footer = ttk.Frame(form, style="Surface.TFrame")
        footer.grid(row=10, column=0, columnspan=6, sticky="e", pady=(14, 0))
        ttk.Label(footer, text=f"当前版本：{APP_VERSION}", background="#FFFFFF", foreground="#6C757D").pack(side="left", padx=(0, 10))
        ttk.Button(footer, text="检查更新", command=self.check_for_update, style="Ghost.TButton").pack(side="left", padx=(0, 8))
        ttk.Button(footer, text="关闭", command=close_window).pack(side="right")

    def open_log_window(self) -> None:
        if self.log_window is not None:
            try:
                if self.log_window.winfo_exists():
                    self.log_window.lift()
                    self.log_window.focus_force()
                    return
            except Exception:
                self.log_window = None
                self.log_text = None

        window = tk.Toplevel(self.root)
        self.log_window = window
        window.title("运行日志")
        window.geometry("860x460")
        window.minsize(720, 320)
        window.configure(background="#EEF2F6")
        window.transient(self.root)

        def close_window() -> None:
            self.log_window = None
            self.log_text = None
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", close_window)
        text = ScrolledText(
            window,
            wrap="word",
            font=("Consolas", 9),
            bg="#17212B",
            fg="#E6EDF3",
            insertbackground="white",
            borderwidth=0,
        )
        text.pack(fill="both", expand=True, padx=10, pady=10)
        text.configure(state="disabled")
        self.log_text = text
        self._populate_log_window()

    def _populate_log_window(self) -> None:
        if self.log_text is None:
            return
        try:
            if not self.log_text.winfo_exists():
                return
            self.log_text.configure(state="normal")
            self.log_text.delete("1.0", "end")
            self.log_text.insert("end", "\n".join(self.log_lines))
            if self.log_lines:
                self.log_text.insert("end", "\n")
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        except Exception:
            self.log_text = None

    def _log(self, message: str, payload: Any | None = None) -> None:
        stamp = time.strftime("%H:%M:%S")
        chunks = [f"[{stamp}] {message}"]
        if payload is not None:
            if isinstance(payload, str):
                chunks.append(payload)
            else:
                chunks.append(json.dumps(payload, ensure_ascii=False, indent=2))
        text = "\n".join(chunks)
        self.log_lines.append(text)
        if len(self.log_lines) > 800:
            self.log_lines = self.log_lines[-800:]
        if self.log_text is not None:
            try:
                if self.log_text.winfo_exists():
                    self.log_text.configure(state="normal")
                    self.log_text.insert("end", text + "\n")
                    self.log_text.see("end")
                    self.log_text.configure(state="disabled")
            except Exception:
                self.log_text = None

    def _set_busy(self, busy: bool, status: str) -> None:
        self.busy = busy
        self.status_var.set(status)
        state = "disabled" if busy else "normal"
        for button in self.buttons:
            try:
                button.configure(state=state)
            except Exception:
                pass
        try:
            self.root.configure(cursor="watch" if busy else "")
        except Exception:
            pass

    def _make_runtime_args(self, *, allow_browser_fallback: bool = True) -> argparse.Namespace:
        return build_runtime_args(
            auth_source=AUTH_SOURCE_FILE,
            auth_file=self.auth_file_var.get().strip() or DEFAULT_AUTH_STATE_FILE,
        )

    def _format_ts_for_auto_refresh(self, timestamp: float | None) -> str:
        if not timestamp:
            return "-"
        return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(timestamp))

    def _compute_next_auto_refresh_ts(self) -> float:
        mode = self.auto_refresh_mode_var.get().strip()
        now = time.time()
        if mode == "每天时间":
            raw_time = self.auto_refresh_time_var.get().strip()
            match = re.fullmatch(r"(\d{1,2}):(\d{2})", raw_time)
            if not match:
                raise ValueError("每天时间格式应为 HH:MM，例如 09:30")
            hour = int(match.group(1))
            minute = int(match.group(2))
            if hour > 23 or minute > 59:
                raise ValueError("每天时间超出范围，应为 00:00 到 23:59")
            parts = list(time.localtime(now))
            parts[3] = hour
            parts[4] = minute
            parts[5] = 0
            target = time.mktime(tuple(parts))
            if target <= now:
                target += 24 * 60 * 60
            return target

        raw_interval = self.auto_refresh_interval_var.get().strip()
        interval_minutes = float(raw_interval)
        if interval_minutes <= 0:
            raise ValueError("间隔分钟必须大于 0")
        return now + interval_minutes * 60

    def _update_auto_refresh_status(self) -> None:
        if not self.auto_refresh_enabled_var.get():
            self.auto_refresh_status_var.set("自动刷新关闭")
            return
        mode = self.auto_refresh_mode_var.get().strip()
        self.auto_refresh_status_var.set(f"已开启：{mode}，下次 {self._format_ts_for_auto_refresh(self.auto_refresh_next_ts)}")

    def _schedule_auto_refresh_tick(self) -> None:
        if not self.auto_refresh_enabled_var.get() or not self.auto_refresh_next_ts:
            return
        delay_ms = int(max(1000, min(30000, (self.auto_refresh_next_ts - time.time()) * 1000)))
        self.auto_refresh_after_id = self.root.after(delay_ms, self._auto_refresh_tick)

    def apply_auto_refresh_settings(self) -> None:
        if self.auto_refresh_after_id:
            try:
                self.root.after_cancel(self.auto_refresh_after_id)
            except Exception:
                pass
            self.auto_refresh_after_id = None

        if not self.auto_refresh_enabled_var.get():
            self.auto_refresh_next_ts = None
            self._update_auto_refresh_status()
            return

        try:
            self.auto_refresh_next_ts = self._compute_next_auto_refresh_ts()
        except Exception as exc:  # noqa: BLE001
            self.auto_refresh_enabled_var.set(False)
            self.auto_refresh_next_ts = None
            self._update_auto_refresh_status()
            messagebox.showerror("自动刷新设置", str(exc))
            return

        self._update_auto_refresh_status()
        self._schedule_auto_refresh_tick()
        self._log(self.auto_refresh_status_var.get())

    def _auto_refresh_tick(self) -> None:
        self.auto_refresh_after_id = None
        if not self.auto_refresh_enabled_var.get():
            self.auto_refresh_next_ts = None
            self._update_auto_refresh_status()
            return

        if self.auto_refresh_next_ts and time.time() >= self.auto_refresh_next_ts:
            if self.busy:
                self._log("自动刷新跳过：当前正在执行任务")
            else:
                self.refresh_summary(auto=True)
            try:
                self.auto_refresh_next_ts = self._compute_next_auto_refresh_ts()
            except Exception as exc:  # noqa: BLE001
                self.auto_refresh_enabled_var.set(False)
                self.auto_refresh_next_ts = None
                self._update_auto_refresh_status()
                self._log("自动刷新已关闭：设置无效", str(exc))
                return
            self._update_auto_refresh_status()

        self._schedule_auto_refresh_tick()

    def _run_task(
        self,
        title: str,
        task,
        on_success=None,
        *,
        quiet_if_busy: bool = False,
        show_error: bool = True,
        on_error=None,
    ) -> None:
        if self.busy:
            if quiet_if_busy:
                self._log(f"跳过：{title}，当前已有任务执行中")
                return
            messagebox.showinfo("请稍候", "当前已有任务执行中。")
            return

        self.current_success_handler = on_success
        self.current_error_handler = on_error
        self.current_show_error_popup = show_error
        self._set_busy(True, f"{title}中...")
        self._log(f"开始：{title}")

        def worker() -> None:
            try:
                result = task()
                self.queue.put(("success", title, result))
            except Exception as exc:  # noqa: BLE001
                self.queue.put(("error", title, str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _process_queue(self) -> None:
        try:
            while True:
                kind, title, payload = self.queue.get_nowait()
                if kind == "success":
                    self._set_busy(False, f"{title}完成")
                    self._log(f"完成：{title}", payload)
                    handler = self.current_success_handler
                    self.current_success_handler = None
                    self.current_error_handler = None
                    self.current_show_error_popup = True
                    if handler:
                        handler(payload)
                else:
                    brief = str(payload).replace("\r", " ").replace("\n", " ").strip()
                    if len(brief) > 80:
                        brief = brief[:77] + "..."
                    self._set_busy(False, f"{title}失败：{brief}" if brief else f"{title}失败")
                    show_error = self.current_show_error_popup
                    handler = self.current_error_handler
                    self.current_success_handler = None
                    self.current_error_handler = None
                    self.current_show_error_popup = True
                    self._log(f"失败：{title}", payload)
                    if handler:
                        handler(payload)
                    if show_error:
                        messagebox.showerror(title, str(payload))
        except Empty:
            pass
        finally:
            self.root.after(150, self._process_queue)

    def _process_image_queue(self) -> None:
        try:
            from PIL import ImageTk

            while True:
                label, image, owner_refs, error = self.image_queue.get_nowait()
                if isinstance(label, tuple) and label and label[0] == "text_image":
                    _kind, text, mark_name, placeholder_tag, generation = label
                    try:
                        if not text.winfo_exists():
                            continue
                        if generation != getattr(self, "payment_render_generation", None):
                            continue
                        previous_state = str(text.cget("state"))
                        text.configure(state="normal")
                        ranges = text.tag_ranges(placeholder_tag)
                        if ranges:
                            text.delete(ranges[0], ranges[1])
                        if error:
                            text.insert(mark_name, error + "\n\n", ("payment_error",))
                        elif image is not None:
                            photo = ImageTk.PhotoImage(image)
                            owner_refs.append(photo)
                            text.image_create(mark_name, image=photo)
                            text.insert(f"{mark_name}+1c", "\n\n")
                        text.configure(state=previous_state)
                    except Exception:
                        try:
                            text.configure(state="disabled")
                        except Exception:
                            pass
                    continue
                try:
                    if not label.winfo_exists():
                        continue
                    if error:
                        label.configure(text=error, image="")
                        continue
                    photo = ImageTk.PhotoImage(image)
                    owner_refs.append(photo)
                    label.configure(image=photo, text="")
                except Exception as exc:  # noqa: BLE001
                    try:
                        if label.winfo_exists():
                            label.configure(text=f"图片显示失败：{exc}", image="")
                    except Exception:
                        pass
        except Empty:
            pass
        finally:
            try:
                self.root.after(100, self._process_image_queue)
            except Exception:
                pass

    def _on_close(self) -> None:
        if self.auto_refresh_after_id:
            try:
                self.root.after_cancel(self.auto_refresh_after_id)
            except Exception:
                pass
            self.auto_refresh_after_id = None
        self._stop_auth_sync_receiver("登录态接收：已关闭")
        try:
            self.image_executor.shutdown(wait=False, cancel_futures=True)
        except TypeError:
            self.image_executor.shutdown(wait=False)
        except Exception:
            pass
        self.root.destroy()

    def _apply_summary(self, payload: dict[str, Any]) -> None:
        account = payload.get("account") or {}
        counts = payload.get("counts") or {}
        username = account.get("username") or "-"
        nickname = account.get("nickname") or ""
        account_text = username if not nickname else f"{username} / {nickname}"
        balance_text = format_money(account.get("balance"))
        self.account_var.set(account_text)
        self.factory_var.set(str(account.get("factoryId") or "-"))
        self.balance_var.set(f"¥ {balance_text}" if balance_text != "-" else "-")
        self._check_low_balance(account.get("balance"), balance_text)
        self.wait_edit_var.set(str(counts.get("waitEdit") or 0))
        self.wait_pay_var.set(str(counts.get("waitPayJit") or 0))
        self.paid_var.set(str(counts.get("paid") or 0))
        self.jit_var.set(str(counts.get("jit") or 0))
        self.vmi_var.set(str(counts.get("vmi") or 0))
        waybill_failed_count = int(counts.get("waybillFailed") or 0)
        self.waybill_failed_var.set(str(waybill_failed_count))
        self._apply_waybill_failed_style(waybill_failed_count)
        self.summary_line_var.set(
            "    ".join(
                [
                    f"账号：{account_text}",
                    f"待编辑：{counts.get('waitEdit') or 0}",
                    f"JIT：{counts.get('jit') or 0}",
                    f"VMI：{counts.get('vmi') or 0}",
                    f"待付款 JIT：{counts.get('waitPayJit') or 0}",
                    f"已支付：{counts.get('paid') or 0}",
                ]
            )
        )
        self.other_var.set(str(counts.get("other") or 0))

    def _check_low_balance(self, raw_balance: Any, balance_text: str) -> None:
        amount = parse_money_amount(raw_balance)
        if amount is None:
            return
        if amount >= LOW_BALANCE_ALERT_THRESHOLD:
            self.low_balance_alerted = False
            return
        message = f"下单余额低于 {LOW_BALANCE_ALERT_THRESHOLD:.0f} 元：当前 ¥ {balance_text}"
        self.status_var.set(message)
        if self.low_balance_alerted:
            return
        self.low_balance_alerted = True
        self._log(message)
        messagebox.showwarning("下单余额不足提醒", message)

    def _apply_waybill_failed_style(self, count: int) -> None:
        failed = count > 0
        bg = "#FEF2F2" if failed else "#F8FAFC"
        border = "#FCA5A5" if failed else "#CBD5E1"
        title_style = "WaybillFailTitle.TLabel" if failed else "WaybillOkTitle.TLabel"
        value_style = "WaybillFailValue.TLabel" if failed else "WaybillOkValue.TLabel"
        if self.waybill_failed_card is not None:
            self.waybill_failed_card.configure(bg=bg, highlightbackground=border)
        if self.waybill_failed_title_label is not None:
            self.waybill_failed_title_label.configure(style=title_style)
        if self.waybill_failed_value_label is not None:
            self.waybill_failed_value_label.configure(style=value_style)

    def _populate_order_table(self, rows: list[dict[str, Any]]) -> None:
        rows_by_status: dict[str, list[dict[str, Any]]] = {str(status): [] for status in GUI_STATUS_TABS}
        for row in rows:
            rows_by_status.setdefault(str(row.get("status") or ""), []).append(row)
        self._populate_order_tabs(rows_by_status)

    def _populate_order_tabs(self, rows_by_status: dict[str, list[dict[str, Any]]]) -> None:
        for tab_index, status in enumerate(GUI_STATUS_TABS):
            rows = rows_by_status.get(str(status)) or []
            if status == 1:
                self._populate_edit_cards(rows)
            elif status == 2:
                self._populate_payment_cards(rows)
            else:
                tree = self.trees[status]
                tree.delete(*tree.get_children())
                self.order_rows_by_status_iid[status].clear()
                tree.tag_configure("even", background="#FFFFFF")
                tree.tag_configure("odd", background="#F8F9FA")
                for index, row in enumerate(rows, start=1):
                    iid = f"{status}-row-{index}"
                    values = (
                        row.get("order_no") or "-",
                        row.get("status_text") or "-",
                        row.get("tag_name") or "-",
                        row.get("shop_name") or "-",
                        row.get("buy_number_count") or 0,
                        row.get("buy_type_count") or 0,
                        row.get("plattime") or "-",
                    )
                    tag = "even" if index % 2 == 0 else "odd"
                    tree.insert("", "end", iid=iid, text="", values=values, tags=(tag,))
                    self.order_rows_by_status_iid[status][iid] = row
            self.notebook.tab(tab_index, text=f"{STATUS_TEXT_MAP.get(status, status)} ({len(rows)})")
        self._on_tree_select()

    def _is_jit_row(self, row: dict[str, Any]) -> bool:
        return str(row.get("tag_name") or "").strip().upper() == "JIT"

    def _set_edit_cards_empty(self, message: str) -> None:
        self.edit_card_frames.clear()
        self.edit_check_vars.clear()
        self.selected_edit_iid = ""
        text = getattr(self, "edit_text", None)
        if text is not None:
            text.configure(state="normal")
            text.delete("1.0", "end")
            text.insert("end", message + "\n", ("edit_meta",))
            text.configure(state="disabled")

    def _toggle_edit_card(self, order_id: str) -> None:
        if not order_id:
            return
        if order_id in self.unchecked_edit_order_ids:
            self.unchecked_edit_order_ids.remove(order_id)
        else:
            self.unchecked_edit_order_ids.add(order_id)
        self._refresh_edit_marks()
        self._on_tree_select()

    def _sync_edit_check_var(self, order_id: str, var: tk.BooleanVar) -> None:
        expected = order_id not in self.unchecked_edit_order_ids
        if var.get() != expected:
            var.set(expected)

    def _select_edit_card(self, iid: str) -> None:
        if not iid:
            return
        self.selected_edit_iid = iid
        self._refresh_edit_card_selection()
        self._on_tree_select()

    def _refresh_edit_card_selection(self) -> None:
        text = getattr(self, "edit_text", None)
        if text is None:
            return
        text.configure(state="normal")
        text.tag_remove("edit_selected", "1.0", "end")
        if self.selected_edit_iid:
            tag_name = f"edit_row_{self.selected_edit_iid}"
            ranges = text.tag_ranges(tag_name)
            for start, end in zip(ranges[0::2], ranges[1::2]):
                text.tag_add("edit_selected", start, end)
                text.see(start)
        text.configure(state="disabled")

    def _bind_edit_card_widget(self, widget: tk.Widget, iid: str) -> None:
        widget.bind("<Button-1>", lambda _event, row_iid=iid: self._select_edit_card(row_iid))
        widget.bind("<Button-3>", lambda event, row_iid=iid: self._show_edit_card_context_menu(event, row_iid))
        if hasattr(self, "edit_canvas"):
            self._bind_widget_mousewheel(widget, self.edit_canvas)

    def _show_edit_card_context_menu(self, event, iid: str) -> str:
        self._select_edit_card(iid)
        menu = tk.Menu(self.root, tearoff=False)
        menu.add_command(label="复制订单号", command=self.copy_selected_order_nos)
        menu.tk_popup(event.x_root, event.y_root)
        return "break"

    def _populate_edit_cards(self, rows: list[dict[str, Any]]) -> None:
        self.edit_card_frames.clear()
        self.edit_check_vars.clear()
        self.order_rows_by_status_iid[1].clear()
        text = getattr(self, "edit_text", None)
        if text is not None:
            text.configure(state="normal")
            text.delete("1.0", "end")
        self.selected_edit_iid = ""

        if not rows:
            self._set_edit_cards_empty("当前没有待编辑订单。")
            return

        for index, row in enumerate(rows, start=1):
            iid = f"1-row-{index}"
            row_tag = f"edit_row_{iid}"
            self.order_rows_by_status_iid[1][iid] = row
            order_id = str(row.get("order_id") or "")
            is_jit = self._is_jit_row(row)
            checked_var = tk.BooleanVar(value=bool(order_id and is_jit and order_id not in self.unchecked_edit_order_ids))
            if order_id:
                self.edit_check_vars[order_id] = checked_var
            if text is None:
                continue

            card_start = text.index("end-1c")
            chk = tk.Checkbutton(
                text,
                variable=checked_var,
                bg="#F8F9FA",
                activebackground="#F8F9FA",
                command=lambda oid=order_id, row_iid=iid: (self._select_edit_card(row_iid), self._toggle_edit_card(oid)),
            )
            if not is_jit or not order_id:
                chk.configure(state="disabled")
            self._bind_widget_mousewheel(chk, text)
            text.window_create("end", window=chk, padx=4, pady=2)
            text.insert(
                "end",
                f"  {row.get('order_no') or '-'}  {row.get('tag_name') or '-'}  {row.get('shop_name') or '-'}\n",
                ("edit_title", row_tag),
            )
            text.insert(
                "end",
                f"状态：{row.get('status_text') or '-'}    件数：{row.get('buy_number_count') or 0}    SKU数：{row.get('buy_type_count') or 0}    平台时间：{row.get('plattime') or '-'}\n",
                ("edit_meta", row_tag),
            )
            text.insert("end", "\n", ("edit_card_gap", row_tag))
            card_end = text.index("end-1c")
            text.tag_add(row_tag, card_start, card_end)
            text.tag_bind(row_tag, "<Button-1>", lambda _event, row_iid=iid: self._select_edit_card(row_iid))
            text.tag_bind(row_tag, "<Button-3>", lambda event, row_iid=iid: self._show_edit_card_context_menu(event, row_iid))

        if rows:
            self.selected_edit_iid = "1-row-1"
            self._refresh_edit_card_selection()
        if text is not None:
            text.configure(state="disabled")

    def _set_payment_cards_empty(self, message: str) -> None:
        self.payment_render_generation += 1
        self.payment_card_images.clear()
        self.payment_card_frames.clear()
        self.payment_check_vars.clear()
        self.selected_payment_iid = ""
        text = getattr(self, "payment_text", None)
        if text is not None:
            text.configure(state="normal")
            text.delete("1.0", "end")
            text.insert("end", message + "\n", ("payment_meta",))
            text.configure(state="disabled")

    def _payment_check_mark_for_row(self, row: dict[str, Any]) -> str:
        order_id = str(row.get("order_id") or "")
        if not order_id or not self._is_jit_row(row):
            return "-"
        return "✓" if order_id not in self.unchecked_payment_order_ids else ""

    def _set_payment_preview_empty(self, message: str) -> None:
        self.payment_card_images.clear()
        text = getattr(self, "payment_text", None)
        if text is not None:
            text.configure(state="normal")
            text.delete("1.0", "end")
            text.insert("end", message + "\n", ("payment_meta",))
            text.configure(state="disabled")

    def _on_payment_tree_click(self, event) -> str | None:
        tree = getattr(self, "payment_tree", None)
        if tree is None:
            return None
        iid = tree.identify_row(event.y)
        if not iid:
            return None
        self._select_payment_card(iid)
        if tree.identify_column(event.x) == "#1":
            row = self.order_rows_by_status_iid.get(2, {}).get(iid)
            if row:
                self._toggle_payment_card(str(row.get("order_id") or ""))
            return "break"
        return None

    def _on_payment_tree_select(self, _event=None) -> None:
        tree = getattr(self, "payment_tree", None)
        if tree is None:
            return
        selection = tree.selection()
        if selection:
            self.selected_payment_iid = selection[0]
        self._refresh_payment_preview()
        self._on_tree_select()

    def _toggle_selected_payment_tree(self, _event=None) -> str:
        self._toggle_selected_payment_order()
        return "break"

    def _toggle_selected_payment_order(self) -> None:
        row = self.get_selected_rows()[0] if self.get_selected_rows() else None
        if not row:
            return
        self._toggle_payment_card(str(row.get("order_id") or ""))

    def _refresh_payment_preview(self) -> None:
        selected = self.selected_payment_iid
        for iid, frame in self.payment_card_frames.items():
            color = "#0D6EFD" if iid == selected else "#E9ECEF"
            try:
                frame.configure(highlightbackground=color, highlightcolor=color)
            except tk.TclError:
                pass

    def _resize_payment_text_cards(self, _event=None) -> None:
        text = getattr(self, "payment_text", None)
        if text is None:
            return
        try:
            width = max(320, text.winfo_width() - 28)
        except tk.TclError:
            return
        for frame in self.payment_card_frames.values():
            try:
                frame.configure(width=width)
            except tk.TclError:
                pass

    def _toggle_payment_card(self, order_id: str) -> None:
        if not order_id:
            return
        if order_id in self.unchecked_payment_order_ids:
            self.unchecked_payment_order_ids.remove(order_id)
        else:
            self.unchecked_payment_order_ids.add(order_id)
        self._refresh_payment_marks()
        self._on_tree_select()

    def _sync_payment_check_var(self, order_id: str, var: tk.BooleanVar) -> None:
        expected = order_id not in self.unchecked_payment_order_ids
        if var.get() != expected:
            var.set(expected)

    def _select_payment_card(self, iid: str) -> None:
        if not iid:
            return
        self.selected_payment_iid = iid
        tree = getattr(self, "payment_tree", None)
        if tree is not None and iid in tree.get_children():
            tree.selection_set(iid)
            tree.focus(iid)
        self._refresh_payment_card_selection()
        self._refresh_payment_preview()
        self._on_tree_select()

    def _refresh_payment_card_selection(self) -> None:
        selected = self.selected_payment_iid
        for iid, frame in self.payment_card_frames.items():
            color = "#0D6EFD" if iid == selected else "#E9ECEF"
            frame.configure(highlightbackground=color, highlightcolor=color)
        tree = getattr(self, "payment_tree", None)
        if tree is not None and selected in tree.get_children():
            tree.selection_set(selected)
            tree.focus(selected)

    def _bind_payment_card_widget(self, widget: tk.Widget, iid: str) -> None:
        widget.bind("<Button-1>", lambda _event, row_iid=iid: self._select_payment_card(row_iid))
        widget.bind("<Button-3>", lambda event, row_iid=iid: self._show_payment_card_context_menu(event, row_iid))
        if hasattr(self, "payment_canvas"):
            self._bind_widget_mousewheel(widget, self.payment_canvas)

    def _show_payment_card_context_menu(self, event, iid: str) -> str:
        self._select_payment_card(iid)
        menu = tk.Menu(self.root, tearoff=False)
        menu.add_command(label="复制订单号", command=self.copy_selected_order_nos)
        menu.add_command(label="放大预览图片", command=self.preview_selected_payment_order_images)
        menu.tk_popup(event.x_root, event.y_root)
        return "break"

    @staticmethod
    def _format_payment_card_size_summary(row: dict[str, Any]) -> str:
        values: list[str] = []
        seen: set[str] = set()
        for detail in row.get("detail") or []:
            if not isinstance(detail, dict):
                continue
            size = str(detail.get("size") or detail.get("spec_size") or detail.get("goods_size") or "").strip()
            if not size:
                continue
            sku = normalize_sku(detail.get("sku") or detail.get("productSku") or detail.get("product_sku") or detail.get("goods_sku"))
            value = f"{sku}：{size}" if sku else size
            if value not in seen:
                values.append(value)
                seen.add(value)
        return "当前尺码：" + "；".join(values) if values else "当前尺码：未返回"

    @staticmethod
    def _generic_size_skus_from_payment_row(row: dict[str, Any]) -> list[str]:
        skus: list[str] = []
        seen: set[str] = set()
        for detail in row.get("detail") or []:
            if not isinstance(detail, dict):
                continue
            size = str(detail.get("size") or detail.get("spec_size") or detail.get("goods_size") or "").strip()
            if size != "通用尺码":
                continue
            sku = normalize_sku(detail.get("sku") or detail.get("productSku") or detail.get("product_sku") or detail.get("goods_sku"))
            value = sku or "未识别 SKU"
            if value not in seen:
                skus.append(value)
                seen.add(value)
        return skus

    def _populate_payment_cards(self, rows: list[dict[str, Any]]) -> None:
        self.payment_render_generation += 1
        self.payment_card_images.clear()
        self.payment_card_frames.clear()
        self.payment_check_vars.clear()
        self.order_rows_by_status_iid[2].clear()
        text = getattr(self, "payment_text", None)
        if text is not None:
            text.configure(state="normal")
            text.delete("1.0", "end")
        self.selected_payment_iid = ""

        if not rows:
            self._set_payment_cards_empty("当前没有待付款订单。")
            return

        for index, row in enumerate(rows, start=1):
            iid = f"2-row-{index}"
            self.order_rows_by_status_iid[2][iid] = row
            order_id = str(row.get("order_id") or "")
            is_jit = self._is_jit_row(row)
            checked_var = tk.BooleanVar(value=bool(order_id and is_jit and order_id not in self.unchecked_payment_order_ids))
            if order_id:
                self.payment_check_vars[order_id] = checked_var
            if text is None:
                continue

            card = tk.Frame(
                text,
                bg="#FFFFFF",
                highlightbackground="#E9ECEF",
                highlightcolor="#E9ECEF",
                highlightthickness=1,
                padx=8,
                pady=6,
            )
            self.payment_card_frames[iid] = card
            self._bind_payment_card_widget(card, iid)

            header = tk.Frame(card, bg="#FFFFFF")
            header.pack(fill="x")
            self._bind_payment_card_widget(header, iid)

            check = tk.Checkbutton(
                header,
                variable=checked_var,
                bg="#FFFFFF",
                activebackground="#FFFFFF",
                command=lambda oid=order_id: self._toggle_payment_card(oid),
            )
            if not is_jit or not order_id:
                check.configure(state="disabled")
            check.pack(side="left", padx=(0, 6))
            self._bind_widget_mousewheel(check, text)

            title = f"{row.get('order_no') or '-'}  {row.get('tag_name') or '-'}  {row.get('shop_name') or '-'}"
            title_label = ttk.Label(header, text=title, font=("Microsoft YaHei UI", 9, "bold"), background="#FFFFFF")
            title_label.pack(side="left")
            self._bind_payment_card_widget(title_label, iid)
            size_summary = self._format_payment_card_size_summary(row)
            size_label = ttk.Label(
                header,
                text=size_summary,
                font=("Microsoft YaHei UI", 9, "bold"),
                background="#FFFFFF",
                foreground="#B45309" if size_summary.endswith("未返回") else "#047857",
            )
            size_label.pack(side="left", padx=(18, 0))
            self._bind_payment_card_widget(size_label, iid)

            meta = f"状态：{row.get('status_text') or '-'}    件数：{row.get('buy_number_count') or 0}    SKU数：{row.get('buy_type_count') or 0}    平台时间：{row.get('plattime') or '-'}"
            meta_label = ttk.Label(card, text=meta, background="#FFFFFF", foreground="#52606D")
            meta_label.pack(anchor="w", pady=(4, 4))
            self._bind_payment_card_widget(meta_label, iid)

            images = self._image_items_for_row(row)
            if not images:
                missing_label = ttk.Label(
                    card,
                    text=f"未找到服务器图片或本地已下载图片。目录：{self._current_output_dir()}",
                    foreground="#B02A37",
                    background="#FFFFFF",
                )
                missing_label.pack(anchor="w", pady=(2, 2))
                self._bind_payment_card_widget(missing_label, iid)
            else:
                image_row = tk.Frame(card, bg="#FFFFFF")
                image_row.pack(fill="x", pady=(4, 2))
                self._bind_payment_card_widget(image_row, iid)
                for item in images:
                    block = tk.Frame(image_row, bg="#FFFFFF", padx=8, pady=4)
                    block.pack(side="left", padx=14)
                    self._bind_payment_card_widget(block, iid)
                    sku_text = str(item.get("sku") or "").strip() or Path(str(item.get("file") or "")).name
                    if sku_text:
                        sku_label = ttk.Label(block, text=sku_text, background="#FFFFFF", foreground="#52606D")
                        sku_label.pack(anchor="center")
                        self._bind_payment_card_widget(sku_label, iid)
                    image_label = ttk.Label(block, text="图片加载中...", background="#FFFFFF", foreground="#6C757D")
                    image_label.pack(anchor="center", pady=(4, 0))
                    image_label.configure(cursor="hand2")
                    self._bind_payment_card_widget(image_label, iid)
                    image_label.bind(
                        "<Button-1>",
                        lambda _event, row_iid=iid: (self._select_payment_card(row_iid), self.preview_selected_payment_order_images()),
                    )
                    self._load_image_async(item, image_label, (760, 520), self.payment_card_images, trim=True, high_res=False)

            text.window_create("end", window=card, padx=0, pady=6)
            text.insert("end", "\n")

        if rows:
            self.selected_payment_iid = "2-row-1"
            self._refresh_payment_card_selection()
            self._refresh_payment_preview()
        self._resize_payment_text_cards()
        if text is not None:
            text.configure(state="disabled")

    def _refresh_payment_marks(self) -> None:
        for order_id, var in self.payment_check_vars.items():
            self._sync_payment_check_var(order_id, var)

    def _refresh_edit_marks(self) -> None:
        for order_id, var in self.edit_check_vars.items():
            self._sync_edit_check_var(order_id, var)

    def bulk_uncheck_payment_orders(self) -> None:
        target_order_nos = parse_order_no_set(self.bulk_uncheck_order_nos_var.get())
        if not target_order_nos:
            messagebox.showinfo("取消勾选", "请先输入 WB 订单号，多个用空格或逗号分隔。")
            return

        rows = self.order_rows_by_status_iid.get(2, {})
        if not rows:
            messagebox.showinfo("取消勾选", "当前待付款列表为空，请先刷新订单。")
            return

        matched_order_nos: list[str] = []
        seen_jit_order_nos: set[str] = set()
        for row in rows.values():
            if not self._is_jit_row(row):
                continue
            order_no_key = normalize_order_no(row.get("order_no"))
            if order_no_key:
                seen_jit_order_nos.add(order_no_key)
            if order_no_key in target_order_nos:
                order_id = str(row.get("order_id") or "")
                if order_id:
                    self.unchecked_payment_order_ids.add(order_id)
                    matched_order_nos.append(str(row.get("order_no") or ""))

        if matched_order_nos:
            if not messagebox.askyesno(
                "确认取消勾选",
                "\n".join(
                    [
                        f"检测到 {len(matched_order_nos)} 个待付款 JIT 可取消勾选。",
                        f"订单：{self._format_order_no_preview(matched_order_nos)}",
                        "",
                        "确定把这些订单从本次支付里排除吗？",
                    ]
                ),
            ):
                return

        self._refresh_payment_marks()
        self._on_tree_select()

        not_found = sorted(target_order_nos - seen_jit_order_nos)
        message = f"已取消勾选 {len(matched_order_nos)} 个待付款 JIT。"
        if not_found:
            message += f"\n未在待付款 JIT 中找到：{', '.join(not_found[:10])}"
            if len(not_found) > 10:
                message += " ..."
        messagebox.showinfo("取消勾选", message)

    def _build_payment_size_items(self) -> list[dict[str, Any]]:
        def first_value(source: dict[str, Any], keys: Iterable[str]) -> str:
            for key in keys:
                value = source.get(key)
                if value not in (None, ""):
                    return normalize_db_key(value)
            return ""

        items: list[dict[str, Any]] = []
        for row in self.order_rows_by_status_iid.get(2, {}).values():
            details = row.get("detail") or []
            if not isinstance(details, list):
                continue
            for detail in details:
                if not isinstance(detail, dict):
                    continue
                order_detail_id = detail.get("id") or detail.get("order_detail_id") or detail.get("item_id")
                sku = normalize_sku(
                    detail.get("sku")
                    or detail.get("sku_id")
                    or detail.get("skuId")
                    or detail.get("sku_code")
                    or detail.get("skuCode")
                    or detail.get("productSku")
                    or detail.get("product_sku")
                    or detail.get("product_sku_id")
                    or detail.get("goods_sku")
                )
                skc_id = first_value(
                    detail,
                    ("skc", "SKC", "skc_id", "skcId", "product_skc_id", "productSkcId", "product_skcId"),
                )
                product_id = first_value(
                    detail,
                    ("product_id", "productId", "spu_id", "spuId", "goods_id", "goodsId", "design_product_id"),
                ) or first_value(row, ("product_id", "productId", "spu_id", "spuId", "goods_id", "goodsId"))
                product_no = get_product_no_from_sources(detail, row)
                match_candidates = []
                # SKU 是唯一规格键，必须优先；SKC/SPU 仅作为兼容旧接口的回退。
                for value in (sku, skc_id, product_id):
                    if value and value not in match_candidates:
                        match_candidates.append(value)
                current_size = str(detail.get("size") or detail.get("spec_size") or detail.get("goods_size") or "").strip()
                current_size_id = normalize_option_id(detail.get("size_id") or detail.get("sizeId"))
                if not order_detail_id or not sku:
                    continue
                items.append(
                    {
                        "order_no": str(row.get("order_no") or ""),
                        "order_id": str(row.get("order_id") or ""),
                        "internal_order_id": str(row.get("id") or ""),
                        "tag_name": str(row.get("tag_name") or ""),
                        "shop_name": str(row.get("shop_name") or ""),
                        "sku": sku,
                        "skc_id": skc_id,
                        "product_id": product_id,
                        "product_no": product_no,
                        "match_candidates": match_candidates,
                        "order_detail_id": str(order_detail_id),
                        "current_size": current_size,
                        "current_size_id": current_size_id,
                        "list_detail": detail,
                    }
                )
        return items

    def open_payment_size_editor(self) -> None:
        if self.size_editor_window is not None:
            try:
                if self.size_editor_window.winfo_exists():
                    self.size_editor_window.lift()
                    self.size_editor_window.focus_force()
                    return
            except Exception:
                self.size_editor_window = None

        if self._active_status() != 2:
            try:
                self.notebook.select(1)
            except Exception:
                pass
        items = self._build_payment_size_items()

        window = tk.Toplevel(self.root)
        self.size_editor_window = window
        window.title("修改待付款订单成分（尺码）")
        window.geometry("1050x650")
        window.minsize(920, 520)
        window.configure(background="#F8F9FA")
        window.transient(self.root)
        window.protocol("WM_DELETE_WINDOW", self._close_size_editor_window)

        file_var = tk.StringVar(value=DEFAULT_COMPOSITION_XLSX)
        status_var = tk.StringVar(
            value=(
                "当前无待付款订单，可先选择并保存成分数据库目录。"
                if not items
                else "可手填目标尺码；优先选择成分数据库目录后点击“开始匹配”。提交时使用“全部关联”。"
            )
        )
        views: list[dict[str, Any]] = []

        toolbar = ttk.Frame(window)
        toolbar.pack(fill="x", padx=8, pady=8)
        ttk.Label(toolbar, text="数据库目录").pack(side="left")
        ttk.Entry(toolbar, textvariable=self.composition_db_folder_var, state="readonly").pack(
            side="left", fill="x", expand=True, padx=6
        )
        ttk.Button(toolbar, text="选择数据库文件夹", command=lambda: self._choose_composition_db_folder(status_var), style="Ghost.TButton").pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(toolbar, text="开始匹配", command=lambda: self._apply_saved_db_composition(views, status_var), style="Match.TButton").pack(
            side="left", padx=(2, 0)
        )

        legacy_toolbar = ttk.Frame(window)
        legacy_toolbar.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Label(legacy_toolbar, text="旧版表格导入（通常不需要）").pack(side="left")
        ttk.Entry(legacy_toolbar, textvariable=file_var).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(legacy_toolbar, text="选择表格", command=lambda: self._choose_composition_file(file_var), style="Ghost.TButton").pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(
            legacy_toolbar,
            text="导入表格快速填写",
            command=lambda: self._apply_composition_to_size_views(file_var.get(), views, status_var),
            style="Ghost.TButton",
        ).pack(side="left")

        status_label = ttk.Label(window, textvariable=status_var, background="#F8F9FA", foreground="#6C757D")
        status_label.pack(fill="x", padx=8, pady=(0, 6))

        col_widths = [40, 160, 140, 80, 100, 120, 100, 240]
        header_texts = ["选", "订单号", "SKU", "标签", "原尺码", "目标尺码", "匹配状态", "成分"]

        header = tk.Frame(window, bg="#E9ECEF", padx=6, pady=4)
        header.pack(fill="x", padx=8)
        for column, (text, width) in enumerate(zip(header_texts, col_widths)):
            header.columnconfigure(column, minsize=width)
            tk.Label(
                header,
                text=text,
                anchor="w",
                bg="#E9ECEF",
                font=("Microsoft YaHei UI", 9, "bold"),
            ).grid(row=0, column=column, sticky="w", padx=2)

        canvas = tk.Canvas(window, bg="#F8F9FA", highlightthickness=0)
        scrollbar = ttk.Scrollbar(window, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        body_window = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(body_window, width=event.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        scrollbar.pack(side="right", fill="y", padx=(0, 8), pady=(0, 8))
        self._bind_widget_mousewheel(canvas, canvas)
        self._bind_widget_mousewheel(body, canvas)

        for index, item in enumerate(items, start=1):
            row_bg = "#FFFFFF" if index % 2 else "#F8F9FA"
            row_frame = tk.Frame(body, bg=row_bg, padx=6, pady=3)
            row_frame.pack(fill="x")
            self._bind_widget_mousewheel(row_frame, canvas)
            for column, width in enumerate(col_widths):
                row_frame.columnconfigure(column, minsize=width)

            checked_var = tk.BooleanVar(value=False)
            target_var = tk.StringVar(value="")
            composition_var = tk.StringVar(value="")
            note_var = tk.StringVar(value="待填写")
            view = {
                "item": item,
                "checked": checked_var,
                "target": target_var,
                "composition": composition_var,
                "note": note_var,
            }
            views.append(view)

            check = tk.Checkbutton(row_frame, variable=checked_var, bg=row_bg, activebackground=row_bg)
            check.grid(row=0, column=0, sticky="w", padx=2)
            self._bind_widget_mousewheel(check, canvas)

            for column, text in (
                (1, item["order_no"]),
                (2, item["sku"]),
                (3, item["tag_name"]),
                (4, item["current_size"] or "-"),
            ):
                label = tk.Label(row_frame, text=text, anchor="w", bg=row_bg)
                label.grid(row=0, column=column, sticky="w", padx=2)
                self._bind_widget_mousewheel(label, canvas)

            combo = ttk.Combobox(row_frame, textvariable=target_var, values=SIZE_TARGET_OPTIONS, width=13, state="readonly")
            combo.grid(row=0, column=5, sticky="w", padx=2)
            combo.bind("<<ComboboxSelected>>", lambda _event, current_view=view: self._mark_size_view_after_manual_edit(current_view))
            self._bind_widget_mousewheel(combo, canvas)

            note_label = ttk.Label(row_frame, textvariable=note_var, background=row_bg, foreground="#6C757D")
            note_label.grid(row=0, column=6, sticky="w", padx=2)
            self._bind_widget_mousewheel(note_label, canvas)
            composition_label = ttk.Label(row_frame, textvariable=composition_var, background=row_bg, foreground="#495057")
            composition_label.grid(row=0, column=7, sticky="w", padx=2)
            self._bind_widget_mousewheel(composition_label, canvas)

        footer = ttk.Frame(window)
        footer.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(footer, text="全选有目标尺码", command=lambda: self._set_size_views_checked(views, True)).pack(side="left")
        ttk.Button(footer, text="全部取消", command=lambda: self._set_size_views_checked(views, False)).pack(side="left", padx=6)
        submit_button = ttk.Button(footer, text="提交勾选修改")
        submit_button.configure(command=lambda: self._submit_payment_size_changes(window, views, status_var, submit_button))
        if not views:
            submit_button.configure(state="disabled")
        submit_button.pack(side="right")

    def _close_size_editor_window(self) -> None:
        window = self.size_editor_window
        self.size_editor_window = None
        if window is not None:
            try:
                window.destroy()
            except Exception:
                pass

    def _choose_composition_file(self, file_var: tk.StringVar) -> None:
        chosen = filedialog.askopenfilename(
            title="选择订单成分匹配结果表",
            initialdir=str(Path.home() / "Documents"),
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if chosen:
            file_var.set(chosen)

    def _get_composition_db_folder(self) -> Path:
        configured = self.composition_db_folder_var.get().strip()
        return Path(configured).expanduser() if configured else COMPOSITION_DB_FOLDER

    def _save_composition_db_folder(self, folder: Path) -> Path:
        selected_folder = folder.expanduser()
        self.composition_db_folder_var.set(str(selected_folder))
        settings = load_app_settings()
        settings["composition_db_folder"] = str(selected_folder)
        save_app_settings(settings)
        return selected_folder

    def _choose_composition_db_folder(self, status_var: tk.StringVar | None = None) -> None:
        chosen = filedialog.askdirectory(title="选择成分数据库文件夹", initialdir=str(self._get_composition_db_folder()))
        if not chosen:
            return
        try:
            selected_folder = self._save_composition_db_folder(Path(chosen))
        except OSError as exc:
            messagebox.showerror("成分数据库", f"保存目录设置失败：{exc}")
            return
        message = f"成分数据库目录已保存：{selected_folder}"
        self.status_var.set(message)
        if status_var is not None:
            status_var.set(message)

    def _apply_saved_db_composition(self, views: list[dict[str, Any]], status_var: tk.StringVar) -> None:
        db_folder = self._get_composition_db_folder()
        if not views:
            status_var.set(f"当前无待付款订单。已保存的成分数据库目录：{db_folder}")
            return
        self._apply_db_composition_to_size_views(views, status_var, db_folder)

    def _match_payment_size_items_with_db(
        self,
        items: list[dict[str, Any]],
        db_folder: Path,
        *,
        client: LandwuClient | None = None,
    ) -> dict[str, Any]:
        query_values: list[str] = []
        for item in items:
            candidates = item.get("match_candidates") or [item.get("sku")]
            for value in candidates:
                key = normalize_db_key(value)
                if key:
                    query_values.append(key)

        payload = load_composition_db_mapping(query_values, db_folder=db_folder)
        mapping = payload["mapping"]
        matched = 0
        changed = 0
        no_target = 0
        no_sku = 0
        same = 0
        fallback_count = 0
        targets: list[dict[str, Any]] = []
        result_by_detail_id: dict[str, dict[str, Any]] = {}

        for item in items:
            detail_id = str(item.get("order_detail_id") or "")
            result = {
                "composition": "",
                "target_size": "",
                "current_size": str(item.get("current_size") or "").strip(),
                "current_size_id": normalize_option_id(item.get("current_size_id")),
                "target_size_id": "",
                "note": "未匹配",
                "checked": False,
            }
            if client and detail_id:
                try:
                    size_state = client.get_order_size_state(
                        detail_id,
                        list_detail=item.get("list_detail") if isinstance(item.get("list_detail"), dict) else None,
                        product_id=item.get("product_id"),
                    )
                    if size_state.get("currentSizeId"):
                        item["current_size_id"] = normalize_option_id(size_state.get("currentSizeId"))
                        result["current_size_id"] = item["current_size_id"]
                    if size_state.get("currentSizeName"):
                        item["current_size"] = str(size_state.get("currentSizeName") or "").strip()
                        result["current_size"] = item["current_size"]
                    item["size_options"] = size_state.get("sizeOptions") or {}
                except Exception as exc:  # noqa: BLE001
                    result["size_state_error"] = str(exc)
            record = None
            for value in item.get("match_candidates") or [item.get("sku")]:
                key = normalize_db_key(value)
                if key and key in mapping:
                    record = mapping[key]
                    break
            if not record:
                record = infer_polyester_fallback_from_product_no(item.get("product_no"))
                if record:
                    fallback_count += 1
                else:
                    no_sku += 1
                    result_by_detail_id[detail_id] = result
                    continue

            matched += 1
            composition = str(record.get("composition") or "")
            target = str(record.get("target_size") or "")
            current = str(item.get("current_size") or "").strip()
            current_size_id = normalize_option_id(item.get("current_size_id"))
            target_option = find_named_option(item.get("size_options") or {}, target)
            target_size_id = normalize_option_id(target_option.get("id"))
            result["composition"] = composition
            result["target_size"] = target
            result["current_size"] = current
            result["current_size_id"] = current_size_id
            result["target_size_id"] = target_size_id
            if not target:
                result["note"] = "无法识别"
                no_target += 1
            elif (current_size_id and target_size_id and current_size_id == target_size_id) or target == current:
                result["note"] = "相同跳过（尺码ID）" if current_size_id and target_size_id else "相同跳过"
                same += 1
            else:
                result["note"] = str(record.get("db_field") or "已匹配")
                result["checked"] = True
                targets.append(
                    {
                        "order_no": item.get("order_no"),
                        "sku": item.get("sku"),
                        "order_detail_id": item.get("order_detail_id"),
                        "target_size": target,
                        "target_size_id": target_size_id,
                    }
                )
                changed += 1
            result_by_detail_id[detail_id] = result

        return {
            "targets": targets,
            "resultByDetailId": result_by_detail_id,
            "matched": matched,
            "changed": changed,
            "same": same,
            "noSku": no_sku,
            "noTarget": no_target,
            "fallbackCount": fallback_count,
            "dbFileCount": payload.get("dbFileCount") or 0,
            "skippedFiles": payload.get("skippedFiles") or [],
            "dbFolder": str(db_folder),
        }

    @staticmethod
    def _format_payment_size_match_message(payload: dict[str, Any]) -> str:
        message = (
            f"成分数据库匹配完成：匹配 {payload.get('matched') or 0}，可修改 {payload.get('changed') or 0}，"
            f"相同 {payload.get('same') or 0}，未匹配 {payload.get('noSku') or 0}，"
            f"无法识别 {payload.get('noTarget') or 0}；读取 {payload.get('dbFileCount') or 0} 个文件"
        )
        fallback_count = int(payload.get("fallbackCount") or 0)
        if fallback_count:
            message += f"，货号兜底 {fallback_count}"
        skipped = payload.get("skippedFiles") or []
        if skipped:
            message += f"，跳过 {len(skipped)} 个文件"
        return message

    def _mark_size_view_after_manual_edit(self, view: dict[str, Any]) -> None:
        target = str(view["target"].get() or "").strip()
        current = str((view.get("item") or {}).get("current_size") or "").strip()
        view["checked"].set(bool(target and target != current))
        if not target:
            view["note"].set("待填写")
        elif target == current:
            view["note"].set("相同跳过")
        else:
            view["note"].set("手填")

    def _apply_composition_to_size_views(
        self,
        file_path: str,
        views: list[dict[str, Any]],
        status_var: tk.StringVar,
    ) -> None:
        try:
            payload = load_composition_xlsx(file_path)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("导入成分表", str(exc))
            return

        mapping = payload["mapping"]
        matched = 0
        changed = 0
        no_target = 0
        no_sku = 0
        same = 0
        for view in views:
            item = view["item"]
            sku = normalize_sku(item.get("sku"))
            record = mapping.get(sku)
            if not record:
                view["composition"].set("")
                view["target"].set("")
                view["checked"].set(False)
                view["note"].set("未匹配")
                no_sku += 1
                continue
            matched += 1
            composition = record.get("composition") or ""
            target = record.get("target_size") or ""
            current = str(item.get("current_size") or "").strip()
            view["composition"].set(composition)
            view["target"].set(target)
            if not target:
                view["checked"].set(False)
                view["note"].set("无法识别")
                no_target += 1
            elif target == current:
                view["checked"].set(False)
                view["note"].set("相同跳过")
                same += 1
            else:
                view["checked"].set(True)
                view["note"].set("已匹配")
                changed += 1

        conflicts = payload.get("conflicts") or []
        message = f"导入完成：SKU匹配 {matched}，可修改 {changed}，相同 {same}，未匹配 {no_sku}，无法识别 {no_target}"
        if conflicts:
            message += f"，冲突 {len(conflicts)}"
        status_var.set(message)

    def _apply_db_composition_to_size_views(
        self,
        views: list[dict[str, Any]],
        status_var: tk.StringVar,
        db_folder: Path,
    ) -> None:
        try:
            status_var.set(f"正在读取数据库：{db_folder}")
            self.root.update_idletasks()
            payload = self._match_payment_size_items_with_db([view["item"] for view in views], db_folder)
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("成分数据库匹配", str(exc))
            return

        result_by_detail_id = payload.get("resultByDetailId") or {}
        for view in views:
            item = view["item"]
            result = result_by_detail_id.get(str(item.get("order_detail_id") or "")) or {}
            view["composition"].set(str(result.get("composition") or ""))
            view["target"].set(str(result.get("target_size") or ""))
            view["checked"].set(bool(result.get("checked")))
            view["note"].set(str(result.get("note") or "未匹配"))

        status_var.set(self._format_payment_size_match_message(payload))

    def _set_size_views_checked(self, views: list[dict[str, Any]], checked: bool) -> None:
        for view in views:
            target = str(view["target"].get() or "").strip()
            current = str((view.get("item") or {}).get("current_size") or "").strip()
            view["checked"].set(bool(checked and target and target != current))

    def _submit_payment_size_targets(
        self,
        targets: list[dict[str, Any]],
        *,
        status_var: tk.StringVar | None = None,
        submit_button: ttk.Button | None = None,
        button_text: str = "提交勾选修改",
        progress_window: tk.Toplevel | None = None,
        close_editor_on_success: bool = False,
        title: str = "修改待付款订单成分尺码",
    ) -> None:
        if self.busy:
            messagebox.showinfo("请稍候", "当前已有任务执行中。")
            return

        pending_message = f"正在提交 {len(targets)} 个 SKU，请稍候..."
        self.status_var.set(pending_message)
        if status_var is not None:
            status_var.set(pending_message)
        if submit_button is not None:
            submit_button.configure(text="正在提交...", state="disabled")
        if progress_window is not None:
            try:
                progress_window.update_idletasks()
            except Exception:
                pass

        def task():
            return with_landwu_session(
                self._make_runtime_args(),
                lambda _session, client: client.change_order_detail_sizes(targets, relation_type=1),
            )

        def on_success(payload: dict[str, Any]) -> None:
            failed = payload.get("failed") or []
            message = payload.get("message") or "尺码修改完成"
            done_message = (
                f"尺码修改完成：失败 {len(failed)} / 共 {len(targets)} 个，请查看弹窗。"
                if failed
                else f"尺码修改完成：成功 {len(targets)} 个，正在刷新订单..."
            )
            self.status_var.set(done_message)
            if status_var is not None:
                status_var.set(done_message)
            if failed and submit_button is not None:
                submit_button.configure(text=button_text, state="normal")
            if failed:
                message += "\n失败示例：" + "；".join(
                    f"{item.get('orderNo')} {item.get('sku')}：{item.get('error')}" for item in failed[:3]
                )
            messagebox.showinfo("尺码修改", message)
            if not failed and close_editor_on_success:
                self._close_size_editor_window()
            self.refresh_summary()

        def on_error(error: str) -> None:
            brief = str(error).replace("\r", " ").replace("\n", " ").strip()
            if len(brief) > 80:
                brief = brief[:77] + "..."
            error_message = f"尺码修改失败：{brief}" if brief else "尺码修改失败"
            self.status_var.set(error_message)
            if status_var is not None:
                status_var.set(error_message)
            if submit_button is not None:
                submit_button.configure(text=button_text, state="normal")

        self._run_task(title, task, on_success=on_success, on_error=on_error)

    def _submit_payment_size_changes(
        self,
        window: tk.Toplevel,
        views: list[dict[str, Any]],
        status_var: tk.StringVar | None = None,
        submit_button: ttk.Button | None = None,
    ) -> None:
        targets: list[dict[str, Any]] = []
        skipped_blank = 0
        skipped_same = 0
        for view in views:
            if not view["checked"].get():
                continue
            item = view["item"]
            target_size = str(view["target"].get() or "").strip()
            current_size = str(item.get("current_size") or "").strip()
            if not target_size:
                skipped_blank += 1
                continue
            if target_size == current_size:
                skipped_same += 1
                continue
            targets.append(
                {
                    "order_no": item.get("order_no"),
                    "sku": item.get("sku"),
                    "order_detail_id": item.get("order_detail_id"),
                    "target_size": target_size,
                }
            )

        if not targets:
            messagebox.showinfo("提交尺码修改", "没有可提交的修改。请先填写目标尺码并勾选。")
            return

        order_nos = [str(item.get("order_no") or "") for item in targets]
        if not messagebox.askyesno(
            "确认修改尺码",
            "\n".join(
                [
                    f"将按“全部关联”修改 {len(targets)} 个待付款 SKU。",
                    f"涉及订单：{self._format_order_no_preview(order_nos)}",
                    f"空目标跳过：{skipped_blank}，相同跳过：{skipped_same}",
                    "",
                    "确定提交吗？",
                ]
            ),
        ):
            return

        self._submit_payment_size_targets(
            targets,
            status_var=status_var,
            submit_button=submit_button,
            progress_window=window,
            close_editor_on_success=True,
        )

    def quick_match_submit_payment_sizes(self) -> None:
        if self._active_status() != 2:
            try:
                self.notebook.select(1)
            except Exception:
                pass
        items = self._build_payment_size_items()
        if not items:
            messagebox.showinfo("一键匹配并提交成分", "当前没有待付款订单。请先刷新订单。")
            return

        db_folder = self._get_composition_db_folder()

        def task():
            def worker(_session, client: LandwuClient) -> dict[str, Any]:
                return self._match_payment_size_items_with_db(items, db_folder, client=client)

            return with_landwu_session(self._make_runtime_args(), worker)

        def on_match_success(payload: dict[str, Any]) -> None:
            message = self._format_payment_size_match_message(payload)
            self.status_var.set(message)
            targets = payload.get("targets") or []
            if not targets:
                messagebox.showinfo("一键匹配并提交成分", message + "\n\n没有可提交的尺码修改。")
                return
            order_nos = [str(item.get("order_no") or "") for item in targets]
            if not messagebox.askyesno(
                "确认提交成分尺码",
                "\n".join(
                    [
                        message,
                        f"将按“全部关联”提交 {len(targets)} 个待付款 SKU。",
                        f"涉及订单：{self._format_order_no_preview(order_nos)}",
                        f"数据库目录：{db_folder}",
                        "",
                        "确定提交吗？",
                    ]
                ),
            ):
                return
            self._submit_payment_size_targets(targets, title="一键提交成分尺码")

        self._run_task("一键匹配成分数据库", task, on_success=on_match_success)

    def _active_status(self) -> int:
        try:
            tab_index = self.notebook.index("current")
        except Exception:  # noqa: BLE001
            return 1
        if 0 <= tab_index < len(self.tab_statuses):
            return self.tab_statuses[tab_index]
        return 1

    def _on_tree_select(self, _event=None) -> None:
        rows = self.get_selected_rows()
        status = self._active_status()
        self._show_toolbar_for_status(status)
        status_text = STATUS_TEXT_MAP.get(status, str(status))
        if not rows:
            if status == 1:
                counts = self.get_edit_counts()
                self.selection_hint_var.set(f"待编辑打勾的 JIT 会进入一键流程。当前勾选 JIT：{counts['checked']} 单。")
            elif status == 2:
                checked_count = len(self.get_checked_payment_order_ids())
                self.selection_hint_var.set(f"待付款每单直接显示预览图；打勾表示会支付。当前勾选 JIT：{checked_count} 单。")
            else:
                self.selection_hint_var.set(f"{status_text}未选择订单。该标签页只用于查看。")
            return

        order_nos = [row.get("order_no") or "-" for row in rows[:3]]
        suffix = "" if len(rows) <= 3 else " ..."
        action_text = "一键流程按打勾的待编辑 JIT 执行。" if status == 1 else "每单已直接显示图片，支付按打勾的 JIT 执行。" if status == 2 else "该标签页只用于查看。"
        self.selection_hint_var.set(
            f"{status_text}已选 {len(rows)} 单：{', '.join(order_nos)}{suffix}。{action_text}"
        )

    def get_selected_rows(self) -> list[dict[str, Any]]:
        status = self._active_status()
        if status == 1:
            row_map = self.order_rows_by_status_iid.get(1, {})
            if self.selected_edit_iid and self.selected_edit_iid in row_map:
                return [row_map[self.selected_edit_iid]]
            return []
        if status == 2:
            row_map = self.order_rows_by_status_iid.get(2, {})
            if self.selected_payment_iid and self.selected_payment_iid in row_map:
                return [row_map[self.selected_payment_iid]]
            return []
        tree = self.trees.get(status)
        if tree is None:
            return []
        row_map = self.order_rows_by_status_iid.get(status, {})
        return [row_map[iid] for iid in tree.selection() if iid in row_map]

    def get_checked_payment_order_ids(self) -> list[str]:
        ids: list[str] = []
        for row in self.order_rows_by_status_iid.get(2, {}).values():
            order_id = str(row.get("order_id") or "")
            if order_id and self._is_jit_row(row) and order_id not in self.unchecked_payment_order_ids:
                ids.append(order_id)
        return ids

    def get_checked_payment_order_nos(self) -> list[str]:
        order_nos: list[str] = []
        for row in self.order_rows_by_status_iid.get(2, {}).values():
            order_id = str(row.get("order_id") or "")
            if order_id and self._is_jit_row(row) and order_id not in self.unchecked_payment_order_ids:
                order_nos.append(str(row.get("order_no") or ""))
        return order_nos

    def get_checked_payment_generic_size_orders(self) -> list[dict[str, Any]]:
        orders: list[dict[str, Any]] = []
        for row in self.order_rows_by_status_iid.get(2, {}).values():
            order_id = str(row.get("order_id") or "")
            if not order_id or not self._is_jit_row(row) or order_id in self.unchecked_payment_order_ids:
                continue
            skus = self._generic_size_skus_from_payment_row(row)
            if skus:
                orders.append({"order_id": order_id, "order_no": str(row.get("order_no") or ""), "skus": skus})
        return orders

    def get_payment_counts(self) -> dict[str, int]:
        jit_total = 0
        checked = 0
        unchecked = 0
        for row in self.order_rows_by_status_iid.get(2, {}).values():
            if not self._is_jit_row(row):
                continue
            jit_total += 1
            order_id = str(row.get("order_id") or "")
            if order_id and order_id not in self.unchecked_payment_order_ids:
                checked += 1
            else:
                unchecked += 1
        return {"jitTotal": jit_total, "checked": checked, "unchecked": unchecked}

    def _confirm_payment_with_size_warning(
        self,
        counts: dict[str, int],
        selected_order_nos: list[str],
        generic_size_orders: list[dict[str, Any]],
    ) -> str | None:
        message_lines = [
            f"待付款 JIT 共 {counts['jitTotal']} 单。",
            f"当前勾选 {counts['checked']} 单，取消勾选 {counts['unchecked']} 单。",
            f"将先预检，预检通过后真实支付：{self._format_order_no_preview(selected_order_nos)}",
        ]
        if not generic_size_orders:
            return "all" if messagebox.askyesno("确认支付", "\n".join([*message_lines, "", "确认图片无误，并继续吗？"])) else None

        warning_orders: list[str] = []
        for item in generic_size_orders:
            order_no = str(item.get("order_no") or "-")
            skus = [str(sku) for sku in item.get("skus") or [] if str(sku)]
            sku_text = "、".join(skus[:3])
            if len(skus) > 3:
                sku_text += "…"
            prefix = "无法确认尺码" if item.get("uncertain") else "通用尺码"
            warning_orders.append(f"{prefix}：{order_no}（{sku_text}）" if sku_text else f"{prefix}：{order_no}")

        safe_count = max(0, counts["checked"] - len(generic_size_orders))
        result: dict[str, str | None] = {"action": None}
        dialog = tk.Toplevel(self.root)
        dialog.title("确认支付")
        dialog.configure(background="#FFFFFF")
        dialog.resizable(False, False)
        dialog.transient(self.root)

        def close(action: str | None = None) -> None:
            result["action"] = action
            try:
                dialog.grab_release()
            except Exception:
                pass
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", close)
        dialog.bind("<Escape>", lambda _event: close())

        body = tk.Frame(dialog, bg="#FFFFFF", padx=24, pady=20)
        body.pack(fill="both", expand=True)
        tk.Label(
            body,
            text="\n".join(message_lines),
            justify="left",
            anchor="w",
            bg="#FFFFFF",
            fg="#1F2933",
            font=("Microsoft YaHei UI", 9),
        ).pack(fill="x", anchor="w")

        warning_frame = tk.Frame(body, bg="#FEF2F2", highlightbackground="#FCA5A5", highlightthickness=1, padx=12, pady=10)
        warning_frame.pack(fill="x", pady=(14, 0))
        tk.Label(
            warning_frame,
            text="有未改成分尺码或无法确认尺码的订单，确认付款吗？",
            justify="left",
            anchor="w",
            bg="#FEF2F2",
            fg="#B91C1C",
            font=("Microsoft YaHei UI", 10, "bold"),
        ).pack(fill="x", anchor="w")
        tk.Label(
            warning_frame,
            text="风险订单：" + "；".join(warning_orders),
            justify="left",
            anchor="w",
            wraplength=500,
            bg="#FEF2F2",
            fg="#B91C1C",
            font=("Microsoft YaHei UI", 9),
        ).pack(fill="x", anchor="w", pady=(6, 0))

        tk.Label(
            body,
            text=f"可选择只支付已改好尺码的 {safe_count} 单，以上通用尺码订单会自动跳过。",
            bg="#FFFFFF",
            fg="#52606D",
        ).pack(anchor="w", pady=(14, 0))
        buttons = ttk.Frame(body)
        buttons.pack(fill="x", pady=(18, 0))
        cancel_button = ttk.Button(buttons, text="取消", command=close)
        cancel_button.pack(side="right")
        ttk.Button(buttons, text="仍要付款全部", command=lambda: close("all"), style="Danger.TButton").pack(side="right", padx=(0, 8))
        safe_button = ttk.Button(
            buttons,
            text=f"只支付已改好尺码（{safe_count}单）",
            command=lambda: close("safe_only"),
            style="Accent.TButton",
        )
        if safe_count <= 0:
            safe_button.configure(state="disabled")
        safe_button.pack(side="right", padx=(0, 8))

        dialog.update_idletasks()
        x = self.root.winfo_rootx() + max(0, (self.root.winfo_width() - dialog.winfo_width()) // 2)
        y = self.root.winfo_rooty() + max(0, (self.root.winfo_height() - dialog.winfo_height()) // 2)
        dialog.geometry(f"+{x}+{y}")
        dialog.grab_set()
        cancel_button.focus_set()
        self.root.wait_window(dialog)
        return result["action"]

    def _format_order_no_preview(self, order_nos: Iterable[str], limit: int = 8) -> str:
        values = [str(item) for item in order_nos if str(item)]
        if not values:
            return "-"
        suffix = "" if len(values) <= limit else f" ... 等 {len(values)} 单"
        return "、".join(values[:limit]) + suffix

    def _local_jit_rows(self, status: int) -> list[dict[str, Any]]:
        return [row for row in self.order_rows_by_status_iid.get(status, {}).values() if self._is_jit_row(row)]

    def get_selected_order_ids(self, *, statuses: Iterable[int] | None = None, tag_name: str | None = None) -> list[str]:
        rows = self.get_selected_rows()
        selected_statuses = {int(item) for item in statuses} if statuses is not None else None
        target_tag = str(tag_name or "").strip().upper() if tag_name else None
        ids: list[str] = []
        for row in rows:
            row_status = int(row.get("status") or 0)
            row_tag = str(row.get("tag_name") or "").strip().upper()
            if selected_statuses is not None and row_status not in selected_statuses:
                continue
            if target_tag and row_tag != target_tag:
                continue
            order_id = str(row.get("order_id") or "")
            if order_id:
                ids.append(order_id)
        return ids

    def get_checked_edit_order_ids(self) -> list[str]:
        ids: list[str] = []
        for row in self.order_rows_by_status_iid.get(1, {}).values():
            order_id = str(row.get("order_id") or "")
            if order_id and self._is_jit_row(row) and order_id not in self.unchecked_edit_order_ids:
                ids.append(order_id)
        return ids

    def get_checked_edit_rows(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in self.order_rows_by_status_iid.get(1, {}).values():
            order_id = str(row.get("order_id") or "")
            if order_id and self._is_jit_row(row) and order_id not in self.unchecked_edit_order_ids:
                rows.append(row)
        return rows

    def get_checked_edit_order_nos(self) -> list[str]:
        return [str(row.get("order_no") or "") for row in self.get_checked_edit_rows()]

    def get_edit_counts(self) -> dict[str, int]:
        jit_total = 0
        checked = 0
        unchecked = 0
        for row in self.order_rows_by_status_iid.get(1, {}).values():
            if not self._is_jit_row(row):
                continue
            jit_total += 1
            order_id = str(row.get("order_id") or "")
            if order_id and order_id not in self.unchecked_edit_order_ids:
                checked += 1
            else:
                unchecked += 1
        return {"jitTotal": jit_total, "checked": checked, "unchecked": unchecked}

    def refresh_summary(self, *, auto: bool = False, startup: bool = False) -> None:
        def task():
            return with_landwu_session(
                self._make_runtime_args(allow_browser_fallback=not (auto or startup)),
                lambda _session, client: client.get_gui_snapshot(),
            )

        def on_success(payload: dict[str, Any]) -> None:
            self._apply_summary(payload.get("summary") or {})
            self._populate_order_tabs(payload.get("ordersByStatus") or {})

        title = "启动刷新订单" if startup else "自动刷新订单" if auto else "刷新统计"
        self._run_task(title, task, on_success=on_success, quiet_if_busy=auto or startup, show_error=not (auto or startup))

    def fetch_auth(self) -> None:
        def task():
            def worker(session, client: LandwuClient) -> dict[str, Any]:
                profile = client.get_user_profile()
                return {
                    "authSource": session.auth_source,
                    "origin": session.origin,
                    "host": session.host,
                    "pageTitle": session.page_title,
                    "pageUrl": session.page_url,
                    "factoryId": session.factory_id,
                    "masterFactoryId": session.master_factory_id,
                    "username": profile.get("username") or session.user_info.get("username"),
                    "nickname": profile.get("nickname") or session.user_info.get("nickname"),
                    "balance": profile.get("money"),
                    "browserPid": session.browser_pid,
                }

            return with_landwu_session(
                self._make_runtime_args(),
                worker,
            )

        def on_success(payload: dict[str, Any]) -> None:
            username = payload.get("username") or "-"
            nickname = payload.get("nickname") or ""
            self.account_var.set(username if not nickname else f"{username} / {nickname}")
            self.factory_var.set(str(payload.get("factoryId") or "-"))
            balance_text = format_money(payload.get("balance"))
            self.balance_var.set(f"¥ {balance_text}" if balance_text != "-" else "-")

        self._run_task("查看登录态", task, on_success=on_success)

    def preview_logistics(self) -> None:
        selected_ids = self.get_checked_edit_order_ids()
        if not selected_ids:
            messagebox.showinfo("JIT改物流预检", "当前没有勾选待编辑 JIT。")
            return

        def task():
            return with_landwu_session(
                self._make_runtime_args(),
                lambda _session, client: client.auto_apply_logistics(commit=False, order_ids=selected_ids),
            )

        self._run_task("JIT改物流预检", task)

    def commit_logistics(self) -> None:
        selected_ids = self.get_checked_edit_order_ids()
        selected_order_nos = self.get_checked_edit_order_nos()
        if not selected_ids:
            messagebox.showinfo("确认改物流", "当前没有勾选待编辑 JIT。")
            return
        counts = self.get_edit_counts()
        if not messagebox.askyesno(
            "确认改物流",
            "\n".join(
                [
                    f"待编辑 JIT 共 {counts['jitTotal']} 单。",
                    f"当前勾选 {counts['checked']} 单，取消勾选 {counts['unchecked']} 单。",
                    f"将真实改物流：{self._format_order_no_preview(selected_order_nos)}",
                    "",
                    "继续吗？",
                ]
            ),
        ):
            return

        def task():
            return with_landwu_session(
                self._make_runtime_args(),
                lambda _session, client: client.auto_apply_logistics(commit=True, order_ids=selected_ids),
            )

        self._log(f"准备真实改物流：勾选 {len(selected_ids)} 个待编辑 JIT")
        self._run_task("JIT真实改物流", task, on_success=lambda _payload: self.refresh_summary())

    def download_images(self) -> None:
        output_dir = self.output_dir_var.get().strip()
        target_rows = self._local_jit_rows(2)
        selected_ids = [str(row.get("order_id") or "") for row in target_rows if row.get("order_id")]
        target_order_nos = [row.get("order_no") or "" for row in target_rows]
        if not target_order_nos:
            messagebox.showinfo("下载图片", "当前没有待付款 JIT 可下载。请先刷新订单。")
            return
        if not selected_ids:
            messagebox.showinfo("下载图片", "当前待付款 JIT 缺少订单ID，请先刷新订单。")
            return
        if not messagebox.askyesno(
            "确认下载图片",
            "\n".join(
                [
                    f"检测到 {len(target_order_nos)} 个待付款 JIT。",
                    f"订单：{self._format_order_no_preview(target_order_nos)}",
                    f"目录：{output_dir or get_default_output_dir()}",
                    "",
                    "确定下载这些订单的图片吗？",
                ]
            ),
        ):
            return

        def task():
            def worker(_session, client: LandwuClient) -> dict[str, Any]:
                download_data = client.download_order_images(
                    output_dir=output_dir,
                    order_ids=selected_ids,
                    plat_order_type=client.get_jit_tag()["id"],
                )
                require_download_success(download_data)
                return download_data

            return with_landwu_session(
                self._make_runtime_args(),
                worker,
            )

        self._run_task("下载待付款 JIT 图片", task)

    def commit_payment(self) -> None:
        selected_ids = self.get_checked_payment_order_ids()
        selected_order_nos = self.get_checked_payment_order_nos()
        if not selected_ids:
            messagebox.showinfo("确认支付", "当前没有勾选待付款 JIT。")
            return
        counts = self.get_payment_counts()
        rows_snapshot = list(self.order_rows_by_status_iid.get(2, {}).values())

        def inspect_task():
            return with_landwu_session(
                self._make_runtime_args(),
                lambda _session, client: client.inspect_payment_size_states(rows_snapshot, selected_ids),
            )

        def continue_payment(inspection: dict[str, Any]) -> None:
            generic_size_orders = inspection.get("genericSizeOrders") or []
            unresolved_size_orders = inspection.get("unresolvedSizeOrders") or []
            risk_orders: list[dict[str, Any]] = list(generic_size_orders)
            known_ids = {str(item.get("order_id") or "") for item in risk_orders}
            for item in unresolved_size_orders:
                order_id = str(item.get("order_id") or "")
                if order_id and order_id in known_ids:
                    continue
                risk_orders.append({**item, "uncertain": True})
            if unresolved_size_orders:
                self._log("支付前有尺码状态无法确认，已按风险订单处理", unresolved_size_orders)
            payment_action = self._confirm_payment_with_size_warning(counts, selected_order_nos, risk_orders)
            if not payment_action:
                return
            payment_ids = list(selected_ids)
            if payment_action == "safe_only":
                risk_order_ids = {str(item.get("order_id") or "") for item in risk_orders if item.get("order_id")}
                payment_ids = [order_id for order_id in payment_ids if order_id not in risk_order_ids]
                if not payment_ids:
                    messagebox.showinfo("确认支付", "没有可支付的已确认尺码订单。请先处理风险订单。")
                    return
                self._log(f"支付跳过风险订单：{len(risk_order_ids)} 单")

            def task():
                def worker(_session, client: LandwuClient) -> dict[str, Any]:
                    jit_tag = client.get_jit_tag()
                    payment = client.resolve_payment_orders(
                        ids=payment_ids,
                        status=2,
                        limit=100,
                        plat_order_type=jit_tag["id"],
                    )
                    ids = payment["ids"]
                    if not ids:
                        return {
                            "message": "当前勾选的待付款 JIT 已不存在或状态已变化，未支付",
                            "ids": [],
                        }
                    preview = client.get_check_order(ids)
                    validation = require_payment_preview_ok(preview, ids)
                    return {
                        "message": f"待付款 JIT 已提交支付：当前勾选 {len(ids)} 单",
                        "ids": ids,
                        "orderNos": [row.get("order_no") for row in payment["rows"]],
                        "preview": preview,
                        "validation": validation,
                        "commitRequested": True,
                        "forceRequested": False,
                        "result": client.order_pay(ids, commit=True, force=False),
                    }

                return with_landwu_session(self._make_runtime_args(), worker)

            self._run_task("待付款 JIT 预检并支付", task, on_success=lambda _payload: self.refresh_summary())

        self._run_task("读取待付款尺码状态", inspect_task, on_success=continue_payment)

    def process_until_review(self) -> None:
        output_dir = self.output_dir_var.get().strip()
        selected_ids = self.get_checked_edit_order_ids()
        target_rows = self.get_checked_edit_rows()
        target_order_nos = [row.get("order_no") or "" for row in target_rows]
        skipped_vmi_count = sum(1 for row in self.order_rows_by_status_iid.get(1, {}).values() if str(row.get("tag_name") or "").strip().upper() == "VMI")
        if not target_order_nos:
            messagebox.showinfo("一键流程", "当前没有勾选待编辑 JIT。VMI 会跳过。")
            return
        counts = self.get_edit_counts()
        if not messagebox.askyesno(
            "确认一键流程",
            "\n".join(
                [
                    f"待编辑 JIT 共 {counts['jitTotal']} 单。",
                    f"当前勾选 {counts['checked']} 单，取消勾选 {counts['unchecked']} 单。",
                    f"订单：{self._format_order_no_preview(target_order_nos)}",
                    f"待编辑 VMI：{skipped_vmi_count} 单，将跳过。",
                    "",
                    "确定真实改物流，并在订单进入待付款后下载图片吗？",
                ]
            ),
        ):
            return

        def task():
            def worker(_session, client: LandwuClient) -> dict[str, Any]:
                apply_data = client.auto_apply_logistics(commit=True, order_ids=selected_ids)
                if apply_data.get("matchedOrders", 0) > 0 and not apply_data.get("commitExecuted"):
                    return {"apply": apply_data, "message": "改物流预检未通过，已暂停"}
                if apply_data.get("matchedOrders", 0) == 0:
                    return {"apply": apply_data, "message": "待编辑无 JIT，未下载图片"}
                download_target_ids = apply_data.get("orderIds") or selected_ids
                jit_tag_id = (apply_data.get("jitTag") or {}).get("id")
                wait_data = client.wait_for_orders_in_status(
                    order_ids=download_target_ids,
                    status=2,
                    plat_order_type=jit_tag_id,
                    timeout_seconds=90,
                    interval_seconds=3,
                )
                if not wait_data.get("ok"):
                    missing = ", ".join(wait_data.get("missingIds") or [])
                    raise RuntimeError(f"改物流已提交，但未等到订单进入待付款 JIT，已暂停下载。缺失订单ID：{missing}")
                download_data = client.download_order_images(
                    output_dir=output_dir,
                    order_ids=download_target_ids,
                    plat_order_type=jit_tag_id,
                )
                require_download_success(download_data)
                return {"apply": apply_data, "wait": wait_data, "download": download_data, "message": download_data["message"]}

            return with_landwu_session(self._make_runtime_args(), worker)

        self._run_task("一键做到验图前", task, on_success=lambda _payload: self.refresh_summary())


def launch_gui() -> int:
    root = tk.Tk()
    LandwuGuiApp(root)
    root.mainloop()
    return 0


def interactive_menu() -> list[str] | None:
    print(print_help_text())
    print()
    print("选择操作：")
    print("1. 查看登录态")
    print("2. 查看平台标签")
    print("3. 改物流 dry-run")
    print("4. 真实改物流")
    print("5. 一直到验图前")
    print("6. 下载待付款 JIT 图片")
    print("7. 预检并真实支付指定待付款 JIT")
    print("0. 退出")
    choice = input("请输入编号：").strip()

    if choice == "0":
        return None
    if choice == "1":
        return ["auth"]
    if choice == "2":
        return ["tags"]
    if choice == "3":
        return ["apply-logistics"]
    if choice == "4":
        confirm = input("确认真实改物流？输入 YES 继续：").strip().upper()
        if confirm != "YES":
            print("已取消。")
            return None
        return ["apply-logistics", "--commit"]
    if choice == "5":
        confirm = input("这会真实改物流并继续下载图片，输入 YES 继续：").strip().upper()
        if confirm != "YES":
            print("已取消。")
            return None
        return ["process-until-review", "--commit-logistics"]
    if choice == "6":
        custom_dir = input("图片目录可留空，直接回车用桌面默认目录：").strip()
        return ["download-images", "--output-dir", custom_dir] if custom_dir else ["download-images"]
    if choice == "7":
        ids = input("请输入订单ID，多个用逗号分隔；真实支付不允许留空：").strip()
        if not ids:
            print("已取消：真实支付必须指定订单ID。")
            return None
        confirm = input("确认真实支付？输入 PAY 继续：").strip().upper()
        if confirm != "PAY":
            print("已取消。")
            return None
        return ["pay-orders", "--ids", ids, "--commit"]

    print("无效编号。")
    return None


def run_command(args: argparse.Namespace) -> dict[str, Any] | None:
    if args.command == "auth":
        def worker(session, client: LandwuClient) -> dict[str, Any]:
            profile = client.get_user_profile()
            return {
                "authSource": session.auth_source,
                "origin": session.origin,
                "host": session.host,
                "pageTitle": session.page_title,
                "pageUrl": session.page_url,
                "factoryId": session.factory_id,
                "masterFactoryId": session.master_factory_id,
                "username": profile.get("username") or session.user_info.get("username"),
                "nickname": profile.get("nickname") or session.user_info.get("nickname"),
                "balance": profile.get("money"),
                "browser": {
                    "shortcut": session.browser_config.shortcut_path,
                    "executablePath": session.browser_config.target_path,
                    "userDataDir": session.browser_config.user_data_dir,
                    "profileDirectory": session.browser_config.profile_directory,
                    "browserPid": session.browser_pid,
                },
            }

        return with_landwu_session(
            args,
            worker,
        )

    if args.command == "tags":
        return with_landwu_session(args, lambda _session, client: client.get_platform_tags())

    if args.command == "company-list":
        return with_landwu_session(
            args,
            lambda _session, client: [
                {
                    "id": item.get("id"),
                    "logistic_id": item.get("logistic_id"),
                    "name": item.get("name"),
                    "code": item.get("code"),
                }
                for item in client.get_company_list(args.plat_id)
            ],
        )

    if args.command == "list":
        def worker(_session, client: LandwuClient) -> dict[str, Any]:
            data = client.list_orders(
                status=args.status,
                page=args.page,
                limit=args.limit,
                plat_order_type=args.plat_order_type,
            )
            return {
                "current_page": data.get("current_page"),
                "last_page": data.get("last_page"),
                "total": data.get("total"),
                "orders": [
                    {
                        "order_no": row.get("order_no"),
                        "order_id": row.get("order_id"),
                        "status": row.get("status"),
                        "status_text": row.get("status_text"),
                        "plat_id": row.get("plat_id"),
                        "plat_name": row.get("plat_name"),
                        "shop_name": row.get("shop_name"),
                        "plat_order_type": row.get("plat_order_type"),
                        "express_name": row.get("express_name"),
                        "remarks": row.get("remarks"),
                    }
                    for row in data.get("data") or []
                ],
            }

        return with_landwu_session(args, worker)

    if args.command == "apply-logistics":
        return with_landwu_session(
            args,
            lambda _session, client: client.auto_apply_logistics(
                status=args.status,
                limit=args.limit,
                commit=args.commit,
                target_keyword=args.target_keyword,
            ),
        )

    if args.command == "download-images":
        def worker(_session, client: LandwuClient) -> dict[str, Any]:
            download_data = client.download_order_images(
                status=args.status,
                output_dir=args.output_dir or "",
                plat_order_type=client.get_jit_tag()["id"] if int(args.status) == 2 else None,
            )
            require_download_success(download_data)
            return download_data

        return with_landwu_session(
            args,
            worker,
        )

    if args.command == "pay-orders":
        explicit_ids = parse_ids(args.ids)
        if args.commit and not explicit_ids:
            raise RuntimeError("真实支付必须显式传入 --ids，禁止命令行空 ids 支付全部待付款。")

        def worker(_session, client: LandwuClient) -> dict[str, Any]:
            jit_tag = client.get_jit_tag()
            payment = client.resolve_payment_orders(
                ids=explicit_ids,
                status=2,
                limit=args.limit,
                plat_order_type=jit_tag["id"],
            )
            ids = payment["ids"]
            if not ids:
                return {"message": "传入订单里没有待付款 JIT", "ids": []}
            preview = client.get_check_order(ids)
            validation = require_payment_preview_ok(preview, ids) if args.commit else validate_payment_preview(preview, ids)
            return {
                "ids": ids,
                "orderNos": [row.get("order_no") for row in payment["rows"]],
                "preview": preview,
                "validation": validation,
                "commitRequested": args.commit,
                "forceRequested": args.force,
                "result": client.order_pay(ids, commit=args.commit, force=args.force),
            }

        return with_landwu_session(args, worker)

    if args.command == "process-until-review":
        def worker(_session, client: LandwuClient) -> dict[str, Any]:
            apply_data = client.auto_apply_logistics(
                status=1,
                limit=args.limit,
                commit=args.commit_logistics,
                target_keyword=args.target_keyword,
            )
            if apply_data.get("matchedOrders", 0) > 0 and args.commit_logistics and not apply_data.get("commitExecuted"):
                return {"apply": apply_data, "message": "改物流预检未通过，已暂停"}
            if apply_data.get("matchedOrders", 0) > 0 and not args.commit_logistics:
                return {"apply": apply_data, "message": "检测到待编辑 JIT，未真实改物流，未进入抓图"}
            if apply_data.get("matchedOrders", 0) == 0:
                return {"apply": apply_data, "message": "待编辑无 JIT，未下载图片"}
            target_ids = apply_data.get("orderIds") or []
            jit_tag_id = (apply_data.get("jitTag") or {}).get("id")
            wait_data = client.wait_for_orders_in_status(
                order_ids=target_ids,
                status=2,
                plat_order_type=jit_tag_id,
                timeout_seconds=90,
                interval_seconds=3,
            )
            if not wait_data.get("ok"):
                missing = ", ".join(wait_data.get("missingIds") or [])
                raise RuntimeError(f"改物流已提交，但未等到订单进入待付款 JIT，已暂停下载。缺失订单ID：{missing}")
            download_data = client.download_order_images(
                status=2,
                output_dir=args.output_dir or "",
                order_ids=target_ids,
                plat_order_type=jit_tag_id,
            )
            require_download_success(download_data)
            return {"apply": apply_data, "wait": wait_data, "download": download_data, "message": download_data["message"]}

        return with_landwu_session(args, worker)

    return None


def main(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    if not argv:
        return launch_gui()

    parser = build_parser()
    args = parser.parse_args(argv)
    if not args.command:
        print(print_help_text())
        return 0

    try:
        result = run_command(args)
        if result is not None:
            json_out(result)
        return 0
    except Exception as exc:  # noqa: BLE001
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as exc:  # noqa: BLE001
        log_path = Path(__file__).with_suffix(".error.log")
        log_path.write_text(traceback.format_exc(), encoding="utf-8")
        try:
            messagebox.showerror("Landwu 做单助手", f"{exc}\n\n错误日志：{log_path}")
        except Exception:  # noqa: BLE001
            print(str(exc), file=sys.stderr)
        raise
