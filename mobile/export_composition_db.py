from __future__ import annotations

import argparse
import csv
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


DEFAULT_SOURCE_DIR = Path(r"D:\匹配成分数据库")
DEFAULT_OUTPUT_FILE = Path(__file__).resolve().parent / "composition-db.json"
SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


def normalize_db_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "undefined"}:
        return ""
    text = text.replace("\t", "").replace("\u3000", " ").strip()
    if re.fullmatch(r"[+-]?\d+(\.0+)?", text):
        return text.split(".")[0]
    if re.fullmatch(r"[+-]?\d+(\.\d+)?[eE][+-]?\d+", text):
        try:
            dec = Decimal(text)
            if dec == dec.to_integral():
                return str(dec.quantize(Decimal("1")))
            return format(dec.normalize(), "f").rstrip("0").rstrip(".")
        except InvalidOperation:
            return text
    return re.sub(r"\s+", "", text)


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def normalize_composition_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text)


def infer_size_from_composition(value: Any) -> str:
    text = normalize_composition_text(value)
    if not text:
        return ""
    if "人棉" in text:
        return "人棉"
    if any(token in text for token in ("涤纶", "聚酯", "聚脂", "polyester")):
        return "涤纶"
    if "棉" in text or "cotton" in text:
        return "棉"
    return ""


def find_col(headers: list[str], *names: str) -> int:
    normalized = [normalize_header(item) for item in headers]
    targets = {normalize_header(name) for name in names}
    for index, header in enumerate(normalized):
        if header in targets:
            return index
    return -1


def read_csv_table(path: Path) -> tuple[list[str], list[list[Any]]]:
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                rows = list(csv.reader(handle))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"无法识别 CSV 编码：{path}")
    if not rows:
        return [], []
    return [str(item or "").strip() for item in rows[0]], rows[1:]


def read_xlsx_table(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("缺少 openpyxl，无法读取 xlsx/xlsm。") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(item or "").strip() for item in next(iterator, [])]
        rows = [list(row) for row in iterator]
    finally:
        workbook.close()
    return headers, rows


def read_table(path: Path) -> tuple[list[str], list[list[Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_table(path)
    raise RuntimeError(f"不支持的文件类型：{path.name}")


def value_at(row: list[Any], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else ""


def export_records(source_dir: Path) -> tuple[list[dict[str, str]], dict[str, Any]]:
    files = sorted(
        path
        for path in source_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES and not path.name.startswith("~$")
    )
    records: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    skipped: list[str] = []
    scanned_rows = 0
    for path in files:
        try:
            headers, rows = read_table(path)
        except Exception:
            skipped.append(path.name)
            continue
        spu_col = find_col(headers, "SPU_ID", "spu_id", "spuId", "product_id", "productId")
        skc_col = find_col(headers, "SKC_ID", "skc_id", "skcId", "skc")
        sku_col = find_col(headers, "SKU", "sku", "productSku", "product_sku")
        composition_col = find_col(headers, "成分", "成份")
        material_col = find_col(headers, "材质")
        if composition_col < 0 and material_col < 0:
            skipped.append(path.name)
            continue
        for row in rows:
            scanned_rows += 1
            spu_id = normalize_db_key(value_at(row, spu_col))
            skc_id = normalize_db_key(value_at(row, skc_col))
            sku = normalize_db_key(value_at(row, sku_col))
            composition = str(value_at(row, composition_col) or value_at(row, material_col) or "").strip()
            target_size = infer_size_from_composition(composition)
            if not (spu_id or skc_id or sku) or not composition:
                continue
            key = (skc_id, spu_id, sku, composition, target_size)
            if key in seen:
                continue
            seen.add(key)
            records.append(
                {
                    "SKC_ID": skc_id,
                    "SPU_ID": spu_id,
                    "SKU": sku,
                    "composition": composition,
                    "target_size": target_size,
                }
            )
    records.sort(key=lambda item: (item["SKC_ID"], item["SPU_ID"], item["SKU"], item["composition"]))
    return records, {
        "sourceDir": str(source_dir),
        "fileCount": len(files),
        "skippedFiles": skipped,
        "scannedRows": scanned_rows,
        "exportedRows": len(records),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="导出手机端公开成分数据库 JSON。")
    parser.add_argument("source_dir", nargs="?", default=str(DEFAULT_SOURCE_DIR), help="本地成分数据库目录")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_FILE), help="输出 JSON 文件")
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_file = Path(args.output)
    if not source_dir.exists():
        raise SystemExit(f"数据库目录不存在：{source_dir}")
    records, summary = export_records(source_dir)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text(json.dumps(records, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({**summary, "output": str(output_file)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
